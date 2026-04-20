"""
cogs/accounts_cog.py — Comandos de conta corrente e lançamentos.

Comandos tratados:
  - saldo / saldo conta / conta
  - desfazer / undo
  - apagar <id> [id2 id3 ...]   (com confirmação)
  - listar lançamentos
  - exportar excel [data_ini data_fim]
"""
import io
import re
from datetime import date

import discord
from discord.ext import commands
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

from db import (
    get_balance,
    list_launches,
    delete_launch_and_rollback,
    set_pending_action,
    get_launches_by_period,
)
from utils_text import fmt_brl
from utils_date import parse_date_str, month_range_today


class AccountsCog(commands.Cog):

    def __init__(self, bot: commands.Bot):
        self.bot = bot

    async def handle(self, message: discord.Message, t: str, uid: int) -> bool:
        """Retorna True se este Cog tratou a mensagem."""

        # ── Saldo conta ───────────────────────────────────────────────────────
        if t in ("saldo", "saldo conta", "saldo da conta", "conta", "saldo geral"):
            bal = get_balance(uid)
            await message.reply(f"🏦 **Conta Corrente:** {fmt_brl(float(bal))}")
            return True

        # ── Desfazer último lançamento ────────────────────────────────────────
        if t in ("desfazer", "undo", "voltar", "excluir"):
            rows = list_launches(uid, limit=1)
            if not rows:
                await message.reply("Você não tem lançamentos para desfazer.")
                return True

            last_id = int(rows[0]["id"])
            try:
                delete_launch_and_rollback(uid, last_id)
            except LookupError:
                await message.reply("Não achei o último lançamento para desfazer.")
                return True
            except ValueError as e:
                await message.reply(f"Não consegui desfazer o último lançamento: {e}")
                return True
            except Exception:
                await message.reply("Deu erro ao desfazer o último lançamento. Veja os logs.")
                return True

            await message.reply(f"↩️ Desfeito: lançamento **#{last_id}** (saldos ajustados).")
            return True

        # ── Listar lançamentos ────────────────────────────────────────────────
        if t in ("listar lancamentos", "listar lançamentos",
                 "ultimos lancamentos", "últimos lançamentos"):
            rows = list_launches(uid, limit=10)
            if not rows:
                await message.reply("Você ainda não tem lançamentos.")
                return True

            lines = []
            for r in rows:
                tipo = r["tipo"]
                valor = r["valor"]
                alvo = r["alvo"] or "-"
                criado = r["criado_em"]
                nota = r["nota"]

                if tipo == "create_investment" and nota and "taxa=" in nota:
                    try:
                        m_taxa = re.search(r"taxa=([0-9.]+)", nota)
                        m_per = re.search(r"periodo=(\w+)", nota)
                        taxa = float(m_taxa.group(1)) * 100 if m_taxa else None
                        per = m_per.group(1) if m_per else ""
                        per = "ao mês" if per.startswith("month") else "ao dia" if per.startswith("day") else per
                        nota = f"{taxa:.4g}% {per}" if taxa is not None else None
                    except Exception:
                        pass

                valor_str = f"R$ {float(valor):.2f}" if valor is not None else "-"
                nota_part = f" • {nota}" if nota else ""
                lines.append(f"#{r['id']} • {tipo} • {valor_str} • {alvo}{nota_part} • {criado}")

            await message.reply("🧾 **Últimos lançamentos:**\n" + "\n".join(lines))
            return True

        # ── Apagar lançamento(s) ──────────────────────────────────────────────
        if t.startswith(("apagar", "remover")):
            ids_found = [int(x) for x in re.findall(r'\d+', t)]
            if not ids_found:
                await message.reply("Me diga o ID do lançamento. Ex: `apagar 3` ou `apagar 3 5 10`")
                return True

            rows = list_launches(uid, limit=1000)
            rows_by_id = {int(r["id"]): r for r in rows}

            found = [rows_by_id[lid] for lid in ids_found if lid in rows_by_id]
            not_found = [lid for lid in ids_found if lid not in rows_by_id]

            if not found:
                ids_str = ", ".join(str(i) for i in ids_found)
                await message.reply(f"Não achei nenhum lançamento com os IDs: {ids_str}")
                return True

            lines = []
            for row in found:
                lid = int(row["id"])
                tipo = (row.get("tipo") or "").lower()
                tipo_label = "Despesa" if tipo == "despesa" else "Receita" if tipo == "receita" else tipo
                valor = float(row.get("valor") or 0)
                alvo = row.get("alvo") or ""
                nota = row.get("nota") or ""
                criado = row.get("criado_em")
                data = criado.strftime("%d/%m/%Y %H:%M") if hasattr(criado, "strftime") else str(criado)
                desc = f" — {alvo or nota}" if (alvo or nota) else ""
                lines.append(f"• **#{lid}** • {tipo_label} • **{fmt_brl(valor)}**{desc} • {data}")

            aviso = ""
            if not_found:
                aviso = f"\n⚠️ IDs não encontrados: {', '.join(str(i) for i in not_found)}"

            launch_ids = [int(r["id"]) for r in found]
            set_pending_action(uid, "delete_launch_bulk", {"launch_ids": launch_ids}, minutes=10)

            plural = "este lançamento" if len(found) == 1 else f"estes {len(found)} lançamentos"
            await message.reply(
                f"⚠️ Você está prestes a apagar {plural}:\n"
                + "\n".join(lines)
                + aviso
                + "\n\nResponda **sim** para confirmar ou **não** para cancelar. (expira em 10 min)"
            )
            return True

        # ── Exportar Excel ────────────────────────────────────────────────────
        if t.startswith(("exportar excel", "export excel")):
            return await self._exportar_excel(message, uid)

        return False

    # ── Excel ─────────────────────────────────────────────────────────────────

    async def _exportar_excel(self, message: discord.Message, uid: int) -> bool:
        parts = message.content.split()
        try:
            if len(parts) == 2:
                start, end = month_range_today()
            else:
                start = parse_date_str(parts[2])
                end = parse_date_str(parts[3])
                if end < start:
                    await message.reply("A data final não pode ser menor que a inicial.")
                    return True
        except Exception:
            await message.reply("Use: `exportar excel` ou `exportar excel 2026-02-01 2026-02-29`")
            return True

        rows = get_launches_by_period(uid, start, end)
        if not rows:
            await message.reply("📭 Nenhum lançamento no período.")
            return True

        wb = Workbook()
        ws_dash = wb.active
        ws_dash.title = "Dashboard"
        ws_rec = wb.create_sheet("Receitas")
        ws_des = wb.create_sheet("Despesas")
        ws_all = wb.create_sheet("Lançamentos")

        headers = ["Data", "Valor", "Descrição", "Observação", "ID"]
        for ws in (ws_rec, ws_des, ws_all):
            ws.append(headers)

        total_rec = total_des = 0.0
        despesas_por_categoria: dict[str, float] = {}

        for r in rows:
            tipo = r["tipo"]
            valor = float(r["valor"])
            desc = r["alvo"] or r["nota"] or ""
            nota = r["nota"] or ""
            data = r["criado_em"].strftime("%d/%m/%Y")
            row_data = [data, valor, desc, nota, r["id"]]

            ws_all.append(row_data)
            if tipo == "receita":
                ws_rec.append(row_data)
                total_rec += valor
            elif tipo == "despesa":
                ws_des.append(row_data)
                total_des += valor
                cat = (r["alvo"] or "Sem categoria").strip()
                despesas_por_categoria[cat] = despesas_por_categoria.get(cat, 0.0) + valor

        saldo_periodo = total_rec - total_des
        saldo_atual = get_balance(uid)

        # ── Dashboard tab ──────────────────────────────────────────────────
        ws_dash.append(["Período", f"{start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}"])
        ws_dash.append(["Total Receitas", total_rec])
        ws_dash.append(["Total Despesas", total_des])
        ws_dash.append(["Saldo do Período", saldo_periodo])
        ws_dash.append(["Saldo Atual", saldo_atual])

        title_fill = PatternFill("solid", fgColor="1F2937")
        card_fill = PatternFill("solid", fgColor="111827")
        label_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=18, color="FFFFFF")

        ws_dash["A1"].value = "Dashboard Financeiro"
        ws_dash["A1"].font = title_font
        ws_dash["A1"].fill = title_fill
        ws_dash.merge_cells("A1:B1")
        ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dash.row_dimensions[1].height = 28
        ws_dash.column_dimensions["A"].width = 22
        ws_dash.column_dimensions["B"].width = 22

        colors = {"A3": "064E3B", "B3": "064E3B", "A4": "7F1D1D", "B4": "7F1D1D",
                  "A5": "1E3A8A", "B5": "1E3A8A", "A6": "1E3A8A", "B6": "1E3A8A"}
        for row in range(2, 7):
            ws_dash[f"A{row}"].fill = card_fill
            ws_dash[f"B{row}"].fill = card_fill
            ws_dash[f"A{row}"].font = label_font
        for cell_ref, color in colors.items():
            ws_dash[cell_ref].fill = PatternFill("solid", fgColor=color)
        for cell in ("B3", "B4", "B5", "B6"):
            ws_dash[cell].number_format = 'R$ #,##0.00'

        # tabela para gráfico de pizza
        start_row = 2
        ws_dash["D2"] = "Categoria (Despesas)"
        ws_dash["E2"] = "Total"
        ws_dash["D2"].font = Font(bold=True)
        ws_dash["E2"].font = Font(bold=True)

        cats_sorted = sorted(despesas_por_categoria.items(), key=lambda x: x[1], reverse=True)
        r0 = start_row + 1
        for i, (cat, total) in enumerate(cats_sorted):
            rr = r0 + i
            ws_dash[f"D{rr}"] = cat
            ws_dash[f"E{rr}"] = float(total)
            ws_dash[f"E{rr}"].number_format = 'R$ #,##0.00'

        ws_dash.column_dimensions["D"].width = 26
        ws_dash.column_dimensions["E"].width = 14

        last_row = r0 + len(cats_sorted) - 1
        if cats_sorted:
            data_ref = Reference(ws_dash, min_col=5, min_row=start_row, max_row=last_row)
            cats_ref = Reference(ws_dash, min_col=4, min_row=r0, max_row=last_row)
            pie = PieChart()
            pie.title = "Distribuição das despesas"
            pie.add_data(data_ref, titles_from_data=True)
            pie.set_categories(cats_ref)
            pie.height = 15
            pie.width = 20
            pie.style = 10
            ws_dash.add_chart(pie, "G1")

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"dashboard_{start.isoformat()}_{end.isoformat()}.xlsx"
        await message.reply(file=discord.File(fp=bio, filename=filename))
        return True


async def setup(bot: commands.Bot):
    await bot.add_cog(AccountsCog(bot))
