import os
import re
from datetime import date, datetime, timedelta, timezone, time
from dateutil.relativedelta import relativedelta
import calendar
import discord
import io
import csv
import traceback
from discord.ext import commands
from db import init_db
from dotenv import load_dotenv
load_dotenv() #carrega o .env
from db import init_db, ensure_user, add_launch_and_update_balance, get_balance, list_launches, list_pockets, pocket_withdraw_to_account, create_pocket, pocket_deposit_from_account, delete_pocket, investment_withdraw_to_account, accrue_all_investments, create_investment, investment_deposit_from_account, delete_launch_and_rollback
from db import create_investment_db, delete_investment, get_pending_action, clear_pending_action, set_pending_action, list_investments, export_launches, get_launches_by_period, upsert_category_rule, get_memorized_category, get_conn, get_latest_cdi_aa, undo_credit_transaction, undo_installment_group
from ai_router import handle_ai_message, classify_category_with_gpt
import io
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from sheets_export import export_rows_to_dados
import unicodedata
from reports import setup_monthly_export
from utils_date import _tz, now_tz, extract_date_from_text, fmt_br
from commands.resumo import handle_resumo
from utils_date import extract_date_from_text, now_tz, parse_date_str, month_range_today, days_between
from handlers.credit import handle_credit_commands
from utils_text import parse_money, normalize_text
from parsers import parse_receita_despesa_natural
from utils_text import (
    normalize_text,
    contains_word,
    LOCAL_RULES,
    STOPWORDS_PT,
    extract_keyword_for_memory,
    fmt_brl,
    fmt_rate,
    DEPOSIT_VERBS,
    normalize_spaces,
    parse_money,
    should_use_ai,
    parse_pocket_deposit_natural,
)
from ofx_import import import_ofx_bytes
from parsers import parse_receita_despesa_natural
from utils_text import (
    guess_category,
    parse_note_after_amount,
    parse_expense_income_natural,
)
from investment_parse import parse_interest
import time as pytime
import asyncio
from help_ui import help_embed, HelpView

















# Lançamento (padrão único)
# {
#   "id": int,
#   "tipo": str,           # receita | despesa | deposito_caixinha | saque_caixinha | aporte_investimento | resgate_investimento | criar_caixinha | criar_investimento
#   "valor": float,        # sempre número (use 0.0 quando não tiver)
#   "alvo": str,           # categoria/caixinha/investimento
#   "nota": str | None,
#   "criado_em": str,      # ISO
#   "delta_conta": float   # impacto na conta corrente (+/-)
# }


# --------- helpers ---------



# --------- bot setup ---------
intents = discord.Intents.default()
intents.message_content = True  # precisa habilitar no Developer Portal também

bot = commands.Bot(command_prefix="!", intents=intents)

HELP_TEXT_SHORT = (
    "❓ **Não entendi.**\n"
    "Digite `ajuda` para ver os comandos ou `tutorial` para aprender em 2 minutos.\n\n"
    "**Exemplos rápidos:**\n"
    "• `gastei 50 mercado`  *(despesa)*\n"
    "• `recebi 1000 salario` *(receita)*\n"
    "• `saldo`\n"
    "• `importar ofx` + anexo\n"
    "• `exportar sheets`\n"
)

HELP_TEXT_FULL = (
    "💰 **Receitas e Despesas (conta corrente)**\n"
    "• `recebi 1000 salario`\n"
    "• `gastei 120 mercado`\n\n"

    "🏦 **Conta Corrente**\n"
    "• `saldo`\n\n"

    "📦 **Caixinhas**\n"
    "• `criar caixinha viagem`\n"
    "• `coloquei 300 na caixinha viagem`\n"
    "• `retirei 100 da caixinha viagem`\n"
    "• `saldo caixinhas`\n"
    "• `listar caixinhas`\n"
    "• `excluir caixinha viagem`\n\n"

    "📈 **Investimentos**\n"
    "• `criar investimento CDB Nubank 1% ao mês`\n"
    "• `criar investimento Tesouro 0,03% ao dia`\n"
    "• `criar investimento CDB 110% CDI`\n"
    "• `apliquei 200 no investimento CDB Nubank`\n"
    "• `retirei 100 do investimento CDB Nubank`\n"
    "• `saldo investimentos`\n"
    "• `listar investimentos`\n"
    "• `excluir investimento CDB Nubank`\n\n"

    "📊 **CDI**\n"
    "• `ver cdi`\n\n"

    "📤 **Exportar para Google Sheets**\n"
    "• `exportar sheets`\n"
    "• `exportar sheets 2026-02-01 2026-02-28`\n\n"

    "🧾 **Lançamentos**\n"
    "• `listar lançamentos`\n"
    "• `desfazer`\n"
    "• `apagar 3`\n\n"

    "⚠️ **Confirmações**\n"
    "• `sim` → confirma ações (ex: apagar lançamento)\n"
    "• `nao` → cancela a ação pendente\n"
)


@bot.event
async def on_ready():
    print(f"✅ Logado como {bot.user}")
    setup_monthly_export(bot)

@bot.event
async def on_message(message: discord.Message):
    # ignora mensagens do próprio bot
    if message.author.bot:
        return

    text = (message.content or "").strip()
    if not text:
        return
    t = text.casefold()


    # Se existir uma ação pendente, processa "sim" / "não"
    pending = get_pending_action(message.author.id)
    if pending:
        ans = t.strip()

        # confirmar
        if ans in ["sim", "s", "yes", "y"]:
            action = pending["action_type"]
            payload = pending["payload"]

            try:
                if action == "delete_launch":
                    delete_launch_and_rollback(message.author.id, int(payload["launch_id"]))
                    await message.reply(f"🗑️ Apagado e revertido: lançamento **#{payload['launch_id']}**.")
                elif action == "delete_pocket":
                    delete_pocket(message.author.id, payload["pocket_name"])
                    await message.reply(f"🗑️ Caixinha deletada: **{payload['pocket_name']}**.")
                elif action == "delete_investment":
                    delete_investment(message.author.id, payload["investment_name"])
                    await message.reply(f"🗑️ Investimento deletado: **{payload['investment_name']}**.")
                else:
                    await message.reply("Ação pendente desconhecida. Cancelando.")
            except Exception as e:
                traceback.print_exc()
                await message.reply("❌ Deu erro ao executar a ação pendente. Veja os logs.")
            finally:
                # tenta limpar a pending action mesmo se algo falhar
                try:
                    clear_pending_action(message.author.id)
                except Exception as e:
                    print("Erro ao limpar pending_action:", e)
            return

        # cancelar
        if ans in ["nao", "não", "n", "no"]:
            try:
                clear_pending_action(message.author.id)
            except Exception as e:
                print("Erro ao limpar pending_action:", e)
            await message.reply("❌ Ação cancelada.")
            return

       # tem ação pendente, mas o usuário respondeu outra coisa
        preview = pending.get("payload", {}).get("preview_text")
        if preview:
            await message.reply(
                preview
                + "\n\nResponda **sim** para confirmar ou **não** para cancelar."
            )
        else:
            await message.reply(
                "⚠️ Existe uma ação pendente.\n"
                "Responda **sim** para confirmar ou **não** para cancelar."
            )
        return
    
        # ---------------- IMPORT OFX ----------------
    if t.startswith("importar ofx") or t.startswith("import ofx"):
        if not message.attachments:
            await message.reply("📎 Anexe um arquivo `.ofx` junto com o comando `importar ofx`.")
            return

        att = next((a for a in message.attachments if (a.filename or "").lower().endswith(".ofx")), None)
        if not att:
            await message.reply("📎 Não achei nenhum arquivo `.ofx` nos anexos.")
            return

        try:
            ofx_bytes = await att.read()
            report = await asyncio.to_thread(import_ofx_bytes, message.author.id, ofx_bytes, att.filename,)
        except Exception:
            traceback.print_exc()
            await message.reply("❌ Erro ao ler/importar o OFX. (Veja os logs)")
            return

        dt_start = report.get("dt_start")
        dt_end = report.get("dt_end")
        total = report.get("total", 0)
        ins = report.get("inserted", 0)
        dup = report.get("duplicates", 0)
        bal = report.get("new_balance")

        period_txt = ""
        try:
            period_txt = f"{fmt_br(dt_start)} → {fmt_br(dt_end)}"
        except Exception:
            period_txt = "período não identificado"

        skipped = report.get("skipped_same_file", False)
        skipped_txt = " (arquivo já importado antes, pulei)" if skipped else ""

        await message.reply(
            f"✅ OFX importado{skipped_txt}\n"
            f"📅 Período: **{period_txt}**\n"
            f"🧾 Transações no arquivo: **{total}**\n"
            f"➕ Inseridas: **{ins}** | ♻️ Duplicadas: **{dup}**\n"
            f"🏦 Saldo atual: **R$ {bal}**"
        )
        return
    

    # comandos de consulta (não são lançamentos)
    if t.startswith("resumo"):
        await handle_resumo(message, message.author.id, t)
        return


    if t in ["listar caixinhas", "saldo caixinhas", "caixinhas"]:
        rows = list_pockets(message.author.id)

        if not rows:
            await message.reply("Você ainda não tem caixinhas.")
            return

        total = sum(float(r["balance"]) for r in rows)
        linhas = [f"• **{r['name']}**: {fmt_brl(float(r['balance']))}" for r in rows]

        await message.reply(
            "📦 **Caixinhas:**\n"
            + "\n".join(linhas)
            + f"\n\nTotal nas caixinhas: **{fmt_brl(total)}**"
        )
        return
    
    # depositar na caixinha (ex: "transferi 200 para caixinha viagem", "adicionar 200 na caixinha viagem")
    if ("caixinha" in t) and any(w in t for w in ["transferi", "transferir", "adicionar", "colocar", "coloquei", "por", "depositar", "aporte", "aportei"]):
        amount = parse_money(text)
        if amount is None:
            await message.reply("Qual valor? Ex: `transferi 200 para caixinha viagem`")
            return

        # nome depois de "caixinha"
        parts = t.split("caixinha", 1)
        name = parts[1].strip() if len(parts) > 1 else ""
        if not name:
            await message.reply("Pra qual caixinha? Ex: `transferi 200 para caixinha viagem`")
            return

        name = re.sub(r'^(a|para|pra|na|no|da|do)\s+', '', name).strip()

        try:
            launch_id, new_acc, new_pocket, canon_name = pocket_deposit_from_account(
                message.author.id,
                pocket_name=name,
                amount=float(amount),
                nota=text
            )
        except LookupError:
            await message.reply(f"Não achei essa caixinha: **{name}**. Use: `criar caixinha {name}`")
            return
        except ValueError as e:
            if str(e) == "INSUFFICIENT_ACCOUNT":
                # pega saldo atual pra mensagem ficar boa
                bal = get_balance(message.author.id)
                await message.reply(f"Saldo insuficiente na conta. Conta: {fmt_brl(float(bal))}")
            else:
                await message.reply("Valor inválido.")
            return
        except Exception:
            await message.reply("Deu erro ao depositar na caixinha (Postgres). Veja os logs.")
            return

        await message.reply(
            f"✅ Depósito na caixinha **{canon_name}**: +{fmt_brl(float(amount))}\n"
            f"🏦 Conta: {fmt_brl(float(new_acc))} • 📦 Caixinha: {fmt_brl(float(new_pocket))}\n"
            f"ID: **#{launch_id}**"
        )
        return

    
    # sacar/retirar/resgatar X da caixinha Y (CAIXINHA -> CONTA)
    if any(w in t for w in ["retirei", "retirar", "sacar", "saquei", "resgatei", "resgatar"]) and "caixinha" in t:
        amount = parse_money(text)
        if amount is None:
            await message.reply("Qual valor? Ex: `retirei 200 da caixinha viagem`")
            return

        parts = t.split("caixinha", 1)
        name = parts[1].strip() if len(parts) > 1 else ""
        name = re.sub(r'^(da|do|de|na|no|para|pra)\s+', '', name).strip()

        if not name:
            await message.reply("De qual caixinha? Ex: `retirei 200 da caixinha viagem`")
            return

        try:
            launch_id, new_acc, new_pocket, canon_name = pocket_withdraw_to_account(
                message.author.id,
                pocket_name=name,
                amount=float(amount),
                nota=None
            )
        except LookupError:
            await message.reply(f"Não achei essa caixinha: **{name}**. Use: `criar caixinha {name}`")
            return
        except ValueError as e:
            if str(e) == "INSUFFICIENT_POCKET":
                await message.reply(f"Saldo insuficiente na caixinha **{name}**.")
            else:
                await message.reply("Valor inválido.")
            return
        except Exception:
            await message.reply("Deu erro ao sacar da caixinha (Postgres). Veja os logs.")
            return

        await message.reply(
            f"📤 Caixinha **{canon_name}**: -R$ {float(amount):.2f}\n"
            f"🏦 Conta: R$ {float(new_acc):.2f} • 📦 Caixinha: R$ {float(new_pocket):.2f}\n"
            f"ID: #{launch_id}"
        )
        return


    # =========================
    # Listar caixinhas (Postgres)
    # =========================
    if t in ["listar caixinhas", "lista caixinhas", "caixinhas"]:
        rows = list_pockets(message.author.id)

        if not rows:
            await message.reply("Você ainda não tem caixinhas. Use: `criar caixinha <nome>`")
            return

        total = sum(float(r["balance"]) for r in rows)
        lines = [f"📦 **{r['name']}**: {fmt_brl(float(r['balance']))}" for r in rows]

        await message.reply(
            "📦 **Suas caixinhas:**\n"
            + "\n".join(lines)
            + f"\n\nTotal em caixinhas: {fmt_brl(total)}"
        )
        return
    
    # handler de crédito (cartão, fatura, parcelamento, desfazer CT/grupo, etc.)
    if await handle_credit_commands(message):
        return



    # excluir caixinha (com confirmação)
    if t.startswith("excluir caixinha") or t.startswith("apagar caixinha") or t.startswith("remover caixinha"):
        parts = text.split("caixinha", 1)
        name = parts[1].strip() if len(parts) > 1 else ""

        if not name:
            await message.reply("Qual caixinha você quer excluir? Ex: `excluir caixinha viagem`")
            return

        # valida existência + pega nome canônico + saldo
        rows = list_pockets(message.author.id)
        pocket = None
        for r in rows:
            if r["name"].lower() == name.lower():
                pocket = r
                break

        if not pocket:
            await message.reply(f"Não achei essa caixinha: **{name}**")
            return

        canon_name = pocket["name"]
        saldo = float(pocket["balance"])

        if saldo != 0.0:
            await message.reply(
                f"⚠️ Não posso excluir a caixinha **{canon_name}** porque o saldo não é zero ({fmt_brl(saldo)}).\n"
                f"Retire o valor antes e tente novamente."
            )
            return

        # cria a ação pendente (expira em 10 min)
        set_pending_action(message.author.id, "delete_pocket", {"pocket_name": canon_name}, minutes=10)

        await message.reply(
            "⚠️ Você está prestes a excluir esta caixinha:\n"
            f"• **{canon_name}** • saldo: **{fmt_brl(0.0)}**\n\n"
            "Responda **sim** para confirmar ou **não** para cancelar. (expira em 10 min)"
        )
        return



    # excluir investimento (com confirmação)
    if t.startswith("excluir investimento") or t.startswith("apagar investimento") or t.startswith("remover investimento"):
        parts = text.split("investimento", 1)
        name = parts[1].strip() if len(parts) > 1 else ""
        if not name:
            await message.reply("Qual investimento você quer excluir? Ex: `excluir investimento CDB`")
            return

        # valida existência + pega nome canônico + saldo
        rows = list_investments(message.author.id)
        inv = None
        for r in rows:
            if r["name"].lower() == name.lower():
                inv = r
                break

        if not inv:
            await message.reply(f"Não achei esse investimento: **{name}**")
            return

        canon = inv["name"]
        saldo = float(inv["balance"])

        if saldo != 0.0:
            await message.reply(
                f"⚠️ Não posso excluir o investimento **{canon}** porque o saldo não é zero ({fmt_brl(saldo)}).\n"
                f"Retire o valor antes e tente novamente."
            )
            return

        rate = inv.get("rate")
        period = inv.get("period")
        taxa = fmt_rate(rate, period)


        preview_text = (
            "⚠️ Você está prestes a excluir este investimento:\n"
            f"• **{canon}** • saldo: **{fmt_brl(saldo)}**"
            + (f" • taxa: **{taxa}**" if taxa else "")
        )

        # cria a ação pendente (expira em 10 min)
        set_pending_action(
            message.author.id,
            "delete_investment",
            {"investment_name": canon, "preview_text": preview_text},
            minutes=10
        )

        await message.reply(
            preview_text + "\n\nResponda **sim** para confirmar ou **não** para cancelar. (expira em 10 min)"
        )
        return



# Gasto/Receita natural (ex: "gastei 35 no ifood", "recebi 2500 salario")
    user_id = message.author.id
    parsed = parse_receita_despesa_natural(user_id, text)
    if parsed:
        ensure_user(user_id)    

        tipo = parsed["tipo"]                 # "despesa" ou "receita"
        valor = float(parsed["valor"])
        categoria = parsed["categoria"]
        nota = parsed.get("nota")
        criado_em = parsed.get("criado_em")



        launch_id, new_balance = add_launch_and_update_balance(
            user_id=user_id,
            tipo=tipo,
            valor=valor,
            alvo=categoria,
            nota=nota,
            criado_em=criado_em
        )

        emoji = "💸" if tipo == "despesa" else "💰"

        await message.reply(
            f"{emoji} **{tipo.capitalize()} registrada**: {fmt_brl(valor)}\n"
            f"🏷 Categoria: {categoria}\n"
            f"🏦 Conta: {fmt_brl(float(new_balance))}\n"
            f"ID: #{launch_id}"
        )

        return

    # ajuda / comandos
    if t in ["ajuda", "help", "comandos", "listar comandos", "menu"]:
        await message.reply(embed=help_embed("start"), view=HelpView(message.author.id))
        return    

    # (Opcional) se você quiser responder só em DM, descomente:
    # if not isinstance(message.channel, discord.DMChannel):
    #     return

    # criar caixinha
    if t.startswith("criar caixinha"):
        parts = text.split("criar caixinha", 1)
        name = parts[1].strip() if len(parts) > 1 else ""
        if not name:
            await message.reply("Qual o nome da caixinha? Ex: `criar caixinha viagem`")
            return

        try:
            launch_id, pocket_id, pocket_name = create_pocket(
                message.author.id,
                name=name,
                nota=text
            )
        except Exception:
            await message.reply("Deu erro ao criar caixinha (Postgres). Veja os logs.")
            return

        if launch_id is None:
            await message.reply(f"ℹ️ A caixinha **{pocket_name}** já existe.")
            return

        await message.reply(f"✅ Caixinha criada: **{pocket_name}** (ID: **#{launch_id}**)")
        return




  # criar investimento (Postgres) — aceita taxa ao dia / ao mês / ao ano / %CDI
    if t.startswith("criar investimento"):
        rest = text[len("criar investimento"):].strip()
        if not rest:
            await message.reply("Use: `criar investimento <nome> <taxa>% ao dia|ao mês|ao ano` ou `criar investimento <nome> <pct>% cdi`")
            return

        m_cdi = re.search(r'(\d+(?:[.,]\d+)?)\s*%\s*(?:do\s*)?cdi\b', rest, flags=re.I)
        m = re.search(r'(\d+(?:[.,]\d+)?)\s*%\s*(?:ao|a)\s*(dia|m[eê]s|ano)\b', rest, flags=re.I)

        if not m_cdi and not m:
            await message.reply(
                "Não entendi a taxa/período. Exemplos:\n"
                "• `criar investimento CDB 1% ao mês`\n"
                "• `criar investimento Tesouro 0,03% ao dia`\n"
                "• `criar investimento IPCA 12% ao ano`\n"
                "• `criar investimento CDB 100% CDI`"
            )
            return

        # --- CDI ---
        if m_cdi:
            num_str = m_cdi.group(1).replace(",", ".")
            try:
                pct_cdi = float(num_str)  # ex: 110
            except ValueError:
                await message.reply("Percentual do CDI inválido. Ex: `criar investimento CDB 110% cdi`")
                return

            rate = pct_cdi / 100.0       # 110% -> 1.10 (multiplicador)
            period = "cdi"
            periodo_str = f"{pct_cdi:.4g}% do CDI"

            name = (rest[:m_cdi.start()] + rest[m_cdi.end():]).strip(" -–—")
            if not name:
                await message.reply("Me diga o nome do investimento também. Ex: `criar investimento CDB 110% cdi`")
                return

        # --- dia/mês/ano ---
        else:
            num_str = m.group(1).replace(",", ".")
            try:
                rate = float(num_str) / 100.0
            except ValueError:
                await message.reply("Taxa inválida. Ex: **1% ao mês**, **0,03% ao dia**, **12% ao ano**")
                return

            period_raw = m.group(2).lower()
            if "dia" in period_raw:
                period = "daily"
                periodo_str = "ao dia"
            elif "ano" in period_raw:
                period = "yearly"
                periodo_str = "ao ano"
            else:
                period = "monthly"
                periodo_str = "ao mês"

            name = (rest[:m.start()] + rest[m.end():]).strip(" -–—")
            if not name:
                await message.reply("Me diga o nome do investimento também. Ex: `criar investimento CDB 1% ao mês`")
                return

        # --- cria no DB (1 única vez) ---
        try:
            launch_id, inv_id, canon = create_investment_db(
                message.author.id,
                name=name,
                rate=rate,
                period=period,
                nota=text
            )
        except Exception as e:
            print("ERRO criar investimento:", repr(e))
            await message.reply("Deu erro ao criar investimento (Postgres). Veja os logs.")
            return

        if launch_id is None:
            await message.reply(f"ℹ️ O investimento **{canon}** já existe.")
            return

        # resposta
        if period == "cdi":
            await message.reply(f"✅ Investimento criado: **{canon}** ({periodo_str}) (ID: #{launch_id})")
        else:
            await message.reply(f"✅ Investimento criado: **{canon}** ({rate*100:.4g}% {periodo_str}) (ID: #{launch_id})")
        return

    # depósito natural em caixinha (ex: "coloquei 300 na emergencia")
    amount, pocket_name = parse_pocket_deposit_natural(text)
    if amount is not None and pocket_name:
        try:
            launch_id, new_acc, new_pocket, canon_name = pocket_deposit_from_account(
                message.author.id,
                pocket_name=pocket_name,
                amount=float(amount),
                nota=text
            )
        except LookupError:
            await message.reply(f"Não achei essa caixinha: **{pocket_name}**. Use: `criar caixinha {pocket_name}`")
            return
        except ValueError as e:
            if str(e) == "INSUFFICIENT_ACCOUNT":
                bal = get_balance(message.author.id)
                await message.reply(f"Saldo insuficiente na conta. Conta: {fmt_brl(float(bal))}")
            else:
                await message.reply("Valor inválido.")
            return
        except Exception:
            await message.reply("Deu erro ao depositar na caixinha (Postgres). Veja os logs.")
            return

        await message.reply(
            f"✅ Depósito na caixinha **{canon_name}**: +{fmt_brl(float(amount))}\n"
            f"🏦 Conta: {fmt_brl(float(new_acc))} • 📦 Caixinha: {fmt_brl(float(new_pocket))}\n"
            f"ID: **#{launch_id}**"
        )
        return

   # aplicar/aporte no investimento (Postgres) — debita conta corrente
    if any(w in t for w in ["apliquei", "aplicar", "aportei", "aporte"]):
        amount = parse_money(text)
        if amount is None:
            await message.reply("Qual valor? Ex: `apliquei 200 no investimento cdb_nubank`")
            return

        raw = text.lower()

        # tenta extrair nome depois de "no investimento"
        name = None
        if "no investimento" in raw:
            name = text.split("no investimento", 1)[1].strip()

        # tenta extrair nome depois de "investimento"
        if not name and "investimento" in raw:
            parts = re.split(r'\binvestimento\b', text, flags=re.I, maxsplit=1)
            name = parts[1].strip() if len(parts) > 1 else None

        # fallback: "apliquei 500 cdb nubank"
        if not name:
            tmp = re.sub(r'^(apliquei|aplicar|aportei|aporte)\b', '', text, flags=re.I).strip()
            tmp = re.sub(r'\b\d[\d\.\,]*\b', '', tmp, count=1).strip()
            name = tmp.strip(" -–—") or None

        if not name:
            await message.reply("Em qual investimento? Ex: `apliquei 200 no investimento cdb_nubank`")
            return

        try:
            launch_id, new_acc, new_inv, canon_name = investment_deposit_from_account(
                message.author.id,
                investment_name=name,
                amount=float(amount),
                nota=text
            )
        except LookupError:
            await message.reply(f"Não achei esse investimento: **{name}**. Use: `criar investimento {name} 1% ao mês`")
            return
        except ValueError as e:
            if str(e) == "INSUFFICIENT_ACCOUNT":
                bal = get_balance(message.author.id)
                await message.reply(f"Saldo insuficiente na conta. Conta: {fmt_brl(float(bal))}")
            else:
                await message.reply("Valor inválido.")
            return
        except Exception:
            await message.reply("Deu erro ao aplicar/aportar no investimento (Postgres). Veja os logs.")
            return

        await message.reply(
            f"✅ Aporte em **{canon_name}**: +{fmt_brl(float(amount))}. Saldo: **{fmt_brl(float(new_inv))}**\n"
            f"🏦 Conta: {fmt_brl(float(new_acc))}\n"
            f"ID: #{launch_id}"
        )
        return

    
    # resgatar/retirar dinheiro do investimento (Postgres) — credita conta corrente
    if any(w in t for w in ["resgatei", "resgatar", "resgate", "retirei", "retirar", "saquei", "sacar"]):
        amount = parse_money(text)
        if amount is None:
            await message.reply("Qual valor? Ex: `resgatei 200 do investimento cdb_nubank`")
            return

        raw = text.lower()

        # tenta extrair nome depois de "do investimento"
        name = None
        if "do investimento" in raw:
            name = text.split("do investimento", 1)[1].strip()

        # tenta extrair nome depois de "investimento"
        if not name and "investimento" in raw:
            parts = re.split(r'\binvestimento\b', text, flags=re.I, maxsplit=1)
            name = parts[1].strip() if len(parts) > 1 else None

        # fallback: "resgatei 200 cdb nubank"
        if not name:
            tmp = re.sub(r'^(resgatei|resgatar|resgate|retirei|retirar|saquei|sacar)\b', '', text, flags=re.I).strip()
            tmp = re.sub(r'\b\d[\d\.\,]*\b', '', tmp, count=1).strip()
            name = tmp.strip(" -–—") or None

        if not name:
            await message.reply("De qual investimento? Ex: `resgatei 200 do investimento cdb_nubank`")
            return

        try:
            launch_id, new_acc, new_inv, canon_name = investment_withdraw_to_account(
                message.author.id,
                investment_name=name,
                amount=float(amount),
                nota=text
            )
        except LookupError:
            await message.reply(f"Não achei esse investimento: **{name}**. Use: `criar investimento {name} 1% ao mês`")
            return
        except ValueError as e:
            if str(e) == "INSUFFICIENT_INVEST":
                await message.reply(f"Saldo insuficiente no investimento **{name}**.")
            else:
                await message.reply("Valor inválido.")
            return
        except Exception:
            await message.reply("Deu erro ao resgatar investimento (Postgres). Veja os logs.")
            return

        await message.reply(
            f"💸 Resgate de **{canon_name}**: -{fmt_brl(float(amount))}. Saldo: **{fmt_brl(float(new_inv))}**\n"
            f"🏦 Conta: {fmt_brl(float(new_acc))}\n"
            f"ID: #{launch_id}"
        )
        return


   # saldo caixinhas (Postgres)
    if t == "saldo caixinhas":
        rows = list_pockets(message.author.id)
        if not rows:
            await message.reply("Você não tem caixinhas ainda. Use: `criar caixinha viagem`")
            return

        lines = "\n".join([f"- **{r['name']}**: {fmt_brl(float(r['balance']))}" for r in rows])
        await message.reply("💰 **Caixinhas:**\n" + lines)
        return


   # saldo investimentos (Postgres + aplica juros antes)
    if t == "saldo investimentos":
        rows = accrue_all_investments(message.author.id)
        if not rows:
            await message.reply("Você não tem investimentos ainda. Use: `criar investimento CDB 1,1% ao mês`")
            return

        lines = "\n".join([f"- **{r['name']}**: {fmt_brl(float(r['balance']))}" for r in rows])
        await message.reply("📈 **Investimentos:**\n" + lines)
        return


   # listar investimentos (Postgres + aplica juros antes)
    if t in ["listar investimentos", "lista investimentos", "investimentos", "meus investimentos"]:
        rows = accrue_all_investments(message.author.id)
        if not rows:
            await message.reply("Você ainda não tem investimentos.")
            return

        lines = ["📈 **Seus investimentos:**"]
        for r in rows:
            rate_pct = float(r["rate"]) * 100
            period = (r["period"] or "monthly").lower()
            period_str = "ao dia" if period == "daily" else ("ao mês" if period == "monthly" else "ao ano")
            bal = float(r["balance"])
            lines.append(f"• **{r['name']}** — {rate_pct:.4g}% {period_str} — saldo: {fmt_brl(bal)}")

        await message.reply("\n".join(lines))
        return

# listar lancamentos (Postgres)
    if t in ["listar lancamentos", "listar lançamentos", "ultimos lancamentos", "últimos lançamentos"]:
        rows = list_launches(message.author.id, limit=10)

        if not rows:
            await message.reply("Você ainda não tem lançamentos.")
            return

        lines = []
        for r in rows:
            tipo = r["tipo"]
            valor = r["valor"]
            alvo = r["alvo"] or "-"
            criado = r["criado_em"]
            nota = r["nota"]

            # mesma limpeza que você já tinha (mantive)
            if tipo == "create_investment" and nota and "taxa=" in nota:
                try:
                    m_taxa = re.search(r"taxa=([0-9.]+)", nota)
                    m_per = re.search(r"periodo=(\w+)", nota)
                    taxa = float(m_taxa.group(1)) * 100 if m_taxa else None
                    per = m_per.group(1) if m_per else ""
                    per = "ao mês" if per.startswith("month") else "ao dia" if per.startswith("day") else per
                    nota = f"{taxa:.4g}% {per}" if taxa is not None else None
                except:
                    pass

            valor_str = f"R$ {float(valor):.2f}" if valor is not None else "-"
            nota_part = f" • {nota}" if nota else ""
            lines.append(f"#{r['id']} • {tipo} • {valor_str} • {alvo}{nota_part} • {criado}")

        await message.reply("🧾 **Últimos lançamentos:**\n" + "\n".join(lines))
        return

    # =========================
    # Apagar lançamento pelo ID (Postgres) - com confirmação
    # =========================
    if t.startswith("apagar") or t.startswith("remover"):
        m = re.search(r'(\d+)', t)
        if not m:
            await message.reply("Me diga o ID do lançamento. Ex: `apagar 3`")
            return

        launch_id = int(m.group(1))

        # (opcional) valida se existe antes de pedir confirmação
        rows = list_launches(message.author.id, limit=1000)
        row = next((r for r in rows if int(r["id"]) == launch_id), None)
        if not row:
            await message.reply(f"Não achei lançamento com ID {launch_id}.")
            return

        tipo = (row.get("tipo") or "").lower()
        tipo_label = "Despesa" if tipo == "despesa" else "Receita" if tipo == "receita" else tipo
        valor = float(row.get("valor") or 0)
        alvo = row.get("alvo") or ""
        nota = row.get("nota") or ""
        criado = row.get("criado_em")
        data = criado.strftime("%d/%m/%Y %H:%M") if hasattr(criado, "strftime") else str(criado)

        desc = alvo if alvo else nota
        if desc:
            desc = f" — {desc}"

        set_pending_action(message.author.id, "delete_launch", {"launch_id": launch_id}, minutes=10)

        await message.reply(
            "⚠️ Você está prestes a apagar este lançamento:\n"
            f"• **#{launch_id}** • **{tipo_label}** • **{fmt_brl(valor)}**{desc} • {data}\n\n"
            "Responda **sim** para confirmar ou **não** para cancelar. (expira em 10 min)"
        )
        return


    # comando para desfazer a última ação (100% Postgres)
    if t in ["desfazer", "undo", "voltar", "excluir"]:
        user_id = message.author.id

        rows = list_launches(user_id, limit=1)
        if not rows:
            await message.reply("Você não tem lançamentos para desfazer.")
            return

        last_id = int(rows[0]["id"])

        try:
            delete_launch_and_rollback(user_id, last_id)
        except LookupError:
            await message.reply("Não achei o último lançamento para desfazer (isso não deveria acontecer).")
            return
        except ValueError as e:
            await message.reply(f"Não consegui desfazer o último lançamento: {e}")
            return
        except Exception:
            await message.reply("Deu erro ao desfazer o último lançamento (Postgres). Veja os logs.")
            return

        await message.reply(f"↩️ Desfeito: lançamento **#{last_id}** (saldos ajustados no banco).")
        return
        
    # comando para ver saldo da conta
    if t in ["saldo", "saldo conta", "saldo da conta", "conta", "saldo geral"]:
        user_id = message.author.id
        bal = get_balance(user_id)
        
        await message.reply(f"🏦 **Conta Corrente:** {fmt_brl(float(bal))}")
        return
    
    # comando para ver CDI
    if t in ["ver cdi", "cdi"]:
        try:
            # abre conexão/cur do jeito que você já usa no bot
         with get_conn() as conn:
            with conn.cursor() as cur:
                res = get_latest_cdi_aa(cur)


            if not res:
                await message.reply("⚠️ Não consegui obter a CDI agora. Tente novamente mais tarde.")
                return

            ref_date, cdi_aa = res
            await message.reply(
                f"📊 **CDI (a.a.)**\n"
                f"Data: **{ref_date.strftime('%d/%m/%Y')}**\n"
                f"Valor: **{cdi_aa:.2f}% ao ano**"
            )
            return

        except Exception as e:
            print("Erro ao buscar CDI:", e)
            await message.reply("❌ Erro ao buscar a CDI. Veja os logs.")
            return

    
   # Exporta para Google Sheets (dashboard)
    if t.startswith("exportar sheets"):
        parts = text.split()

        try:
            # Opção B: por padrão, exporta TUDO
            if len(parts) == 2:
                from datetime import date
                start = date(1970, 1, 1)
                end = month_range_today()[1]  # hoje (fim do mês atual no seu helper)
            # ainda permite exportar um recorte, se você quiser usar
            elif len(parts) == 4:
                start = parse_date_str(parts[2])
                end = parse_date_str(parts[3])
                if end < start:
                    await message.reply("A data final não pode ser menor que a inicial.")
                    return
            else:
                raise ValueError("args")

        except Exception:
            await message.reply("Use: `exportar sheets` ou `exportar sheets 2026-02-01 2026-02-28`")
            return

        rows = get_launches_by_period(message.author.id, start, end)
        if not rows:
            await message.reply("📭 Nenhum lançamento no período.")
            return

        try:
            sheet_link = export_rows_to_dados(message.author.id, rows)
        except Exception as e:
            await message.reply(f"❌ Erro ao exportar para o Sheets: {e}")
            return

        await message.reply(f"✅ Exportado para o dashboard (aba **DADOS**).\n🔗 {sheet_link}")
        return
    
    # Exporta dashboard financeiro em Excel
    if t.startswith("exportar excel") or t.startswith("export excel"):
        parts = text.split()

        try:
            if len(parts) == 2:
                start, end = month_range_today()
            else:
                start = parse_date_str(parts[2])
                end = parse_date_str(parts[3])
                if end < start:
                    await message.reply("A data final não pode ser menor que a inicial.")
                    return
        except Exception:
            await message.reply("Use: `exportar excel` ou `exportar excel 2026-02-01 2026-02-29`")
            return

        rows = get_launches_by_period(message.author.id, start, end)
        if not rows:
            await message.reply("📭 Nenhum lançamento no período.")
            return

        wb = Workbook()
        ws_dash = wb.active
        ws_dash.title = "Dashboard"

        ws_rec = wb.create_sheet("Receitas")
        ws_des = wb.create_sheet("Despesas")
        ws_all = wb.create_sheet("Lançamentos")

        # headers
        headers = ["Data", "Valor", "Descrição", "Observação", "ID"]
        for ws in (ws_rec, ws_des, ws_all):
            ws.append(headers)

        total_rec = 0
        total_des = 0

        for r in rows:
            tipo = r["tipo"]
            valor = float(r["valor"])
            desc = r["alvo"] or r["nota"] or ""
            nota = r["nota"] or ""
            data = r["criado_em"].strftime("%d/%m/%Y")

            row = [data, valor, desc, nota, r["id"]]
            ws_all.append(row)

            if tipo == "receita":
                ws_rec.append(row)
                total_rec += valor
            elif tipo == "despesa":
                ws_des.append(row)
                total_des += valor

        # agrega despesas por categoria (usa alvo como categoria)
        despesas_por_categoria = {}
        for r in rows:
            if r["tipo"] != "despesa":
                continue
            cat = (r["alvo"] or "Sem categoria").strip()
            despesas_por_categoria[cat] = despesas_por_categoria.get(cat, 0.0) + float(r["valor"])

        saldo_periodo = total_rec - total_des
        saldo_atual = get_balance(message.author.id)

        ws_dash.append(["Período", f"{start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}"])
        ws_dash.append(["Total Receitas", total_rec])
        ws_dash.append(["Total Despesas", total_des])
        ws_dash.append(["Saldo do Período", saldo_periodo])
        ws_dash.append(["Saldo Atual", saldo_atual])

        # estilo simples tipo "card"
        title_fill = PatternFill("solid", fgColor="1F2937")  # cinza escuro
        card_fill  = PatternFill("solid", fgColor="111827")  # mais escuro
        label_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=18, color="FFFFFF")

        ws_dash["A1"].value = "Dashboard Financeiro"
        ws_dash["A1"].font = title_font
        ws_dash["A1"].fill = title_fill
        ws_dash.merge_cells("A1:B1")
        ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dash.row_dimensions[1].height = 28

        ws_dash.column_dimensions["A"].width = 22
        ws_dash.column_dimensions["B"].width = 22

        for row in range(2, 7):
            ws_dash[f"A{row}"].fill = card_fill
            ws_dash[f"B{row}"].fill = card_fill
            ws_dash[f"A{row}"].font = label_font
            ws_dash[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            ws_dash[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center")

        rec_fill   = PatternFill("solid", fgColor="064E3B")  # verde escuro
        des_fill   = PatternFill("solid", fgColor="7F1D1D")  # vermelho escuro
        saldo_fill = PatternFill("solid", fgColor="1E3A8A")  # azul escuro

        # A3 = Total Receitas
        ws_dash["A3"].fill = rec_fill
        ws_dash["B3"].fill = rec_fill

        # A4 = Total Despesas
        ws_dash["A4"].fill = des_fill
        ws_dash["B4"].fill = des_fill

        # A5/A6 = Saldos
        ws_dash["A5"].fill = saldo_fill
        ws_dash["B5"].fill = saldo_fill
        ws_dash["A6"].fill = saldo_fill
        ws_dash["B6"].fill = saldo_fill


        for cell in ["B3", "B4", "B5", "B6"]:
            ws_dash[cell].number_format = 'R$ #,##0.00'

        # tabela auxiliar para gráfico (Categoria x Total)
        start_row = 2
        cat_col = "D"
        val_col = "E"

        ws_dash[f"{cat_col}{start_row}"] = "Categoria (Despesas)"
        ws_dash[f"{val_col}{start_row}"] = "Total"
        ws_dash[f"{cat_col}{start_row}"].font = Font(bold=True)
        ws_dash[f"{val_col}{start_row}"].font = Font(bold=True)

        cats_sorted = sorted(despesas_por_categoria.items(), key=lambda x: x[1], reverse=True)

        r0 = start_row + 1
        for i, (cat, total) in enumerate(cats_sorted):
            rr = r0 + i
            ws_dash[f"{cat_col}{rr}"] = cat
            ws_dash[f"{val_col}{rr}"] = float(total)
            ws_dash[f"{val_col}{rr}"].number_format = 'R$ #,##0.00'

        ws_dash.column_dimensions[cat_col].width = 26
        ws_dash.column_dimensions[val_col].width = 14

        last_row = r0 + len(cats_sorted) - 1
        if len(cats_sorted) > 0:

            data = Reference(ws_dash, min_col=5, min_row=start_row, max_row=last_row)
            cats = Reference(ws_dash, min_col=4, min_row=r0, max_row=last_row)
            
            # grafico de barras
            # bar = BarChart()
            # bar.type = "col"
            # bar.title = "Despesas por categoria"
            # bar.y_axis.title = "R$"
            # bar.x_axis.title = "Categoria"
            # bar.style = 10
            # bar.y_axis.majorGridlines = None
            # bar.x_axis.majorGridlines = None

            # bar.add_data(data, titles_from_data=True)
            # bar.set_categories(cats)
            # bar.height = 10
            # bar.width = 22s

            # ws_dash.add_chart(bar, "D8")

            # grafico de pizza
            pie = PieChart()
            pie.title = "Distribuição das despesas"
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(cats)
            pie.height = 15
            pie.width = 20
            pie.style = 10

            ws_dash.add_chart(pie, "G1")

        # salva em memória
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"dashboard_{start.isoformat()}_{end.isoformat()}.xlsx"
        await message.reply(file=discord.File(fp=bio, filename=filename))
        return

    # fallback com IA (apenas se fizer sentido financeiro)
    if should_use_ai(message.content):
        ai_reply = await handle_ai_message(message.author.id, message.content)
        if ai_reply:
            await message.reply(ai_reply)
            return
        
    # fallback
    await message.reply("❓ **Não entendi seu comando. Tente um destes exemplos:**\n\n" + HELP_TEXT_SHORT)



# --------- run ---------
if __name__ == "__main__":
    token = os.getenv("DISCORD_BOT_TOKEN")
    if not token:
        raise RuntimeError("DISCORD_BOT_TOKEN não definido.")

    # 1) init_db com log + falha explícita
    try:
        print("🗄️ Inicializando banco de dados (init_db)...")
        init_db()
        print("✅ Banco inicializado com sucesso!")
    except Exception as e:
        print("❌ Falha no init_db:", e)
        traceback.print_exc()
        # não adianta ficar tentando logar no Discord se o DB tá quebrado
        raise

    # 2) retry com backoff para evitar 429/crash loop
    wait = 15  # começa leve
    while True:
        try:
            print("🤖 Conectando no Discord...")
            bot.run(token)
            # se bot.run retornar (normalmente só retorna ao encerrar), zera backoff
            wait = 15
        except Exception as e:
            msg = str(e)
            print("❌ Bot caiu:", msg)
            traceback.print_exc()

            # se for rate limit / 429, espera mais
            if "429" in msg or "Too Many Requests" in msg:
                wait = max(wait, 60)

            print(f"⏳ Aguardando {wait}s para tentar de novo...")
            pytime.sleep(wait)
            wait = min(wait * 2, 600)  # dobra até 10 min



