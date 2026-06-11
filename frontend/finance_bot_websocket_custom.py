"""
Finance Bot WebSocket Dashboard Server
Real-time financial data via WebSocket + FastAPI

Endpoints:
  GET  /                        → serves dashboard.html
  GET  /manifest.json           → PWA manifest
  GET  /service-worker.js       → PWA service worker
  GET  /health
  GET  /data/{user_id}          → snapshot (query: year, month)
  GET  /history/{user_id}       → last N months summary (query: months)
  GET  /budgets/{user_id}       → list budgets
  POST /budgets/{user_id}       → set budget {categoria, budget}
  DEL  /budgets/{user_id}/{cat} → delete budget
  POST /export/{user_id}        → envia extrato (PDF+XLSX+CSV) p/ email (query: year, month)
  WS   /ws/{user_id}            → real-time updates
"""

import asyncio
import base64
import csv
import io
import json
import logging
import os
import pathlib
import secrets
import sys
import time as _startup_time
import urllib.parse
from decimal import Decimal
from datetime import datetime, date, timedelta, timezone
from typing import Any, Dict

import psycopg
from psycopg.rows import dict_row
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, HTTPException, Depends, Query, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse, RedirectResponse, HTMLResponse, JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import uvicorn
from pydantic import BaseModel, EmailStr
import jwt as pyjwt
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from config.env import load_app_env
from token_utils import decode_dashboard_token, decode_dashboard_token_full, make_dashboard_token
from utils_phone import normalize_phone_e164
from core.admin_dashboard import (
    ensure_admin_tables,
    log_auth_login_event,
    log_system_event,
    log_admin_startup_warnings,
    admin_error_logging_middleware,
    register_admin_routes,
)
from core.audit import (
    AuditEvent,
    list_audit_events,
    maybe_record_login_from_new_ip,
    record_audit_event,
)
from core.sessions import (
    create_session,
    device_label,
    get_active_session,
    list_user_sessions,
    revoke_other_sessions,
    revoke_session,
    touch_session,
)
from db import (
    accrue_all_pockets,
    accrue_all_investments,
    create_card,
    create_investment_db,
    create_pocket,
    delete_pocket,
    delete_investment,
    create_mock_open_finance_connection,
    disconnect_open_finance_connection,
    get_dashboard_market_rates,
    get_open_finance_snapshot,
    get_auth_user,
    build_user_export_zip,
    verify_user_password,
    get_user_email,
    create_data_export_token,
    consume_data_export_token,
    has_recent_export_request,
    ensure_account_deletion_columns,
    get_daily_report_prefs,
    is_account_scheduled_for_deletion,
    list_identities_by_user,
    process_due_account_deletions,
    save_pluggy_open_finance_item,
    schedule_account_deletion,
    set_daily_report_enabled,
    set_daily_report_hour,
    set_engagement_opt_out,
    set_tip_email_opt_out,
    set_insight_email_opt_out,
    set_whatsapp_updates_opt_out,
    sync_engagement_opt_out,
    update_pluggy_open_finance_item_status,
    investment_deposit_from_account,
    investment_withdraw_to_account,
    pocket_deposit_from_account,
    pocket_withdraw_to_account,
    update_launch_category,
    update_launch_fields,
    update_credit_transaction_fields,
    undo_credit_transaction,
    delete_launch_and_rollback,
)
from core.services.pluggy import (
    PluggyApiError,
    PluggyConfigError,
    create_pluggy_connect_token,
)
from frontend.routes.shared import DASHBOARD_URL
from frontend.routes.static_pages import router as static_pages_router

load_app_env()

DATABASE_URL      = os.getenv("DATABASE_URL")
DASHBOARD_USER_ID = os.getenv("DASHBOARD_USER_ID")
TZ                = os.getenv("TZ", "America/Sao_Paulo")
JWT_SECRET              = (os.getenv("JWT_SECRET") or "").strip()
# DASHBOARD_URL (leitura + sanitização do env) vem de frontend/routes/shared.py
# Em dev local (http://localhost) o navegador rejeita cookies Secure. Em prod
# DASHBOARD_URL é https → Secure=True como sempre.
COOKIE_SECURE = DASHBOARD_URL.startswith("https://")
WHATSAPP_NUMBER         = os.getenv("WHATSAPP_NUMBER", "")
STRIPE_SECRET_KEY       = os.getenv("STRIPE_SECRET_KEY", "")
STRIPE_WEBHOOK_SECRET   = os.getenv("STRIPE_WEBHOOK_SECRET", "")
STRIPE_PRICE_ID_PRO     = os.getenv("STRIPE_PRICE_ID_PRO", "")            # legacy: usado como fallback do mensal
STRIPE_PRICE_ID_PRO_MENSAL = os.getenv("STRIPE_PRICE_ID_PRO_MENSAL", "")  # price_xxx Pro mensal (R$ 19,90)
STRIPE_PRICE_ID_PRO_ANUAL  = os.getenv("STRIPE_PRICE_ID_PRO_ANUAL", "")   # price_xxx Pro anual  (R$ 199,00)
PLUGGY_INCLUDE_SANDBOX  = os.getenv("PLUGGY_INCLUDE_SANDBOX", "1") != "0"
DASHBOARD_MAGIC_LINK_MINUTES = int(os.getenv("DASHBOARD_MAGIC_LINK_MINUTES", "5"))
DASHBOARD_SESSION_HOURS = float(os.getenv("DASHBOARD_SESSION_HOURS", "12"))
DB_CONNECT_TIMEOUT = int(os.getenv("DB_CONNECT_TIMEOUT", "5"))
STARTUP_STEP_TIMEOUT = int(os.getenv("STARTUP_STEP_TIMEOUT", "12"))
RUN_BACKGROUND_TASKS = os.getenv("RUN_BACKGROUND_TASKS", "1") != "0"
ENABLE_DEV_ENDPOINTS = (os.getenv("ENABLE_DEV_ENDPOINTS") or "").strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}

HERE = pathlib.Path(__file__).parent  # directory of this file

if not DATABASE_URL:
    print("ERROR: DATABASE_URL not set. Check your .env file.", file=sys.stderr)
    sys.exit(1)

if not JWT_SECRET:
    print("ERROR: JWT_SECRET not set. Refusing to start with insecure default.", file=sys.stderr)
    sys.exit(1)

# ─── JSON serializer ─────────────────────────────────────────────────────────

class FinanceEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):  return float(obj)
        if isinstance(obj, datetime): return obj.isoformat()
        if isinstance(obj, date):     return obj.isoformat()
        return super().default(obj)

def jdump(data: dict) -> str:
    return json.dumps(data, cls=FinanceEncoder, ensure_ascii=False)

# ─── DB helpers (com connection pool) ────────────────────────────────────────
# Pool global de conexões assíncronas. Em vez de abrir nova conn a cada query
# (custa 1-2s no Railway), reusa de um pool. O `_PooledConn` mantém a interface
# antiga (`async with await db_connect() as conn:`) intacta — todos os callers
# antigos continuam funcionando sem mudança.
from psycopg_pool import AsyncConnectionPool

_db_pool: AsyncConnectionPool | None = None
_db_pool_lock = asyncio.Lock()


async def _get_db_pool() -> AsyncConnectionPool:
    global _db_pool
    if _db_pool is not None:
        return _db_pool
    async with _db_pool_lock:
        if _db_pool is not None:  # double-check após pegar lock
            return _db_pool
        pool = AsyncConnectionPool(
            DATABASE_URL,
            min_size=1,
            max_size=int(os.getenv("DB_POOL_MAX", "8")),
            timeout=DB_CONNECT_TIMEOUT,
            kwargs={"row_factory": dict_row},
            open=False,
        )
        await pool.open(wait=True, timeout=DB_CONNECT_TIMEOUT)
        _db_pool = pool
        return _db_pool


class _PooledConn:
    """Adapter pra preservar a interface `async with await db_connect() as conn`.
    `pool.connection()` retorna um async-context-manager direto, mas o caller
    legado faz `await db_connect()` antes de entrar no async-with — esse wrapper
    casa os dois protocolos."""
    def __init__(self, pool: AsyncConnectionPool):
        self._pool = pool
        self._cm = None

    async def __aenter__(self):
        self._cm = self._pool.connection()
        return await self._cm.__aenter__()

    async def __aexit__(self, exc_type, exc, tb):
        if self._cm is None:
            return False
        return await self._cm.__aexit__(exc_type, exc, tb)


async def db_connect():
    pool = await _get_db_pool()
    return _PooledConn(pool)

async def list_users() -> list:
    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute("SELECT id FROM users ORDER BY created_at")
            return await cur.fetchall()

def _month_range(year: int, month: int):
    """Returns (start_date, exclusive_end_date) for the given month."""
    start = date(year, month, 1)
    end   = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    return start, end

# ─── Core data fetcher ───────────────────────────────────────────────────────

_DASHBOARD_CURRENT_CACHE_TTL_SECONDS = 45
_dashboard_current_cache: Dict[int, tuple[float, Any, Any, Any]] = {}


def _invalidate_dashboard_current_cache(user_id: int) -> None:
    _dashboard_current_cache.pop(int(user_id), None)


async def _get_dashboard_current_state(user_id: int):
    now_mono = _startup_time.monotonic()
    cached = _dashboard_current_cache.get(int(user_id))
    if cached and now_mono - cached[0] < _DASHBOARD_CURRENT_CACHE_TTL_SECONDS:
        return cached[1], cached[2], cached[3]

    current_pockets, current_investments, market_rates = await asyncio.gather(
        asyncio.to_thread(accrue_all_pockets, user_id),
        asyncio.to_thread(accrue_all_investments, user_id),
        asyncio.to_thread(get_dashboard_market_rates),
    )
    _dashboard_current_cache[int(user_id)] = (
        _startup_time.monotonic(),
        current_pockets,
        current_investments,
        market_rates,
    )
    return current_pockets, current_investments, market_rates


def _dashboard_launch_filter_sql(filter_type: str | None, query: str | None) -> tuple[list[str], list]:
    clauses: list[str] = []
    params: list = []

    filter_type = (filter_type or "all").strip().lower()
    query = (query or "").strip()

    if filter_type == "receita":
        clauses.append("tipo IN ('receita', 'entrada') AND is_internal_movement = false")
    elif filter_type == "despesa":
        clauses.append("tipo IN ('despesa', 'saida') AND is_internal_movement = false")
    elif filter_type == "investimento":
        clauses.append("tipo IN ('aporte_investimento', 'resgate_investimento')")
    elif filter_type == "interno":
        clauses.append("is_internal_movement = true")

    if query:
        clauses.append(
            """
            (
              lower(coalesce(nota, '')) LIKE %s
              OR lower(coalesce(alvo, '')) LIKE %s
              OR lower(coalesce(categoria, '')) LIKE %s
              OR lower(coalesce(tipo, '')) LIKE %s
            )
            """
        )
        like = f"%{query.lower()}%"
        params.extend([like, like, like, like])

    return clauses, params


async def get_financial_data(
    user_id: int,
    year: int = None,
    month: int = None,
    page: int = 1,
    limit: int = 25,
    filter_type: str | None = None,
    query: str | None = None,
) -> dict:
    """
    Fetch full financial snapshot for user_id.
    If year/month are given, income/expenses/categories/launches
    are scoped to that month. Balance, pockets and investments
    always reflect the current state.
    """
    now = datetime.now(timezone.utc)
    y   = year  or now.year
    m   = month or now.month
    month_start, month_end = _month_range(y, m)
    is_current = (y == now.year and m == now.month)
    page = max(int(page or 1), 1)
    limit = max(min(int(limit or 25), 100), 1)
    offset = (page - 1) * limit
    current_pockets, current_investments, market_rates = await _get_dashboard_current_state(user_id)
    launch_filter_clauses, launch_filter_params = _dashboard_launch_filter_sql(filter_type, query)
    launch_filter_sql = "".join(f"\n                  AND ({clause})" for clause in launch_filter_clauses)

    # Helper que roda 1 query num conn próprio do pool. Permite paralelizar
    # via asyncio.gather (cada gather pega uma conn diferente do pool).
    async def _q(sql: str, params: tuple = ()):
        async with await db_connect() as conn:
            async with conn.cursor() as cur:
                await cur.execute(sql, params)
                return await cur.fetchall()

    pockets = current_pockets
    investments = current_investments  # do cache

    # Compras no crédito viram linhas virtuais com tipo='credito' no
    # histórico — só quando o filtro permitir (no filtro "all" ou sem filtro).
    include_credit = (filter_type or "all").strip().lower() in ("", "all")

    credit_union_sql = ""
    credit_union_params: list = []
    if include_credit:
        credit_union_sql = """
            UNION ALL
            SELECT t.id AS id,
                   'credito' AS tipo,
                   t.valor AS valor,
                   c.name AS alvo,
                   t.nota AS nota,
                   t.categoria AS categoria,
                   t.created_at AS criado_em,
                   false AS is_internal_movement,
                   t.installments_total AS installments_total,
                   t.installment_no AS installment_no
            FROM credit_transactions t
            JOIN credit_cards c ON c.id = t.card_id
            WHERE t.user_id = %s
              AND t.purchased_at >= %s::date
              AND t.purchased_at < %s::date
              AND t.is_refund = false
        """
        credit_union_params = [user_id, month_start, month_end]

    # ───── Paraleliza queries independentes via asyncio.gather ─────
    # Cada _q() pega uma conn do pool. Antes era sequencial dentro de UMA
    # conn → 10 round-trips somados. Agora roda simultâneo → tempo total
    # ≈ tempo da query mais lenta. Ganho enorme em DB com latência alta.
    (
        account_rows,
        launches_total_rows,
        launches,
        monthly,
        categories,
        allocations_rows,
        card_rows,
        daily_rows,
        budget_rows,
    ) = await asyncio.gather(
        # 1) Account balance
        _q("SELECT balance FROM accounts WHERE user_id = %s", (user_id,)),
        # 3) Total launches (com filtros + credit union)
        _q(
            f"""
            SELECT COUNT(*) AS total FROM (
                SELECT id, tipo, valor, alvo, nota, categoria, criado_em, is_internal_movement,
                       NULL::int AS installments_total,
                       NULL::int AS installment_no
                FROM launches
                WHERE user_id = %s
                  AND criado_em >= %s AND criado_em < %s
                  AND tipo NOT IN ('criar_caixinha', 'delete_pocket', 'create_investment', 'delete_investment')
                  {launch_filter_sql}
                {credit_union_sql}
            ) merged
            """,
            (user_id, month_start, month_end, *launch_filter_params, *credit_union_params),
        ),
        # 4) Launches paginado
        _q(
            f"""
            SELECT id, tipo, valor, alvo, nota, categoria, criado_em, is_internal_movement,
                   installments_total, installment_no
            FROM (
                SELECT id, tipo, valor, alvo, nota, categoria, criado_em, is_internal_movement,
                       NULL::int AS installments_total,
                       NULL::int AS installment_no
                FROM launches
                WHERE user_id = %s
                  AND criado_em >= %s AND criado_em < %s
                  AND tipo NOT IN ('criar_caixinha', 'delete_pocket', 'create_investment', 'delete_investment')
                  {launch_filter_sql}
                {credit_union_sql}
            ) merged
            ORDER BY criado_em DESC, id ASC
            LIMIT %s OFFSET %s
            """,
            (user_id, month_start, month_end, *launch_filter_params, *credit_union_params, limit, offset),
        ),
        # 5) Monthly income/expense totals (sem internas).
        # Compras no cartão entram como 'despesa' alocadas pelo mês em que a
        # FATURA fecha (`credit_bills.period_end`), não pelo `purchased_at`.
        # Assim parcelamento aparece distribuído (1/3 maio, 2/3 junho, 3/3 julho)
        # em vez de tudo no mês da compra. Pagamento da fatura é launch interna,
        # então não dobra.
        _q(
            """
            SELECT tipo, SUM(valor) AS total FROM (
                SELECT tipo, valor
                FROM launches
                WHERE user_id = %s
                  AND criado_em >= %s AND criado_em < %s
                  AND is_internal_movement = false
                UNION ALL
                SELECT 'despesa' AS tipo, ct.valor
                FROM credit_transactions ct
                JOIN credit_bills b ON b.id = ct.bill_id
                WHERE ct.user_id = %s
                  AND ct.is_refund = false
                  AND b.period_end >= %s AND b.period_end < %s
            ) merged
            GROUP BY tipo
            """,
            (
                user_id, month_start, month_end,
                user_id, month_start, month_end,
            ),
        ),
        # 6) Categories (despesas do mês — credit_transactions alocadas por
        # `bill.period_end`, igual query 5).
        _q(
            """
            SELECT COALESCE(categoria, 'sem categoria') AS categoria,
                   SUM(valor) AS total,
                   SUM(cnt)   AS count
            FROM (
                SELECT categoria, valor, 1 AS cnt
                FROM launches
                WHERE user_id = %s
                  AND tipo = 'despesa'
                  AND is_internal_movement = false
                  AND criado_em >= %s AND criado_em < %s
                UNION ALL
                SELECT ct.categoria, ct.valor, 1 AS cnt
                FROM credit_transactions ct
                JOIN credit_bills b ON b.id = ct.bill_id
                WHERE ct.user_id = %s
                  AND ct.is_refund = false
                  AND b.period_end >= %s AND b.period_end < %s
            ) merged
            GROUP BY COALESCE(categoria, 'sem categoria')
            ORDER BY total DESC
            LIMIT 10
            """,
            (
                user_id, month_start, month_end,
                user_id, month_start, month_end,
            ),
        ),
        # 7) Allocations (aportes do mês)
        _q(
            """
            SELECT
                CASE
                    WHEN tipo = 'deposito_caixinha' THEN 'pockets'
                    ELSE 'investments'
                END AS bucket,
                alvo,
                SUM(valor) AS total,
                COUNT(*)   AS count
            FROM launches
            WHERE user_id = %s
              AND criado_em >= %s AND criado_em < %s
              AND is_internal_movement = true
              AND (
                tipo IN ('aporte_investimento', 'deposito_caixinha')
                OR (tipo = 'despesa' AND LOWER(REPLACE(COALESCE(categoria, ''), ' ', '_')) IN (
                    'investimentos', 'investimento_aporte', 'criptomoedas'
                ))
              )
            GROUP BY bucket, alvo
            ORDER BY bucket, total DESC
            """,
            (user_id, month_start, month_end),
        ),
        # 8) Cards + faturas do mês via LATERAL JOIN (1 query, sem N+1)
        _q(
            """
            SELECT
                c.id, c.name, c.closing_day, c.due_day, c.color,
                b.status, b.total, b.paid_amount, b.due_amount,
                b.period_start, b.period_end
            FROM credit_cards c
            LEFT JOIN LATERAL (
                SELECT
                    status,
                    total,
                    COALESCE(paid_amount, 0)                          AS paid_amount,
                    GREATEST(0, total - COALESCE(paid_amount, 0))     AS due_amount,
                    period_start,
                    period_end
                FROM credit_bills
                WHERE card_id = c.id
                  AND period_start >= %s
                  AND period_start < %s
                ORDER BY period_start DESC
                LIMIT 1
            ) b ON TRUE
            WHERE c.user_id = %s
            ORDER BY c.display_order NULLS LAST, c.name
            """,
            (month_start, month_end, user_id),
        ),
        # 9) Daily expenses (bar chart)
        _q(
            f"""
            SELECT EXTRACT(DAY FROM criado_em AT TIME ZONE '{TZ}')::int AS dia,
                   SUM(valor) AS total
            FROM launches
            WHERE user_id = %s
              AND tipo IN ('despesa', 'saida')
              AND is_internal_movement = false
              AND criado_em >= %s AND criado_em < %s
            GROUP BY dia
            ORDER BY dia
            """,
            (user_id, month_start, month_end),
        ),
        # 10) Budgets per category
        _q("SELECT categoria, budget FROM category_budgets WHERE user_id = %s", (user_id,)),
    )

    # Desempacota fetchone-style
    account = account_rows[0] if account_rows else None
    launches_total = int(launches_total_rows[0]["total"] or 0) if launches_total_rows else 0

    # Reformat cards (era loop dentro do bloco de queries)
    cards = []
    for r in card_rows:
        period_end = r["period_end"]
        period_label = (
            f"{_months_pt()[period_end.month - 1]}/{period_end.year}"
            if period_end else None
        )
        cards.append({
            "id": r["id"],
            "name": r["name"],
            "closing_day": r["closing_day"],
            "due_day": r["due_day"],
            "color": r["color"] or "purple",
            "status": r["status"] or "open",
            "total": float(r["total"]) if r["total"] is not None else 0.0,
            "paid_amount": float(r["paid_amount"]) if r["paid_amount"] is not None else 0.0,
            "due_amount": float(r["due_amount"]) if r["due_amount"] is not None else 0.0,
            "period_start": r["period_start"],
            "period_end": period_end,
            "period_label": period_label,
        })

    # Build maps
    monthly_map = {row["tipo"]: float(row["total"]) for row in monthly}
    budget_map  = {r["categoria"]: float(r["budget"]) for r in budget_rows}

    # Merge budgets into categories + detect alerts
    cat_list = []
    alerts   = []
    for c in categories:
        cat      = dict(c)
        cat_name = cat["categoria"]
        spent    = float(cat["total"])
        cat["total"] = spent

        if cat_name in budget_map:
            bgt = budget_map[cat_name]
            pct = round(spent / bgt * 100, 1)
            cat["budget"]     = bgt
            cat["budget_pct"] = pct

            if spent > bgt:
                alerts.append({
                    "type":      "budget_exceeded",
                    "categoria": cat_name,
                    "spent":     spent,
                    "budget":    bgt,
                    "pct":       pct,
                })
            elif pct >= 85:
                alerts.append({
                    "type":      "budget_warning",
                    "categoria": cat_name,
                    "spent":     spent,
                    "budget":    bgt,
                    "pct":       pct,
                })

        cat_list.append(cat)

    # Alertas de cobranças automáticas (recurring_charges) ainda não vistas.
    try:
        async with await db_connect() as _alert_conn:
            async with _alert_conn.cursor() as _alert_cur:
                await _alert_cur.execute(
                    """
                    select rc.id, rc.amount, rc.charged_at, rc.ym,
                           r.name, r.payment_type, r.id as recurring_id
                    from recurring_charges rc
                    join recurring_expenses r on r.id = rc.recurring_id
                    where rc.user_id = %s and rc.acknowledged = false
                    order by rc.charged_at desc
                    limit 10
                    """,
                    (user_id,),
                )
                for r in await _alert_cur.fetchall():
                    alerts.append({
                        "type":         "recurring_charged",
                        "charge_id":    r["id"],
                        "recurring_id": r["recurring_id"],
                        "name":         r["name"],
                        "amount":       float(r["amount"]),
                        "payment_type": r["payment_type"],
                        "ym":           r["ym"],
                        "charged_at":   r["charged_at"].isoformat() if r["charged_at"] else None,
                    })
    except Exception:
        # Tabela pode não existir ainda no init_db da primeira subida — silencia.
        pass

    inc = monthly_map.get("receita", 0.0)
    exp = monthly_map.get("despesa", 0.0)

    allocations = {"investments": {"total": 0.0, "count": 0, "by_target": []},
                   "pockets":     {"total": 0.0, "count": 0, "by_target": []}}
    for r in (allocations_rows or []):
        bucket = r["bucket"]
        if bucket not in allocations:
            continue
        v = float(r["total"] or 0)
        allocations[bucket]["total"] += v
        allocations[bucket]["count"] += int(r["count"] or 0)
        allocations[bucket]["by_target"].append({
            "alvo":  r.get("alvo") or "—",
            "total": v,
            "count": int(r["count"] or 0),
        })
    for bucket in allocations:
        allocations[bucket]["by_target"].sort(key=lambda x: -x["total"])

    return {
        "user_id":            user_id,
        "timestamp":          datetime.now(timezone.utc).isoformat(),
        "year":               y,
        "month":              m,
        "is_current_month":   is_current,
        "balance":            float(account["balance"]) if account else 0.0,
        "pockets":            [dict(r) for r in pockets],
        "investments":        [dict(r) for r in investments],
        "market_rates":       market_rates,
        "recent_launches":    [dict(r) for r in launches],
        "launches_pagination": {
            "page": page,
            "limit": limit,
            "total": launches_total,
            "total_pages": max((launches_total + limit - 1) // limit, 1),
            "filter_type": (filter_type or "all").strip().lower(),
            "query": (query or "").strip(),
        },
        "monthly_income":     inc,
        "monthly_expense":    exp,
        "monthly_allocations": allocations,
        "expense_categories": cat_list,
        "credit_cards":       [dict(r) for r in cards],
        "budgets":            budget_map,
        "alerts":             alerts,
        "daily_expenses":     [{"day": int(r["dia"]), "total": float(r["total"])} for r in daily_rows],
    }

# ─── Monthly history ─────────────────────────────────────────────────────────

async def get_monthly_history(user_id: int, n_months: int = 6) -> list:
    """Returns last n_months of income/expense totals, oldest first."""
    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                f"""
                SELECT TO_CHAR(DATE_TRUNC('month', criado_em), 'YYYY-MM') AS mes,
                       tipo, SUM(valor) AS total
                FROM launches
                WHERE user_id = %s
                  AND criado_em >= NOW() - INTERVAL '{int(n_months)} months'
                  AND tipo IN ('receita', 'despesa')
                  AND is_internal_movement = false
                GROUP BY mes, tipo
                ORDER BY mes
                """,
                (user_id,),
            )
            rows = await cur.fetchall()

    history: dict = {}
    for row in rows:
        k = row["mes"]
        if k not in history:
            history[k] = {"month": k, "income": 0.0, "expense": 0.0}
        if row["tipo"] == "receita":
            history[k]["income"] = float(row["total"])
        elif row["tipo"] == "despesa":
            history[k]["expense"] = float(row["total"])

    return list(history.values())

# ─── CSV export ──────────────────────────────────────────────────────────────

_MESES_PT = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

# Classificação por natureza (alinha com o resumo do dashboard). A natureza
# define a cor no relatório: despesa→vermelho, receita→verde, aporte→azul.
_EXPORT_DESPESA_TIPOS = {"despesa", "saida"}
_EXPORT_RECEITA_TIPOS = {"receita", "entrada"}
_EXPORT_APORTE_OUT    = {"aporte_investimento", "deposito_caixinha"}   # sai da conta
_EXPORT_APORTE_IN     = {"resgate_investimento", "saque_caixinha"}     # volta pra conta
_EXPORT_INVEST_CATS   = {"investimentos", "investimento_aporte", "criptomoedas"}
_EXPORT_TIPO_LABEL = {
    "aporte_investimento": "Aporte", "deposito_caixinha": "Depósito",
    "resgate_investimento": "Resgate", "saque_caixinha": "Saque",
}


def _classify_launch(tipo: str, is_internal: bool, categoria: str = ""):
    """Retorna (natureza, sinal, label) ou None se a ação não entra no relatório.

    natureza ∈ {despesa, receita, aporte}; sinal '+'/'-' = entrou/saiu na conta.
    'aporte' cobre movimentações de investimento e caixinha (is_internal),
    INCLUSIVE as registradas como tipo 'despesa'/'receita' com categoria de
    investimento — mesma regra do resumo do dashboard. Exclui ações
    não-monetárias (criar/deletar) e internas que não são aporte
    (saldo_inicial, ajuste, pagamento_fatura — esta já entra via cartão)."""
    t = (tipo or "").lower()
    if is_internal:
        cat_norm = (categoria or "").lower().replace(" ", "_")
        if t in _EXPORT_APORTE_OUT:
            return ("aporte", "-", _EXPORT_TIPO_LABEL[t])
        if t in _EXPORT_APORTE_IN:
            return ("aporte", "+", _EXPORT_TIPO_LABEL[t])
        if cat_norm in _EXPORT_INVEST_CATS:
            if t in _EXPORT_RECEITA_TIPOS:
                return ("aporte", "+", "Resgate")
            return ("aporte", "-", "Aporte")
        return None   # demais movimentações internas ficam de fora
    if t in _EXPORT_DESPESA_TIPOS:
        return ("despesa", "-", "Despesa")
    if t in _EXPORT_RECEITA_TIPOS:
        return ("receita", "+", "Receita")
    return None


def _item_sort_key(it: dict):
    d = it.get("data")
    if d is None:
        return (0, 0, 0, 0, 0)
    return (d.year, d.month, d.day, getattr(d, "hour", 0), getattr(d, "minute", 0))


async def _fetch_export_items(user_id: int, year: int, month: int) -> list[dict]:
    """Itens monetários do mês p/ relatório (CSV/XLSX/PDF compartilham):
    despesas/receitas reais da conta, aportes/movimentações de investimento e
    caixinha, e compras no cartão (alocadas por bill.period_end, igual ao
    dashboard). Ações não-monetárias ficam de fora."""
    month_start, month_end = _month_range(year, month)
    items: list[dict] = []
    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                """
                SELECT tipo, valor, alvo, nota, categoria, criado_em, is_internal_movement
                FROM launches
                WHERE user_id = %s
                  AND criado_em >= %s AND criado_em < %s
                """,
                (user_id, month_start, month_end),
            )
            for r in await cur.fetchall():
                cls = _classify_launch(r["tipo"], r.get("is_internal_movement"), r.get("categoria"))
                if not cls:
                    continue
                natureza, sign, label = cls
                items.append({
                    "data": r["criado_em"],
                    "natureza": natureza,
                    "label": label,
                    "sign": sign,
                    "categoria": (r.get("categoria") or "").strip(),
                    "descricao": (r.get("alvo") or r.get("nota") or "").strip(),
                    "valor": float(r["valor"]),
                })

            # Compras no cartão de crédito, alocadas pelo fechamento da fatura.
            # `data` exibida = b.period_end (dia em que a fatura fecha) — assim
            # uma compra de 20/04 cuja fatura fecha em maio aparece com data de
            # maio no relatório de maio, em vez de confundir o leitor com a data
            # da compra. Parcelas ganham sufixo "(i/N)" pra distinguir linhas.
            await cur.execute(
                """
                SELECT ct.valor, ct.categoria, ct.nota,
                       ct.installment_no, ct.installments_total,
                       b.period_end, c.name AS card_name
                FROM credit_transactions ct
                JOIN credit_bills b ON b.id = ct.bill_id
                JOIN credit_cards c ON c.id = ct.card_id
                WHERE ct.user_id = %s
                  AND ct.is_refund = false
                  AND b.period_end >= %s AND b.period_end < %s
                """,
                (user_id, month_start, month_end),
            )
            for r in await cur.fetchall():
                desc = (r.get("nota") or "").strip()
                card = (r.get("card_name") or "").strip()
                inst_no = r.get("installment_no")
                inst_total = r.get("installments_total")
                if inst_total and int(inst_total) > 1:
                    desc = f"{desc} ({inst_no}/{inst_total})".strip()
                items.append({
                    "data": r["period_end"],
                    "natureza": "despesa",
                    "label": "Cartão",
                    "sign": "-",
                    "categoria": (r.get("categoria") or "").strip(),
                    "descricao": f"{desc} · {card}".strip(" ·") if card else desc,
                    "valor": float(r["valor"]),
                })

    items.sort(key=_item_sort_key, reverse=True)
    return items


async def build_csv(user_id: int, year: int, month: int) -> str | None:
    items = await _fetch_export_items(user_id, year, month)
    if not items:
        return None
    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(["data", "tipo", "valor", "categoria", "descricao"])
    for it in items:
        d = it["data"]
        w.writerow([
            d.strftime("%Y-%m-%d") if d else "",
            it["label"],
            f"{it['sign']}{it['valor']:.2f}",
            it["categoria"],
            it["descricao"],
        ])
    return buf.getvalue()


def _export_summary(items: list[dict]) -> dict:
    """Receitas, despesas (conta + cartão), aportes e despesas/categoria.

    'aportes' considera só as saídas pra investimento/caixinha."""
    from collections import defaultdict
    receitas = sum(it["valor"] for it in items if it["natureza"] == "receita")
    despesas = sum(it["valor"] for it in items if it["natureza"] == "despesa")
    aportes  = sum(it["valor"] for it in items if it["natureza"] == "aporte" and it["sign"] == "-")
    by_cat: dict[str, float] = defaultdict(float)
    for it in items:
        if it["natureza"] == "despesa":
            by_cat[it["categoria"] or "sem categoria"] += it["valor"]
    cats = sorted(by_cat.items(), key=lambda kv: kv[1], reverse=True)[:10]
    return {
        "receitas": receitas,
        "despesas": despesas,
        "aportes": aportes,
        "by_category": cats,
        "count": len(items),
    }


async def build_xlsx(user_id: int, year: int, month: int) -> bytes | None:
    items = await _fetch_export_items(user_id, year, month)
    if not items:
        return None
    return await asyncio.to_thread(_render_xlsx, items)


def _render_xlsx(items: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    BRAND = "2563EB"   # azul PigBank (cabeçalho)
    ZEBRA = "EFF6FF"   # azul bem claro
    POS   = "16A34A"   # verde = entrou
    NEG   = "DC2626"   # vermelho = saiu

    wb = Workbook()
    ws = wb.active
    ws.title = "Lançamentos"
    ws.append(["Data", "Tipo", "Categoria", "Descrição", "Valor"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BRAND)
        cell.alignment = Alignment(vertical="center")

    for idx, it in enumerate(items, start=2):
        d = it["data"]
        signed = it["valor"] if it["sign"] == "+" else -it["valor"]
        color = POS if it["sign"] == "+" else NEG   # verde = entrou, vermelho = saiu
        ws.append([
            d.strftime("%d/%m/%Y") if d else "",
            it["label"],
            it["categoria"],
            it["descricao"],
            round(signed, 2),
        ])
        ws.cell(row=idx, column=2).font = Font(bold=True, color=color)
        valor_cell = ws.cell(row=idx, column=5)
        valor_cell.font = Font(bold=True, color=color)
        valor_cell.number_format = '"R$" #,##0.00'
        if idx % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=idx, column=col).fill = PatternFill("solid", fgColor=ZEBRA)

    for i, width in enumerate((14, 12, 18, 38, 16), start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(items) + 1}"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


async def build_pdf(user_id: int, year: int, month: int) -> bytes | None:
    items = await _fetch_export_items(user_id, year, month)
    if not items:
        return None
    # Saldo atual da conta (não escopado ao mês) — mesmo número do card do dashboard.
    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute("SELECT balance FROM accounts WHERE user_id = %s", (user_id,))
            row = await cur.fetchone()
    balance = float(row["balance"]) if row else 0.0
    return await asyncio.to_thread(_render_pdf, items, year, month, balance)


def _render_pdf(items: list[dict], year: int, month: int, balance: float = 0.0) -> bytes:
    from datetime import datetime as _dt
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.graphics.shapes import Drawing, Rect
    from utils_text import fmt_brl

    # Paleta "Azul PigBank" (sai do roxo cara-de-Nubank)
    BRAND    = colors.HexColor("#2563EB")
    BRAND_D  = colors.HexColor("#1E40AF")
    POS      = colors.HexColor("#16A34A")
    NEG      = colors.HexColor("#DC2626")
    INK      = colors.HexColor("#1E293B")
    MUTED    = colors.HexColor("#64748B")
    LINE     = colors.HexColor("#E2E8F0")
    ZEBRA    = colors.HexColor("#EFF6FF")
    POS_BG   = colors.HexColor("#ECFDF5")
    NEG_BG   = colors.HexColor("#FEF2F2")
    BAR_BG   = colors.HexColor("#E2E8F0")
    HEAD_SUB = colors.HexColor("#DBEAFE")

    base = getSampleStyleSheet()

    def par(text, size=9, color=INK, bold=False, align=TA_LEFT):
        return Paragraph(str(text), ParagraphStyle(
            "p", parent=base["Normal"],
            fontName="Helvetica-Bold" if bold else "Helvetica",
            fontSize=size, textColor=color, alignment=align, leading=size + 3,
        ))

    summary = _export_summary(items)
    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio, pagesize=A4,
        topMargin=14 * mm, bottomMargin=15 * mm, leftMargin=15 * mm, rightMargin=15 * mm,
        title=f"Extrato {_MESES_PT[month]}/{year}", author="PigBank AI",
    )
    W = doc.width
    el = []

    # ── Banner ───────────────────────────────────────────────────────────
    banner = Table(
        [[par("PigBank AI", 22, colors.white, bold=True)],
         [par(f"Extrato de {_MESES_PT[month]} de {year}", 11, HEAD_SUB)]],
        colWidths=[W],
    )
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BRAND),
        ("LEFTPADDING", (0, 0), (-1, -1), 18),
        ("TOPPADDING", (0, 0), (0, 0), 16),
        ("BOTTOMPADDING", (0, 0), (0, 0), 1),
        ("TOPPADDING", (0, 1), (0, 1), 0),
        ("BOTTOMPADDING", (0, 1), (0, 1), 16),
    ]))
    el.append(banner)
    el.append(Spacer(1, 16))

    # ── Cards de resumo (Receitas | Despesas | Saldo) ────────────────────
    # Despesas = despesas + aportes (aporte é saída de caixa).
    # Saldo = saldo atual da conta (≠ "sobrou do mês"); mesmo número do dashboard.
    gap = 8
    card_w = (W - 2 * gap) / 3.0
    SALDO_BG = colors.HexColor("#F1F5F9")
    saldo_color = POS if balance >= 0 else NEG
    despesas_total = summary["despesas"] + summary["aportes"]
    kpi = Table(
        [[par("RECEITAS", 8, MUTED, bold=True), "", par("DESPESAS", 8, MUTED, bold=True), "",
          par("SALDO", 8, MUTED, bold=True)],
         [par(fmt_brl(summary["receitas"]), 13, POS, bold=True), "",
          par(fmt_brl(despesas_total), 13, NEG, bold=True), "",
          par(fmt_brl(balance), 13, saldo_color, bold=True)]],
        colWidths=[card_w, gap, card_w, gap, card_w],
    )
    kpi.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 1), POS_BG),
        ("BACKGROUND", (2, 0), (2, 1), NEG_BG),
        ("BACKGROUND", (4, 0), (4, 1), SALDO_BG),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 11),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 2),
        ("TOPPADDING", (0, 1), (-1, 1), 0),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 11),
        ("LEFTPADDING", (1, 0), (1, -1), 0), ("RIGHTPADDING", (1, 0), (1, -1), 0),
        ("LEFTPADDING", (3, 0), (3, -1), 0), ("RIGHTPADDING", (3, 0), (3, -1), 0),
    ]))
    el.append(kpi)

    # ── Despesas por categoria (com barra de proporção) ──────────────────
    if summary["by_category"]:
        el.append(Spacer(1, 20))
        el.append(par("Despesas por categoria", 13, BRAND_D, bold=True))
        el.append(Spacer(1, 7))
        total_desp = summary["despesas"] or 1.0
        bar_w = W * 0.26

        def make_bar(frac):
            d = Drawing(bar_w, 9)
            d.add(Rect(0, 0, bar_w, 9, fillColor=BAR_BG, strokeColor=None))
            d.add(Rect(0, 0, max(3.0, bar_w * min(frac, 1.0)), 9, fillColor=BRAND, strokeColor=None))
            d.hAlign = "LEFT"
            return d

        cat_data = []
        for cat, total in summary["by_category"]:
            frac = float(total) / total_desp
            cat_data.append([
                par(cat, 9, INK),
                make_bar(frac),
                par(f"{fmt_brl(total)}  ({frac * 100:.0f}%)", 9, MUTED, align=TA_RIGHT),
            ])
        cat_table = Table(cat_data, colWidths=[W * 0.36, W * 0.30, W * 0.34])
        cat_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, LINE),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LEFTPADDING", (1, 0), (1, -1), 6),
            ("RIGHTPADDING", (1, 0), (1, -1), 10),
        ]))
        el.append(cat_table)

    # ── Lançamentos ──────────────────────────────────────────────────────
    el.append(Spacer(1, 20))
    el.append(par("Lançamentos", 13, BRAND_D, bold=True))
    el.append(Spacer(1, 7))
    head = [par(h, 8, colors.white, bold=True, align=(TA_RIGHT if h == "Valor" else TA_LEFT))
            for h in ("Data", "Tipo", "Categoria", "Descrição", "Valor")]
    lanc_rows = [head]
    for it in items:
        color = POS if it["sign"] == "+" else NEG   # verde = entrou, vermelho = saiu
        d = it["data"]
        lanc_rows.append([
            par(d.strftime("%d/%m") if d else "", 8, MUTED),
            par(it["label"], 8, color, bold=True),
            par(it["categoria"] or "-", 8, INK),
            par((it["descricao"] or "-")[:55], 8, INK),
            par(f"{it['sign']} {fmt_brl(it['valor'])}", 8, color, bold=True, align=TA_RIGHT),
        ])
    lanc = Table(
        lanc_rows,
        colWidths=[W * 0.14, W * 0.13, W * 0.20, W * 0.38, W * 0.15],
        repeatRows=1,
    )
    lanc.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ZEBRA]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    el.append(lanc)

    el.append(Spacer(1, 16))
    el.append(par(
        f"Gerado em {_dt.now().strftime('%d/%m/%Y às %H:%M')}  •  pigbankai.com  •  {summary['count']} lançamentos",
        8, MUTED, align=TA_CENTER,
    ))

    doc.build(el)
    return bio.getvalue()

# ─── Connection manager ───────────────────────────────────────────────────────

class ConnectionManager:
    # Cap por usuário: evita um cliente abrir conexões em loop e esgotar
    # memória/file descriptors do servidor. 5 cobre multi-aba legítimo.
    MAX_CONNECTIONS_PER_USER = 5

    def __init__(self):
        # active[user_id][ws] = {"year": int, "month": int}
        self.active: Dict[int, Dict[WebSocket, dict]] = {}

    async def connect(self, ws: WebSocket, user_id: int, year: int, month: int) -> bool:
        await ws.accept()
        if len(self.active.get(user_id, {})) >= self.MAX_CONNECTIONS_PER_USER:
            await ws.close(code=1008, reason="Limite de conexões simultâneas atingido")
            return False
        self.active.setdefault(user_id, {})[ws] = {"year": year, "month": month}
        return True

    def disconnect(self, ws: WebSocket, user_id: int):
        if user_id in self.active:
            self.active[user_id].pop(ws, None)
            if not self.active[user_id]:
                self.active.pop(user_id, None)

    def set_month(self, ws: WebSocket, user_id: int, year: int, month: int):
        if user_id in self.active and ws in self.active[user_id]:
            self.active[user_id][ws]["year"] = year
            self.active[user_id][ws]["month"] = month

    def get_month(self, ws: WebSocket, user_id: int):
        info = self.active.get(user_id, {}).get(ws)
        if info:
            return info["year"], info["month"]

        now = datetime.now(timezone.utc)
        return now.year, now.month

    async def send_to(self, ws: WebSocket, payload: str):
        await ws.send_text(payload)

manager = ConnectionManager()

# ─── App startup ──────────────────────────────────────────────────────────────

from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app: FastAPI):
    _t0 = _startup_time.monotonic()

    async def _startup_required(label: str, coro):
        try:
            return await asyncio.wait_for(coro, timeout=STARTUP_STEP_TIMEOUT)
        except asyncio.TimeoutError as exc:
            print(
                f"ERROR: Startup travou em '{label}' por mais de {STARTUP_STEP_TIMEOUT}s.",
                file=sys.stderr,
                flush=True,
            )
            raise RuntimeError(f"Startup timeout: {label}") from exc

    async def _startup_optional(label: str, coro, default=None):
        try:
            return await asyncio.wait_for(coro, timeout=STARTUP_STEP_TIMEOUT)
        except asyncio.TimeoutError:
            print(
                f"WARNING: Startup ignorou '{label}' após {STARTUP_STEP_TIMEOUT}s.",
                file=sys.stderr,
                flush=True,
            )
            return default
        except Exception as exc:
            print(f"WARNING: Startup ignorou '{label}': {exc}", file=sys.stderr, flush=True)
            return default

    # ── 1. DB health check ────────────────────────────────────────────────────
    try:
        async def _db_health_check():
            async with await db_connect() as conn:
                async with conn.cursor() as cur:
                    await cur.execute("SELECT 1")

        await _startup_required("database health check", _db_health_check())
        print("OK: Database connected", flush=True)
    except Exception as exc:
        print(f"ERROR: Database connection failed: {exc}", file=sys.stderr, flush=True)
        raise

    # ── 2. Schema completo (idempotente) ─────────────────────────────────────
    # init_db() era chamado apenas pelo bot.py (Discord); se o bot falhasse ou
    # subisse depois do dashboard, tabelas novas (auth_sessions, audit_events…)
    # nao existiam quando a primeira request chegasse. Roda aqui para nao
    # depender do bot estar saudavel.
    #
    # SKIP_INIT_DB=1: pula migrations no startup. Util quando o dashboard local
    # aponta pra Railway prod (so pra ver /admin): o deploy do Railway ja roda
    # init_db, nao precisa repetir via internet — round-trips somam >12s e
    # adquirem locks em tabelas reais.
    skip_init = (os.getenv("SKIP_INIT_DB") or "").strip() in ("1", "true", "True")
    if skip_init:
        print("⚠️  SKIP_INIT_DB=1 — pulando migrations no startup", flush=True)
    else:
        from db import init_db
        await _startup_required("init_db", asyncio.to_thread(init_db))
        print("OK: Schema migrado", flush=True)

        # ── 3. Setup de tabelas adicionais em paralelo ───────────────────────
        await _startup_required(
            "setup de tabelas",
            asyncio.gather(
                asyncio.to_thread(ensure_account_deletion_columns),
                ensure_admin_tables(),
            ),
        )
        print("OK: Account deletion controls ready", flush=True)
        print("OK: Admin observability tables ready", flush=True)

    # ── 3. Warnings + detecção de usuário em paralelo ─────────────────────────
    async def _resolve_uid() -> int | None:
        if DASHBOARD_USER_ID:
            return int(DASHBOARD_USER_ID)
        rows = await list_users()
        if not rows:
            print("WARNING: No users found in database.", flush=True)
            return None
        if len(rows) == 1:
            uid = rows[0]["id"]
            print(f"INFO: Auto-detected user ID: {uid}", flush=True)
        else:
            ids = [r["id"] for r in rows]
            uid = ids[0]
            print(f"INFO: Multiple users {ids}. Using first: {uid}", flush=True)
        return uid

    uid, _ = await asyncio.gather(
        _startup_optional("detecção de usuário padrão", _resolve_uid()),
        _startup_optional("avisos administrativos de startup", log_admin_startup_warnings()),
    )

    app.state.default_user_id = uid
    _port = int(os.environ.get("PORT", "8000"))
    if uid:
        print(f"Dashboard: http://localhost:{_port}/", flush=True)
        print(f"WebSocket: ws://localhost:{_port}/ws/{uid}", flush=True)

    # ── 4. Background tasks com imports lazy ──────────────────────────────────
    # Os imports pesados ficam dentro dos wrappers: só são carregados quando
    # a coroutine executa (após o yield), fora da janela de startup.
    async def _wa_worker():
        try:
            await asyncio.sleep(1)
            from adapters.whatsapp.wa_app import _worker_loop  # noqa: PLC0415
            await _worker_loop()
        except Exception as exc:
            print(f"[wa_worker] erro: {exc}", file=sys.stderr)

    async def _wa_daily():
        try:
            await asyncio.sleep(1)
            from adapters.whatsapp.wa_app import _daily_report_loop  # noqa: PLC0415
            await _daily_report_loop()
        except Exception as exc:
            print(f"[wa_daily] erro: {exc}", file=sys.stderr)

    async def _engagement():
        try:
            await asyncio.sleep(1)
            from core.services.engagement_scheduler import run_engagement_loop  # noqa: PLC0415
            await run_engagement_loop()
        except Exception as exc:
            print(f"[engagement] erro: {exc}", file=sys.stderr)

    async def _investment_accrual():
        try:
            await asyncio.sleep(1)
            from core.services.investment_scheduler import run_investment_accrual_loop  # noqa: PLC0415
            await run_investment_accrual_loop()
        except Exception as exc:
            print(f"[investment_accrual] erro: {exc}", file=sys.stderr)

    async def _recurring_charger():
        try:
            await asyncio.sleep(5)
            from core.services.recurring_charger import run_recurring_charger_loop  # noqa: PLC0415
            await run_recurring_charger_loop()
        except Exception as exc:
            print(f"[recurring_charger] erro: {exc}", file=sys.stderr)

    async def _proactive_ai():
        try:
            await asyncio.sleep(1)
            from core.services.proactive_ai_scheduler import run_proactive_ai_loop  # noqa: PLC0415
            await run_proactive_ai_loop()
        except Exception as exc:
            print(f"[proactive_ai] erro: {exc}", file=sys.stderr)

    async def _account_deletion_worker():
        while True:
            try:
                await asyncio.sleep(10)
                results = await asyncio.to_thread(process_due_account_deletions)
                if results:
                    print(f"[account_deletion] resultados processados: {len(results)}", flush=True)
                    from core.services.email_service import send_account_deletion_completed_email  # noqa: PLC0415
                    for result in results:
                        if (result or {}).get("error"):
                            print(f"[account_deletion] erro ao remover user_id={result.get('user_id')}: {result.get('error')}", file=sys.stderr)
                            continue
                        email = (result or {}).get("email")
                        if (result or {}).get("deleted") and email:
                            await asyncio.to_thread(send_account_deletion_completed_email, email)
                await asyncio.sleep(60 * 60)
            except asyncio.CancelledError:
                raise
            except Exception as exc:
                print(f"[account_deletion] erro: {exc}", file=sys.stderr)

    _elapsed = _startup_time.monotonic() - _t0
    print(f"[app] Startup interno concluído em {_elapsed:.1f}s.", flush=True)

    tasks = []
    if RUN_BACKGROUND_TASKS:
        tasks.extend(
            [
                asyncio.create_task(_wa_worker(), name="wa_worker"),
                asyncio.create_task(_wa_daily(), name="wa_daily"),
                asyncio.create_task(_engagement(), name="engagement"),
                asyncio.create_task(_investment_accrual(), name="investment_accrual"),
                asyncio.create_task(_account_deletion_worker(), name="account_deletion"),
                asyncio.create_task(_recurring_charger(), name="recurring_charger"),
                asyncio.create_task(_proactive_ai(), name="proactive_ai"),
            ]
        )
    else:
        print("[app] Background tasks desativadas neste processo.", flush=True)

    yield

    # Shutdown: cancela tasks e aguarda com timeout para não travar
    for t in tasks:
        t.cancel()
    if tasks:
        await asyncio.wait(tasks, timeout=5)
    for t in tasks:
        if not t.done():
            print(f"[lifespan] task '{t.get_name()}' não encerrou a tempo.", file=sys.stderr)

app = FastAPI(
    title="Finance Dashboard",
    docs_url=None,
    redoc_url=None,
    openapi_url=None,
    lifespan=lifespan,
)

# Middleware de log de erros HTTP (definido em core/admin_dashboard.py)
app.middleware("http")(admin_error_logging_middleware)

CSRF_COOKIE_NAME = "csrf_token"
CSRF_HEADER_NAME = "x-csrf-token"
CSRF_COOKIE_MAX_AGE = 86400
CSRF_SAFE_METHODS = {"GET", "HEAD", "OPTIONS", "TRACE"}
CSRF_EXEMPT_PATHS = {
    "/billing/webhook",
    "/open-finance/pluggy/webhook",
    "/wa/webhook",
    "/webhook",
}

_SECURITY_HEADERS = {
    "Strict-Transport-Security": "max-age=31536000; includeSubDomains",
    "X-Frame-Options": "DENY",
    "X-Content-Type-Options": "nosniff",
    "Referrer-Policy": "strict-origin-when-cross-origin",
    "Permissions-Policy": "camera=(), microphone=(), geolocation=()",
    "Content-Security-Policy": (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' "
        "https://cdnjs.cloudflare.com https://cdn.pluggy.ai https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' "
        "https://cdnjs.cloudflare.com https://cdn.jsdelivr.net; "
        "img-src 'self' data: blob: https:; "
        "font-src 'self' data: "
        "https://cdnjs.cloudflare.com https://cdn.jsdelivr.net; "
        "connect-src 'self' https: wss:; "
        "frame-src 'self' https://cdn.pluggy.ai; "
        "frame-ancestors 'none'; "
        "base-uri 'self'; "
        "form-action 'self'; "
        "object-src 'none'"
    ),
}

@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    for header, value in _SECURITY_HEADERS.items():
        response.headers.setdefault(header, value)
    return response


def _make_csrf_token() -> str:
    return secrets.token_urlsafe(32)


def _set_csrf_cookie(response: Response, token: str) -> None:
    response.set_cookie(
        CSRF_COOKIE_NAME,
        token,
        httponly=False,
        secure=COOKIE_SECURE,
        samesite="strict",
        max_age=CSRF_COOKIE_MAX_AGE,
    )


def _csrf_exempt(path: str) -> bool:
    return path in CSRF_EXEMPT_PATHS


@app.middleware("http")
async def csrf_middleware(request: Request, call_next):
    token = request.cookies.get(CSRF_COOKIE_NAME) or ""

    if request.method.upper() not in CSRF_SAFE_METHODS and not _csrf_exempt(request.url.path):
        header_token = request.headers.get(CSRF_HEADER_NAME) or ""
        if not token or not header_token or not secrets.compare_digest(token, header_token):
            return JSONResponse(
                status_code=403,
                content={"detail": "Token CSRF inválido ou ausente."},
                headers={"Cache-Control": "no-store"},
            )

    response = await call_next(request)
    if request.method.upper() in CSRF_SAFE_METHODS and not token:
        _set_csrf_cookie(response, _make_csrf_token())
    return response

# ─── WhatsApp webhook routes (lazy import) ───────────────────────────────────
# Importar wa_app no nível de módulo puxava toda a cadeia de lógica do bot
# (wa_client, wa_runtime, handle_incoming, db, bcrypt, requests…) adicionando
# ~1s ao startup. Com wrappers lazy, o import só acontece na 1ª requisição.

async def _wa_verify(request: Request):
    from adapters.whatsapp.wa_app import wa_verify  # noqa: PLC0415
    return await wa_verify(request)

async def _wa_webhook(request: Request):
    from adapters.whatsapp.wa_app import wa_webhook  # noqa: PLC0415
    return await wa_webhook(request)

async def _wa_simulate(request: Request):
    from adapters.whatsapp.wa_app import wa_simulate  # noqa: PLC0415
    payload = await request.json()
    return await wa_simulate(payload)

# Mantem compatibilidade com a rota antiga `/webhook`, usada em configs legadas.
app.add_api_route("/wa/webhook",     _wa_verify,   methods=["GET"])
app.add_api_route("/wa/webhook",     _wa_webhook,  methods=["POST"])
app.add_api_route("/webhook",        _wa_verify,   methods=["GET"])
app.add_api_route("/webhook",        _wa_webhook,  methods=["POST"])
if ENABLE_DEV_ENDPOINTS:
    app.add_api_route("/wa/dev/simulate", _wa_simulate, methods=["POST"])

# ─── Rate limiting ────────────────────────────────────────────────────────────

RATE_LIMIT_DETAIL = "Muitas tentativas. Aguarde alguns minutos e tente novamente."
EMAIL_RATE_LIMITS = {
    "register": (3, 60 * 60),
    "login": (5, 60),
    "forgot-password": (3, 60 * 60),
}


def _normalize_rate_limit_email(email: str) -> str:
    return (email or "").strip().lower()


async def _check_persistent_rate_limit(bucket: str, identifier: str, max_attempts: int, window_seconds: int) -> None:
    if not identifier:
        return

    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                """
                INSERT INTO auth_rate_limits (bucket, identifier, window_started_at, attempts, updated_at)
                VALUES (%s, %s, NOW(), 1, NOW())
                ON CONFLICT (bucket, identifier) DO UPDATE SET
                    window_started_at = CASE
                        WHEN auth_rate_limits.window_started_at <= NOW() - (%s * INTERVAL '1 second')
                        THEN NOW()
                        ELSE auth_rate_limits.window_started_at
                    END,
                    attempts = CASE
                        WHEN auth_rate_limits.window_started_at <= NOW() - (%s * INTERVAL '1 second')
                        THEN 1
                        ELSE auth_rate_limits.attempts + 1
                    END,
                    updated_at = NOW()
                RETURNING
                    attempts,
                    EXTRACT(EPOCH FROM (NOW() - window_started_at)) AS elapsed_seconds
                """,
                (bucket, identifier, window_seconds, window_seconds),
            )
            row = await cur.fetchone()
        await conn.commit()

    attempts = int(row["attempts"] or 0)
    if attempts > max_attempts:
        elapsed_seconds = float(row["elapsed_seconds"] or 0)
        retry_after = max(1, int(window_seconds - elapsed_seconds))
        raise HTTPException(
            status_code=429,
            detail=RATE_LIMIT_DETAIL,
            headers={"Retry-After": str(retry_after)},
        )


async def _check_auth_rate_limits(action: str, request: Request, email: str) -> None:
    limit = EMAIL_RATE_LIMITS.get(action)
    normalized_email = _normalize_rate_limit_email(email)
    if not limit:
        return

    max_attempts, window_seconds = limit
    client_ip = get_remote_address(request)
    await _check_persistent_rate_limit(
        action,
        f"ip:{client_ip}",
        max_attempts,
        window_seconds,
    )
    await _check_persistent_rate_limit(
        action,
        f"email:{normalized_email}",
        max_attempts,
        window_seconds,
    )


limiter = Limiter(key_func=get_remote_address, default_limits=["200/minute"])
app.state.limiter = limiter


async def rate_limit_exceeded_handler(request: Request, exc: RateLimitExceeded):
    return JSONResponse(
        status_code=429,
        content={"detail": RATE_LIMIT_DETAIL},
        headers={"Retry-After": "60"},
    )


app.add_exception_handler(RateLimitExceeded, rate_limit_exceeded_handler)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://pigbankai.com"],
    allow_methods=["GET", "POST", "PATCH", "DELETE"],
    allow_credentials=True,
    allow_headers=["*"],
)

# ─── Admin dashboard routes (delegado para core/admin_dashboard.py) ───────────
register_admin_routes(app, HERE, JWT_SECRET, limiter)

# ─── Auth helpers ────────────────────────────────────────────────────────────

_bearer = HTTPBearer(auto_error=False)
AUTH_COOKIE_NAME = "auth_token"
# Access token TTL: 15min (curto, vai em toda request).
# Renovação automática via refresh_token (14d, idle 7d).
AUTH_COOKIE_MAX_AGE = 15 * 60
DASHBOARD_COOKIE_NAME = "dashboard_token"
REFRESH_COOKIE_NAME = "refresh_token"
REFRESH_COOKIE_MAX_AGE = 14 * 24 * 3600  # 14 dias absolutos
REFRESH_COOKIE_PATH = "/auth/refresh"     # só vai em request específica

def _make_jwt(user_id: int, email: str, *, jti: str | None = None) -> str:
    from datetime import timedelta
    payload: dict[str, Any] = {
        "sub": str(user_id),
        "email": email,
        "type": "auth",
        "exp": datetime.now(timezone.utc) + timedelta(minutes=15),
    }
    if jti:
        payload["jti"] = jti
    return pyjwt.encode(payload, JWT_SECRET, algorithm="HS256")


def _issue_session_token(user_id: int, email: str, request: Request) -> tuple[str, str, str]:
    """Cria uma sessao em auth_sessions + emite access JWT + refresh token.

    Retorna (access_jwt, jti, refresh_token_plain). O jti deve ser passado para
    `_set_dashboard_cookie` (dashboard_token também session-bound) e o refresh
    deve ser setado no cookie `refresh_token` via `_set_refresh_cookie`.

    Use em todo lugar que emite o cookie auth_token (login, OAuth, signup,
    magic_link, etc).
    """
    from core.refresh_tokens import create_refresh_token

    ip = get_remote_address(request) or None
    ua = request.headers.get("user-agent") or None
    jti = create_session(user_id, ip=ip, user_agent=ua)
    access = _make_jwt(user_id, email, jti=jti)
    refresh = create_refresh_token(user_id, session_jti=jti, ip=ip, user_agent=ua)
    return access, jti, refresh


def _decode_jwt(token: str) -> dict | None:
    try:
        return pyjwt.decode(token, JWT_SECRET, algorithms=["HS256"])
    except Exception:
        return None


def _set_auth_cookie(response: Response, token: str) -> None:
    # samesite=lax (vs strict): permite que clicar em link externo (WhatsApp,
    # email, Slack) que aponte pra pigbankai chegue com o cookie ja anexado.
    # Strict bloqueava magic links do bot — user caia em /?login_required.
    # Protecao CSRF mantida pelo csrf_token cookie + header x-csrf-token
    # (validados no csrf_middleware, mesma protecao que Stripe/GitHub usam).
    response.set_cookie(
        AUTH_COOKIE_NAME,
        token,
        httponly=True,
        secure=COOKIE_SECURE,
        samesite="lax",
        max_age=AUTH_COOKIE_MAX_AGE,
    )


def _set_dashboard_cookie(response: Response, user_id: int, *, jti: str | None = None) -> str:
    token = make_dashboard_token(user_id, hours=DASHBOARD_SESSION_HOURS, jti=jti)
    response.set_cookie(
        DASHBOARD_COOKIE_NAME,
        token,
        httponly=True,
        secure=COOKIE_SECURE,
        samesite="lax",
        max_age=int(DASHBOARD_SESSION_HOURS * 3600),
    )
    return token


def _set_refresh_cookie(response: Response, refresh_token: str) -> None:
    """Seta o refresh_token cookie com path restrito a /auth/refresh.
    Significa que esse cookie só viaja na 1 request específica de renovação —
    muito menos exposto que o access cookie (que vai em toda request)."""
    response.set_cookie(
        REFRESH_COOKIE_NAME,
        refresh_token,
        httponly=True,
        secure=COOKIE_SECURE,
        samesite="lax",
        max_age=REFRESH_COOKIE_MAX_AGE,
        path=REFRESH_COOKIE_PATH,
    )


def _expire_cookie(response: Response, name: str, domain: str | None = None) -> None:
    response.delete_cookie(
        name,
        path="/",
        domain=domain,
        httponly=True,
        secure=COOKIE_SECURE,
        samesite="strict",
    )


def _clear_session_cookies(response: Response) -> None:
    domains: list[str | None] = [None]
    host = urllib.parse.urlparse(DASHBOARD_URL).hostname
    if host:
        domains.extend([host, f".{host}"])
    for domain in domains:
        _expire_cookie(response, AUTH_COOKIE_NAME, domain)
        _expire_cookie(response, DASHBOARD_COOKIE_NAME, domain)
        _expire_cookie(response, CSRF_COOKIE_NAME, domain)
    # Refresh cookie tem path restrito — limpa com mesmo path pra remover
    response.delete_cookie(
        REFRESH_COOKIE_NAME, path=REFRESH_COOKIE_PATH,
        httponly=True, secure=COOKIE_SECURE, samesite="lax",
    )


def _no_store(response: Response) -> Response:
    response.headers["Cache-Control"] = "no-store"
    response.headers["Pragma"] = "no-cache"
    return response


def _dashboard_url(path: str = "/app", view: str | None = None) -> str:
    url = f"{DASHBOARD_URL}{path}"
    if view:
        url += f"?view={urllib.parse.quote(view)}"
    return url


def _post_login_url() -> str:
    """URL para a qual o usuário deve ser direcionado logo após login.
    O campo `dashboard_url` nas respostas de auth aponta para cá."""
    return _dashboard_url("/home")


def _get_auth_token_from_request(
    request: Request,
    creds: HTTPAuthorizationCredentials | None = None,
) -> str | None:
    if creds and creds.credentials:
        return creds.credentials
    cookie_token = (request.cookies.get(AUTH_COOKIE_NAME) or "").strip()
    return cookie_token or None


def _build_whatsapp_onboarding_link(user_id: int, minutes_valid: int = 15) -> str:
    if not WHATSAPP_NUMBER:
        return ""
    safe_number = "".join(ch for ch in WHATSAPP_NUMBER if ch.isdigit())
    if not safe_number:
        return ""
    text = urllib.parse.quote("Olá")
    return f"https://api.whatsapp.com/send?phone={safe_number}&text={text}"


def _raise_if_account_scheduled_for_deletion(user_id: int) -> None:
    deletion = is_account_scheduled_for_deletion(int(user_id))
    if deletion:
        scheduled = deletion.get("deletion_scheduled_for")
        scheduled_txt = scheduled.isoformat() if hasattr(scheduled, "isoformat") else str(scheduled)
        raise HTTPException(
            status_code=403,
            detail=f"Esta conta está agendada para exclusão em {scheduled_txt}.",
        )

async def _get_current_user(
    request: Request,
    creds: HTTPAuthorizationCredentials | None = Depends(_bearer),
) -> int:
    token = _get_auth_token_from_request(request, creds)
    if not token:
        raise HTTPException(status_code=401, detail="Token não fornecido.")
    payload = _decode_jwt(token)
    if not payload or payload.get("type") != "auth":
        raise HTTPException(status_code=401, detail="Token inválido ou expirado.")
    request.state.auth_payload = payload
    user_id = int(payload["sub"])

    # Sessao por jti: tokens novos sempre carregam jti; tokens legados (sem
    # jti) sao grandfathered (rollout sem kick mass) ate expirarem em 24h.
    jti = payload.get("jti")
    if jti:
        session = await asyncio.to_thread(get_active_session, jti)
        if not session or int(session.get("user_id") or 0) != user_id:
            raise HTTPException(status_code=401, detail="Sessão encerrada. Faça login novamente.")
        request.state.session_jti = jti
        # Atualiza last_seen com debounce; falha silenciosa.
        asyncio.create_task(asyncio.to_thread(touch_session, jti))

    _raise_if_account_scheduled_for_deletion(user_id)
    return user_id


def require_pro_feature(feature: str = "generic"):
    """
    Dependency factory que valida JWT (via _get_current_user) e exige plano Pro.
    Uso: `Depends(require_pro_feature("ofx_import"))` em rotas Pro-only.
    Bloqueio retorna 403 com payload `{"error": "pro_required", "feature": ...}`
    para o frontend abrir modal de upgrade contextual.
    """
    from core.services.plan_service import is_pro

    async def _dep(user_id: int = Depends(_get_current_user)) -> int:
        if not is_pro(user_id):
            raise HTTPException(
                status_code=403,
                detail={"error": "pro_required", "feature": feature},
            )
        return user_id

    return _dep


def _require_pro(user_id: int, feature: str) -> None:
    """
    Variante inline pra endpoints que ja fazem _authorize_dashboard_access no body
    (em vez de Depends). Mesmo payload de 403 que require_pro_feature.
    """
    from core.services.plan_service import is_pro

    if not is_pro(user_id):
        raise HTTPException(
            status_code=403,
            detail={"error": "pro_required", "feature": feature},
        )


def _extract_bearer_token(request: Request) -> str | None:
    auth = request.headers.get("authorization", "").strip()
    if not auth:
        return None
    scheme, _, token = auth.partition(" ")
    if scheme.lower() != "bearer" or not token:
        return None
    return token.strip()


def _resolve_dashboard_user_id(request: Request) -> int:
    token = (
        _extract_bearer_token(request)
        or (request.cookies.get(DASHBOARD_COOKIE_NAME) or "").strip()
    )
    payload = decode_dashboard_token_full(token or "")
    if not payload:
        raise HTTPException(status_code=401, detail="Token de dashboard inválido ou expirado.")
    user_id = payload["user_id"]
    jti = payload.get("jti")
    # Tokens com jti: validar contra auth_sessions (revogacao instantanea).
    # Tokens sem jti (legacy / rollout) sao grandfathered ate expirarem.
    if jti:
        session = get_active_session(jti)
        if not session or int(session.get("user_id") or 0) != user_id:
            raise HTTPException(status_code=401, detail="Sessão encerrada. Faça login novamente.")
        request.state.session_jti = jti
    return int(user_id)


def _authorize_dashboard_access(request: Request, user_id: int) -> int:
    current_user_id = _resolve_dashboard_user_id(request)
    if current_user_id != int(user_id):
        raise HTTPException(status_code=403, detail="Acesso negado para este usuário.")
    _raise_if_account_scheduled_for_deletion(current_user_id)
    return current_user_id

# ─── Auth models ─────────────────────────────────────────────────────────────

class RegisterBody(BaseModel):
    email: str
    password: str
    phone: str
    name: str | None = None

class LoginBody(BaseModel):
    email: str
    password: str

class EmailBody(BaseModel):
    email: str

class VerifyEmailBody(BaseModel):
    email: str
    code: str

class ResetPasswordBody(BaseModel):
    token: str
    new_password: str


class DeleteAccountBody(BaseModel):
    password: str


class DataExportBody(BaseModel):
    password: str


class SecurityContactPayload(BaseModel):
    email: str | None = None
    phone: str | None = None
    display_name: str | None = None


class DashboardLinkBody(BaseModel):
    code: str


# ─── Auth endpoints ──────────────────────────────────────────────────────────

@app.get("/auth/validate")
@limiter.limit("120/minute")
async def auth_validate(request: Request, response: Response):
    """
    Valida uma sessão de dashboard usando apenas cookie HttpOnly ou Bearer.
    Magic links devem passar pela rota /d/{code}, que consome o código e redireciona sem token na URL.
    """
    user_id = _resolve_dashboard_user_id(request)
    if not user_id:
        raise HTTPException(status_code=401, detail="Token inválido")
    _raise_if_account_scheduled_for_deletion(int(user_id))
    _no_store(response)

    # Anexa flag de onboarding pra o frontend decidir se mostra o prompt.
    from db import should_show_mfa_onboarding
    show_onboarding = await asyncio.to_thread(should_show_mfa_onboarding, int(user_id))
    return {"user_id": user_id, "show_mfa_onboarding": show_onboarding}


@app.get("/auth/dashboard-profile")
@limiter.limit("60/minute")
async def auth_dashboard_profile(request: Request, response: Response):
    """Retorna dados mínimos da conta para a UI do dashboard."""
    user_id = _resolve_dashboard_user_id(request)
    _raise_if_account_scheduled_for_deletion(int(user_id))
    auth_user = await asyncio.to_thread(get_auth_user, int(user_id))
    _no_store(response)
    return {
        "user_id": user_id,
        "email": (auth_user or {}).get("email"),
        "display_name": (auth_user or {}).get("display_name"),
        "plan": (auth_user or {}).get("plan"),
        "whatsapp_linked": bool((auth_user or {}).get("whatsapp_verified_at")),
    }


@app.post("/auth/register")
@limiter.limit("3/hour")
async def auth_register(request: Request, body: RegisterBody):
    """
    Inicia o cadastro: valida os dados, gera código de 6 dígitos e envia por e-mail.
    A conta só é criada após confirmação via /auth/verify-email.
    """
    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from db import create_email_verification
    from core.services.email_service import send_verification_email

    await _check_auth_rate_limits("register", request, body.email)

    if len(body.password) < 8:
        raise HTTPException(status_code=400, detail="Senha deve ter pelo menos 8 caracteres.")

    name = (body.name or "").strip() or None
    if name is not None:
        if len(name) < 2:
            raise HTTPException(status_code=400, detail="O nome deve ter pelo menos 2 caracteres.")
        if len(name) > 50:
            raise HTTPException(status_code=400, detail="O nome deve ter no máximo 50 caracteres.")

    try:
        normalize_phone_e164(body.phone)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    try:
        code = create_email_verification(
            body.email, body.password, body.phone, display_name=name,
        )
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))

    sent = send_verification_email(body.email.strip().lower(), code)
    if not sent:
        raise HTTPException(status_code=500, detail="Não foi possível enviar o e-mail de verificação. Tente novamente.")

    return {"status": "verification_sent", "email": body.email.strip().lower()}


@app.post("/auth/verify-email")
@limiter.limit("10/minute")
async def auth_verify_email(request: Request, response: Response, body: VerifyEmailBody):
    """
    Confirma o código de verificação e cria a conta.
    Retorna JWT + link_code igual ao registro anterior.
    """
    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from db import confirm_email_verification

    try:
        result = confirm_email_verification(body.email, body.code)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    user_id    = result["user_id"]
    link_code  = result["link_code"]
    token, jti, refresh = _issue_session_token(user_id, body.email.strip().lower(), request)
    _set_auth_cookie(response, token)
    _set_refresh_cookie(response, refresh)
    _set_dashboard_cookie(response, int(user_id), jti=jti)

    wa_link = _build_whatsapp_onboarding_link(user_id)

    return {
        "user_id": user_id,
        "email": body.email.strip().lower(),
        "link_code": link_code,
        "whatsapp_link": wa_link,
        "dashboard_url": _post_login_url(),
    }


@app.post("/auth/login")
@limiter.limit("5/minute")
async def auth_login(request: Request, response: Response, body: LoginBody):
    """Login via email+senha. Retorna JWT + link_code novo para vincular o bot."""
    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from db import login_auth_user, create_link_code, find_user_id_by_email, email_has_password

    await _check_auth_rate_limits("login", request, body.email)

    # Conta criada apenas via Google (sem password_hash) → orienta usar o botão correto
    existing_user_id = await asyncio.to_thread(find_user_id_by_email, body.email)
    if existing_user_id and not await asyncio.to_thread(email_has_password, body.email):
        # Mensagem genérica pra não diferenciar "conta existe via Google" de
        # "email não cadastrado" — evita enumeração de e-mails. O motivo real
        # fica no audit log (failure_reason="google_only_account").
        await log_auth_login_event(
            body.email,
            False,
            ip_address=get_remote_address(request),
            user_agent=request.headers.get("user-agent"),
            failure_reason="google_only_account",
        )
        raise HTTPException(status_code=401, detail="E-mail ou senha incorretos.")

    result = login_auth_user(body.email, body.password)
    if not result:
        await log_auth_login_event(
            body.email,
            False,
            ip_address=get_remote_address(request),
            user_agent=request.headers.get("user-agent"),
            failure_reason="invalid_credentials",
        )
        raise HTTPException(status_code=401, detail="E-mail ou senha incorretos.")

    user_id    = result["user_id"]
    _raise_if_account_scheduled_for_deletion(user_id)

    # Se o usuario tem MFA ativado, emite challenge token e retorna 200 com
    # mfa_required=true (sem cookies de sessao). O cliente deve chamar
    # /auth/mfa/verify-login para completar o login.
    from db import get_mfa_status, mfa_create_login_challenge
    mfa_status = await asyncio.to_thread(get_mfa_status, user_id)
    if mfa_status.get("enabled"):
        challenge = await asyncio.to_thread(mfa_create_login_challenge, user_id)
        await log_auth_login_event(
            result["email"],
            True,
            user_id=user_id,
            ip_address=get_remote_address(request),
            user_agent=request.headers.get("user-agent"),
            failure_reason="mfa_pending",
        )
        return {
            "mfa_required": True,
            "mfa_challenge": challenge,
            "email": result["email"],
        }

    link_code  = create_link_code(user_id, minutes_valid=15)
    token, jti, refresh = _issue_session_token(user_id, result["email"], request)
    _set_auth_cookie(response, token)
    _set_refresh_cookie(response, refresh)
    _set_dashboard_cookie(response, int(user_id), jti=jti)

    # Fire ANTES do log_auth_login_event: senao o IP atual ja vira "conhecido".
    await asyncio.to_thread(maybe_record_login_from_new_ip, user_id, request=request)

    await log_auth_login_event(
        result["email"],
        True,
        user_id=user_id,
        ip_address=get_remote_address(request),
        user_agent=request.headers.get("user-agent"),
    )

    wa_link = _build_whatsapp_onboarding_link(user_id)

    return {
        "user_id": user_id,
        "email": result["email"],
        "plan": result["plan"],
        "link_code": link_code,
        "whatsapp_link": wa_link,
        "dashboard_url": _post_login_url(),
    }


@app.post("/auth/logout")
@limiter.limit("30/minute")
async def auth_logout(request: Request, response: Response):
    # Revoga a sessao corrente (se houver jti no JWT) antes de limpar cookies.
    # Idempotente: token expirado / sem jti / sessao ja revogada = no-op.
    token = _get_auth_token_from_request(request, None)
    if token:
        payload = _decode_jwt(token)
        if payload and payload.get("type") == "auth":
            jti = payload.get("jti")
            user_id_raw = payload.get("sub")
            if jti and user_id_raw:
                try:
                    await asyncio.to_thread(revoke_session, int(user_id_raw), jti)
                    # Revoga também TODOS os refresh tokens dessa sessão
                    from core.refresh_tokens import revoke_session_refresh_tokens
                    await asyncio.to_thread(revoke_session_refresh_tokens, jti)
                except Exception:
                    # Cookies são limpos mesmo assim, mas a sessão segue válida
                    # no servidor — precisa aparecer no log.
                    logging.getLogger(__name__).warning(
                        "logout: falha ao revogar sessão/refresh tokens (user %s, jti %s)",
                        user_id_raw, jti, exc_info=True,
                    )

    # Revoga também o refresh_token específico do cookie (caso a sessão já
    # não bata — defesa em profundidade).
    refresh_in_cookie = (request.cookies.get(REFRESH_COOKIE_NAME) or "").strip()
    if refresh_in_cookie:
        try:
            from core.refresh_tokens import revoke_refresh_token
            await asyncio.to_thread(revoke_refresh_token, refresh_in_cookie)
        except Exception:
            logging.getLogger(__name__).warning(
                "logout: falha ao revogar refresh token do cookie", exc_info=True,
            )

    _clear_session_cookies(response)
    response.headers["Clear-Site-Data"] = '"cookies", "storage"'
    _no_store(response)
    return {"ok": True}


@app.post("/auth/refresh")
@limiter.limit("60/minute")
async def auth_refresh(request: Request, response: Response):
    """Renova o access token usando o refresh_token do cookie.

    Fluxo:
      1. Lê refresh_token do cookie (path=/auth/refresh).
      2. Rotaciona: marca antigo como usado, emite novo refresh com mesmo session_jti.
      3. Emite novo access token (15min) + dashboard_token.
      4. Atualiza auth_sessions.last_seen_at.

    Falhas (token revogado, expirado, replay, idle, sessão revogada): retorna 401
    e limpa cookies. Frontend deve mandar pro login.
    """
    refresh_in_cookie = (request.cookies.get(REFRESH_COOKIE_NAME) or "").strip()
    if not refresh_in_cookie:
        raise HTTPException(status_code=401, detail="missing_refresh_token")

    from core.refresh_tokens import consume_refresh_token
    ip = get_remote_address(request) or None
    ua = request.headers.get("user-agent") or None
    result = await asyncio.to_thread(
        consume_refresh_token, refresh_in_cookie, ip=ip, user_agent=ua,
    )
    if not result:
        # Limpa cookies — qualquer motivo de falha vira deslogue.
        _clear_session_cookies(response)
        raise HTTPException(status_code=401, detail="invalid_refresh_token")

    user_id = int(result["user_id"])
    session_jti = result["session_jti"]
    new_refresh = result["new_refresh_token"]

    # Recupera email pra montar o access JWT
    try:
        from db import get_auth_user
        u = await asyncio.to_thread(get_auth_user, user_id)
        email = (u or {}).get("email") or ""
    except Exception:
        email = ""

    access = _make_jwt(user_id, email, jti=session_jti)
    _set_auth_cookie(response, access)
    _set_refresh_cookie(response, new_refresh)
    # Renova dashboard_token também (mesmo jti)
    _set_dashboard_cookie(response, user_id, jti=session_jti)
    _no_store(response)
    return {"ok": True}


@app.post("/auth/forgot-password")
@limiter.limit("3/hour")
async def auth_forgot_password(request: Request, body: EmailBody):
    """
    Solicita recuperação de senha. Envia e-mail com link se o e-mail existir.
    Sempre retorna 200 para não revelar se o e-mail está cadastrado.
    """
    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from db import create_password_reset_token
    from core.services.email_service import send_password_reset_email

    await _check_auth_rate_limits("forgot-password", request, body.email)

    token = create_password_reset_token(body.email)
    if token:
        reset_url = f"{DASHBOARD_URL}/reset-password#token={token}"
        send_password_reset_email(body.email.strip().lower(), reset_url)

    # sempre retorna 200 — não revela se o e-mail existe ou não
    return {"message": "Se este e-mail estiver cadastrado, você receberá as instruções em breve."}


@app.post("/auth/reset-password")
@limiter.limit("5/minute")
async def auth_reset_password(request: Request, body: ResetPasswordBody):
    """
    Redefine a senha usando o token recebido por e-mail.
    """
    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from db import consume_password_reset_token

    if len(body.new_password) < 8:
        raise HTTPException(status_code=400, detail="Senha deve ter pelo menos 8 caracteres.")

    user_id = consume_password_reset_token(body.token, body.new_password)
    if not user_id:
        raise HTTPException(status_code=400, detail="Link inválido ou expirado. Solicite um novo.")

    await asyncio.to_thread(
        record_audit_event,
        user_id,
        AuditEvent.PASSWORD_RESET_COMPLETED,
        request=request,
    )

    return {"message": "Senha redefinida com sucesso! Faça login com sua nova senha."}


@app.post("/auth/link-code")
async def auth_new_link_code(user_id: int = Depends(_get_current_user)):
    """Gera um novo link_code para o usuário autenticado vincular uma nova plataforma."""
    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from db import create_link_code

    link_code = create_link_code(user_id, minutes_valid=15)
    wa_link = _build_whatsapp_onboarding_link(user_id)

    return {
        "link_code": link_code,
        "whatsapp_link": wa_link,
        "expires_in_minutes": 15,
    }


@app.get("/auth/me")
async def auth_me(user_id: int = Depends(_get_current_user)):
    """Retorna dados do usuário autenticado."""
    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from db import get_auth_user, should_show_mfa_onboarding, get_mfa_status

    user = get_auth_user(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")

    show_onboarding = await asyncio.to_thread(should_show_mfa_onboarding, user_id)
    mfa = await asyncio.to_thread(get_mfa_status, user_id)
    return {
        "user_id": user_id,
        **dict(user),
        "show_mfa_onboarding": show_onboarding,
        "mfa_enabled": bool(mfa.get("enabled")),
    }


# ── MFA (TOTP) ───────────────────────────────────────────────────────────────

class MFASetupBody(BaseModel):
    password: str


class MFAEnableBody(BaseModel):
    code: str


class MFADisableBody(BaseModel):
    password: str
    code: str | None = None


class MFAVerifyLoginBody(BaseModel):
    challenge: str
    code: str
    use_backup: bool = False


def _qr_data_url(uri: str) -> str:
    """Gera SVG do QR code como data URL (sem dependencia de PIL)."""
    import qrcode
    import qrcode.image.svg
    factory = qrcode.image.svg.SvgPathImage
    img = qrcode.make(uri, image_factory=factory, box_size=10, border=2)
    import io
    buf = io.BytesIO()
    img.save(buf)
    svg = buf.getvalue().decode()
    encoded = base64.b64encode(svg.encode()).decode()
    return f"data:image/svg+xml;base64,{encoded}"


@app.get("/auth/mfa/status")
async def auth_mfa_status(user_id: int = Depends(_get_current_user)):
    """Retorna status atual do MFA do usuario logado."""
    from db import get_mfa_status
    return await asyncio.to_thread(get_mfa_status, user_id)


@app.post("/auth/mfa/onboarding-seen")
@limiter.limit("30/hour")
async def auth_mfa_onboarding_seen(request: Request, user_id: int = Depends(_get_current_user)):
    """Grava que o usuario viu a tela de onboarding (independente da escolha).
    Idempotente: chamadas subsequentes nao reescrevem o timestamp."""
    from db import mark_mfa_onboarding_shown
    await asyncio.to_thread(mark_mfa_onboarding_shown, user_id)
    return {"ok": True}


@app.post("/auth/mfa/setup")
@limiter.limit("5/hour")
async def auth_mfa_setup(request: Request, body: MFASetupBody, user_id: int = Depends(_get_current_user)):
    """Inicia setup do MFA: pede senha, gera secret, retorna QR + URI."""
    _block_setup_if_unsupported_user(user_id)

    password_ok = await asyncio.to_thread(verify_user_password, user_id, body.password)
    if not password_ok:
        raise HTTPException(status_code=401, detail="Senha incorreta.")

    from db import get_auth_user, mfa_setup_secret
    user = await asyncio.to_thread(get_auth_user, user_id)
    if not user or not user.get("email"):
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")

    try:
        result = await asyncio.to_thread(mfa_setup_secret, user_id, user["email"])
    except ValueError as exc:
        if str(exc) == "MFA_ALREADY_ENABLED":
            raise HTTPException(status_code=409, detail="MFA já está ativado.") from exc
        raise

    return {
        "secret": result["secret"],
        "uri": result["uri"],
        "qr_code": _qr_data_url(result["uri"]),
    }


@app.post("/auth/mfa/enable")
@limiter.limit("10/hour")
async def auth_mfa_enable(request: Request, body: MFAEnableBody, user_id: int = Depends(_get_current_user)):
    """Confirma o primeiro codigo TOTP e ativa MFA. Retorna codigos de backup."""
    from db import mfa_verify_and_enable
    try:
        backup_codes = await asyncio.to_thread(mfa_verify_and_enable, user_id, body.code)
    except ValueError as exc:
        msg = str(exc)
        if msg == "MFA_NOT_INITIALIZED":
            raise HTTPException(status_code=400, detail="Inicie o setup primeiro.") from exc
        if msg == "MFA_ALREADY_ENABLED":
            raise HTTPException(status_code=409, detail="MFA já está ativado.") from exc
        if msg == "MFA_CODE_INVALID":
            raise HTTPException(status_code=400, detail="Código inválido. Tente novamente.") from exc
        raise

    await asyncio.to_thread(
        record_audit_event,
        user_id,
        AuditEvent.MFA_ENABLED,
        request=request,
    )

    return {"ok": True, "backup_codes": backup_codes}


@app.post("/auth/mfa/disable")
@limiter.limit("5/hour")
async def auth_mfa_disable(request: Request, body: MFADisableBody, user_id: int = Depends(_get_current_user)):
    """Desativa MFA (pede senha + codigo TOTP atual ou backup code)."""
    password_ok = await asyncio.to_thread(verify_user_password, user_id, body.password)
    if not password_ok:
        raise HTTPException(status_code=401, detail="Senha incorreta.")

    from db import get_mfa_status, mfa_verify_totp, mfa_consume_backup_code, disable_mfa
    status = await asyncio.to_thread(get_mfa_status, user_id)
    if not status.get("enabled"):
        # Idempotente: ja desativado.
        return {"ok": True}

    # Exige codigo TOTP ou backup code para desligar (defesa contra sequestro
    # de sessao: roubar cookie nao basta).
    code = (body.code or "").strip()
    if not code:
        raise HTTPException(status_code=400, detail="Código de autenticação é obrigatório.")
    ok = await asyncio.to_thread(mfa_verify_totp, user_id, code)
    if not ok:
        ok = await asyncio.to_thread(mfa_consume_backup_code, user_id, code)
    if not ok:
        raise HTTPException(status_code=400, detail="Código inválido.")

    await asyncio.to_thread(disable_mfa, user_id)

    await asyncio.to_thread(
        record_audit_event,
        user_id,
        AuditEvent.MFA_DISABLED,
        request=request,
    )

    return {"ok": True}


@app.post("/auth/mfa/regenerate-backup-codes")
@limiter.limit("3/hour")
async def auth_mfa_regenerate(request: Request, body: MFADisableBody, user_id: int = Depends(_get_current_user)):
    """Gera novos backup codes (invalida os antigos). Pede senha + TOTP."""
    password_ok = await asyncio.to_thread(verify_user_password, user_id, body.password)
    if not password_ok:
        raise HTTPException(status_code=401, detail="Senha incorreta.")

    code = (body.code or "").strip()
    if not code:
        raise HTTPException(status_code=400, detail="Código de autenticação é obrigatório.")

    from db import mfa_verify_totp, mfa_regenerate_backup_codes
    ok = await asyncio.to_thread(mfa_verify_totp, user_id, code)
    if not ok:
        raise HTTPException(status_code=400, detail="Código TOTP inválido.")

    try:
        codes = await asyncio.to_thread(mfa_regenerate_backup_codes, user_id)
    except ValueError as exc:
        if str(exc) == "MFA_NOT_ENABLED":
            raise HTTPException(status_code=400, detail="MFA não está ativado.") from exc
        raise

    await asyncio.to_thread(
        record_audit_event,
        user_id,
        AuditEvent.MFA_BACKUP_CODES_REGENERATED,
        request=request,
    )

    return {"backup_codes": codes}


@app.post("/auth/mfa/verify-login")
# 5/min é apertado mas não atrapalha usuário legítimo (que erra 1-2 vezes).
# TOTP tem só 10^6 valores — 10/min seria brute-force viável em ~16 dias.
@limiter.limit("5/minute")
async def auth_mfa_verify_login(request: Request, response: Response, body: MFAVerifyLoginBody):
    """Completa o login apos /auth/login retornar mfa_required=true.

    O cliente envia (challenge, code). Se OK, emite JWT auth + cookie.
    """
    from db import (
        mfa_consume_login_challenge,
        mfa_verify_totp,
        mfa_consume_backup_code,
        get_auth_user,
        create_link_code,
    )

    user_id = await asyncio.to_thread(mfa_consume_login_challenge, body.challenge)
    if not user_id:
        raise HTTPException(status_code=400, detail="Sessão MFA expirada. Faça login novamente.")

    code = (body.code or "").strip()
    verified = False
    if body.use_backup:
        verified = await asyncio.to_thread(mfa_consume_backup_code, user_id, code)
    else:
        verified = await asyncio.to_thread(mfa_verify_totp, user_id, code)

    if not verified:
        raise HTTPException(status_code=400, detail="Código inválido.")

    user = await asyncio.to_thread(get_auth_user, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")

    link_code = await asyncio.to_thread(create_link_code, user_id, 15)
    token, jti, refresh = _issue_session_token(user_id, user["email"], request)
    _set_auth_cookie(response, token)
    _set_refresh_cookie(response, refresh)
    _set_dashboard_cookie(response, int(user_id), jti=jti)

    # New-IP audit ANTES do log_auth_login_event de sucesso. Rows de
    # mfa_pending (criadas em /auth/login) sao filtradas pelo helper.
    await asyncio.to_thread(maybe_record_login_from_new_ip, user_id, request=request)

    await log_auth_login_event(
        user["email"],
        True,
        user_id=user_id,
        ip_address=get_remote_address(request),
        user_agent=request.headers.get("user-agent"),
    )

    wa_link = _build_whatsapp_onboarding_link(user_id)
    return {
        "user_id": user_id,
        "email": user["email"],
        "plan": user.get("plan", "free"),
        "link_code": link_code,
        "whatsapp_link": wa_link,
        "dashboard_url": _post_login_url(),
    }


def _block_setup_if_unsupported_user(user_id: int) -> None:
    """Reserva pra futuras restricoes (ex: nao permitir MFA em conta deletada).
    Por enquanto so respeita schedule de delecao."""
    _raise_if_account_scheduled_for_deletion(user_id)


@app.post("/auth/account/export")
@limiter.limit("3/hour")
async def auth_account_export_request(request: Request, body: DataExportBody):
    """
    Solicita exportação completa de dados.

    Camadas de proteção (LGPD + segurança):
      - Rate limit por IP (3/hora) — evita abuso por automação.
      - Re-autenticação por senha — sessão roubada (cookie/token) não basta.
      - Cooldown por usuário — máximo 1 link válido pendente por hora.
      - Não devolve o ZIP nesta resposta. Em vez disso, gera um token de uso
        único (15 min) e envia por e-mail um link de download.
      - Audit log + e-mail de notificação confirmam quem solicitou.
    """
    user_id = _resolve_dashboard_user_id(request)
    _raise_if_account_scheduled_for_deletion(user_id)

    client_ip = get_remote_address(request)
    user_agent = (request.headers.get("user-agent") or "").strip() or None

    # 1) Re-auth por senha
    password_ok = await asyncio.to_thread(verify_user_password, user_id, body.password)
    if not password_ok:
        await log_system_event(
            "warning",
            "data_export_password_failed",
            f"Senha incorreta ao solicitar exportação para user_id={user_id}",
            source="auth_account_export",
            user_id=user_id,
            details={"ip": client_ip, "user_agent": user_agent},
        )
        raise HTTPException(status_code=401, detail="Senha incorreta.")

    # 2) Cooldown por usuário (1 link pendente por hora)
    has_pending = await asyncio.to_thread(has_recent_export_request, user_id, 60)
    if has_pending:
        raise HTTPException(
            status_code=429,
            detail=(
                "Já existe um link de exportação pendente enviado para o seu e-mail. "
                "Aguarde alguns minutos antes de solicitar novamente."
            ),
            headers={"Retry-After": "900"},
        )

    # 3) Email do usuário (obrigatório para entregar o link)
    email = await asyncio.to_thread(get_user_email, user_id)
    if not email:
        raise HTTPException(
            status_code=400,
            detail="E-mail não cadastrado para esta conta. Vincule um e-mail antes de exportar seus dados.",
        )

    # 4) Gera token + persiste request
    token, expires_at = await asyncio.to_thread(
        create_data_export_token,
        user_id,
        minutes_valid=15,
        request_ip=client_ip,
        request_user_agent=user_agent,
        delivered_to_email=email,
    )

    download_url = _dashboard_url(f"/auth/account/export/download/{token}")

    # 5) Envia e-mail (em thread pra não bloquear)
    from core.services.email_service import send_data_export_link_email  # noqa: PLC0415
    sent = await asyncio.to_thread(
        send_data_export_link_email,
        email,
        download_url,
        15,
        client_ip,
        user_agent,
    )

    # 6) Audit log
    await log_system_event(
        "info" if sent else "error",
        "data_export_requested" if sent else "data_export_email_failed",
        f"Solicitação de exportação para user_id={user_id} ({'enviada' if sent else 'falha no envio'})",
        source="auth_account_export",
        user_id=user_id,
        details={"ip": client_ip, "user_agent": user_agent, "email": email},
    )

    if not sent:
        raise HTTPException(
            status_code=500,
            detail="Não foi possível enviar o e-mail com o link de exportação. Tente novamente em alguns minutos.",
        )

    return {
        "status": "email_sent",
        "message": "Enviamos um link de download para o seu e-mail. O link expira em 15 minutos e só pode ser usado uma vez.",
        "expires_in_minutes": 15,
        "expires_at": expires_at.isoformat(),
    }


@app.get("/auth/account/export/download/{token}")
@limiter.limit("10/hour")
async def auth_account_export_download(request: Request, token: str):
    """
    Consome o token de uso único e devolve o ZIP com os dados do usuário.

    Não exige re-autenticação aqui — a posse do token (entregue por e-mail
    de uma solicitação válida com senha) já é a credencial. O token é
    invalidado de forma atômica antes de qualquer trabalho pesado, então
    cliques duplicados ou tentativas concorrentes não baixam duas vezes.
    """
    user_id = await asyncio.to_thread(consume_data_export_token, token)
    if not user_id:
        await log_system_event(
            "warning",
            "data_export_token_invalid",
            "Tentativa de download com token inválido, expirado ou já usado.",
            source="auth_account_export_download",
            details={"ip": get_remote_address(request)},
        )
        raise HTTPException(
            status_code=410,
            detail="Link de exportação inválido, expirado ou já utilizado. Solicite uma nova exportação.",
        )

    _raise_if_account_scheduled_for_deletion(user_id)

    client_ip = get_remote_address(request)
    user_agent = (request.headers.get("user-agent") or "").strip() or None

    content = await asyncio.to_thread(build_user_export_zip, user_id)

    completed_at_dt = datetime.now(timezone.utc)
    filename = f"pigbank_dados_usuario_{user_id}_{completed_at_dt:%Y%m%d}.zip"

    # Notifica o dono por e-mail (auditoria) e loga o evento — em background
    # pra não atrasar o stream do ZIP.
    async def _notify_completed():
        try:
            email = await asyncio.to_thread(get_user_email, user_id)
            if email:
                from core.services.email_service import send_data_export_completed_email  # noqa: PLC0415
                await asyncio.to_thread(
                    send_data_export_completed_email,
                    email,
                    completed_at_dt.strftime("%d/%m/%Y %H:%M UTC"),
                    client_ip,
                    user_agent,
                )
        except Exception as exc:
            logging.getLogger(__name__).warning(
                "Falha ao enviar e-mail de confirmação de export para user_id=%s: %s",
                user_id, exc,
            )
        await log_system_event(
            "info",
            "data_export_completed",
            f"Exportação baixada por user_id={user_id}",
            source="auth_account_export_download",
            user_id=user_id,
            details={"ip": client_ip, "user_agent": user_agent},
        )

    asyncio.create_task(_notify_completed())

    return StreamingResponse(
        iter([content]),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )


@app.delete("/auth/account")
@limiter.limit("5/minute")
async def auth_delete_account(request: Request, response: Response, body: DeleteAccountBody):
    """Agenda a exclusão definitiva da conta após o período de carência."""
    user_id = _resolve_dashboard_user_id(request)
    auth_user = await asyncio.to_thread(get_auth_user, user_id)
    email = (auth_user or {}).get("email")
    try:
        result = await asyncio.to_thread(schedule_account_deletion, user_id, body.password, 7)
    except PermissionError as exc:
        raise HTTPException(status_code=401, detail=str(exc)) from exc
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if email:
        from core.services.email_service import send_account_deletion_scheduled_email

        scheduled = result.get("deletion_scheduled_for")
        scheduled_txt = scheduled.isoformat() if hasattr(scheduled, "isoformat") else str(scheduled)
        await asyncio.to_thread(send_account_deletion_scheduled_email, email.strip().lower(), scheduled_txt)

    _clear_session_cookies(response)
    _no_store(response)
    return json.loads(jdump({
        "ok": True,
        **result,
        "message": "Conta agendada para exclusão. Seus dados serão removidos definitivamente após o período de carência.",
    }))


@app.post("/auth/dashboard-token")
async def auth_dashboard_token(response: Response, request: Request, user_id: int = Depends(_get_current_user)):
    """Troca o token de login por um cookie HttpOnly de acesso ao dashboard."""
    jti = getattr(request.state, "session_jti", None)
    _set_dashboard_cookie(response, int(user_id), jti=jti)
    auth_payload = getattr(request.state, "auth_payload", {}) or {}
    return {
        "email": auth_payload.get("email"),
        "dashboard_url": _post_login_url(),
        "expires_in": int(DASHBOARD_SESSION_HOURS * 3600),
    }


@app.post("/auth/dashboard-link")
async def auth_dashboard_link(response: Response, request: Request, body: DashboardLinkBody, user_id: int = Depends(_get_current_user)):
    """
    Consome um magic link do dashboard e libera acesso apenas
    se o usuário logado for o dono desse link.
    """
    from db import consume_dashboard_session

    target_user_id = consume_dashboard_session(body.code.strip())
    if not target_user_id:
        raise HTTPException(status_code=401, detail="Link de dashboard inválido ou expirado.")
    if int(target_user_id) != int(user_id):
        raise HTTPException(status_code=403, detail="Este link pertence a outra conta.")

    jti = getattr(request.state, "session_jti", None)
    _set_dashboard_cookie(response, int(user_id), jti=jti)
    auth_payload = getattr(request.state, "auth_payload", {}) or {}
    return {
        "email": auth_payload.get("email"),
        "dashboard_url": _post_login_url(),
        "expires_in": int(DASHBOARD_SESSION_HOURS * 3600),
    }


# ─── Login social (Google OAuth) ─────────────────────────────────────────────

GOOGLE_OAUTH_STATE_COOKIE = "google_oauth_state"
GOOGLE_OAUTH_STATE_MAX_AGE = 600  # 10 minutos


class GoogleSignupCompleteBody(BaseModel):
    token: str
    name: str
    phone: str
    accepted_terms: bool = False


def _google_redirect_to_landing(message: str) -> RedirectResponse:
    """Volta pra landing com flag de erro pra UI mostrar."""
    qs = urllib.parse.urlencode({"google_error": message})
    return RedirectResponse(url=f"/?{qs}", status_code=302)


@app.get("/auth/google/start")
@limiter.limit("10/minute")
async def auth_google_start(request: Request):
    """Gera state, salva em cookie short-lived e redireciona pro Google."""
    from core.services.google_oauth import (
        GoogleOAuthError,
        build_authorization_url,
        is_configured,
    )

    if not is_configured():
        raise HTTPException(
            status_code=503,
            detail="Login com Google ainda não está configurado neste ambiente.",
        )

    state = secrets.token_urlsafe(32)
    try:
        url = build_authorization_url(state)
    except GoogleOAuthError as exc:
        raise HTTPException(status_code=503, detail=str(exc))

    response = RedirectResponse(url=url, status_code=302)
    # SameSite=lax: o cookie precisa retornar quando o Google fizer redirect cross-site.
    response.set_cookie(
        GOOGLE_OAUTH_STATE_COOKIE,
        state,
        httponly=True,
        secure=COOKIE_SECURE,
        samesite="lax",
        max_age=GOOGLE_OAUTH_STATE_MAX_AGE,
        path="/auth/google",
    )
    return response


@app.get("/auth/google/callback")
async def auth_google_callback(
    request: Request,
    response: Response,
    code: str | None = None,
    state: str | None = None,
    error: str | None = None,
):
    """
    Recebe o redirect do Google. Valida state, troca code por id_token,
    decide login / vincular / criar pendente, e redireciona o usuário.
    """
    import logging as _logging
    import traceback as _traceback

    from core.services.google_oauth import (
        GoogleOAuthError,
        exchange_code_for_tokens,
        verify_id_token,
    )
    from db import (
        find_user_by_google_sub,
        find_user_id_by_email,
        link_google_identity,
        create_pending_google_signup,
    )

    _log = _logging.getLogger("auth.google")

    cookie_state = request.cookies.get(GOOGLE_OAUTH_STATE_COOKIE) or ""

    if error:
        return _google_redirect_to_landing(f"Login com Google cancelado: {error}")

    if not code or not state or not cookie_state or not secrets.compare_digest(state, cookie_state):
        return _google_redirect_to_landing("Sessão de login expirou. Tente novamente.")

    try:
        try:
            tokens = await exchange_code_for_tokens(code)
            claims = verify_id_token(tokens["id_token"])
        except GoogleOAuthError as exc:
            return _google_redirect_to_landing(str(exc))

        sub = claims["sub"]
        email = (claims.get("email") or "").strip().lower()
        email_verified = bool(claims.get("email_verified"))
        name_hint = claims.get("name") or claims.get("given_name") or None

        if not email or not email_verified:
            return _google_redirect_to_landing(
                "Sua conta Google não tem e-mail verificado. Verifique no Google e tente novamente."
            )

        # 1) Já existe vínculo? → login direto
        user_id = await asyncio.to_thread(find_user_by_google_sub, sub)

        # 2) Não existe vínculo, mas existe conta com este email? → auto-link
        if not user_id:
            existing = await asyncio.to_thread(find_user_id_by_email, email)
            if existing:
                await asyncio.to_thread(link_google_identity, existing, sub, email)
                user_id = existing

        # 3) Conta totalmente nova → cria pendente e manda pra /onboarding
        if not user_id:
            token = await asyncio.to_thread(create_pending_google_signup, sub, email, name_hint)
            signup_response = RedirectResponse(url=f"/onboarding?token={token}", status_code=302)
            signup_response.delete_cookie(GOOGLE_OAUTH_STATE_COOKIE, path="/auth/google")
            return signup_response

        # Usuário existente: bloqueia se conta agendada para deletar
        try:
            _raise_if_account_scheduled_for_deletion(user_id)
        except HTTPException as exc:
            return _google_redirect_to_landing(exc.detail)

        # Login bem-sucedido → cookies + redirect pra home
        jwt_token, jti, refresh = _issue_session_token(user_id, email, request)
        success_response = RedirectResponse(url=_post_login_url(), status_code=302)
        success_response.delete_cookie(GOOGLE_OAUTH_STATE_COOKIE, path="/auth/google")
        _set_auth_cookie(success_response, jwt_token)
        _set_refresh_cookie(success_response, refresh)
        _set_dashboard_cookie(success_response, int(user_id), jti=jti)

        await asyncio.to_thread(maybe_record_login_from_new_ip, user_id, request=request)

        await log_auth_login_event(
            email,
            True,
            user_id=user_id,
            ip_address=get_remote_address(request),
            user_agent=request.headers.get("user-agent"),
        )

        return success_response

    except Exception as exc:
        _log.error("Falha inesperada no /auth/google/callback: %s\n%s",
                   exc, _traceback.format_exc())
        return _google_redirect_to_landing(
            f"Falha inesperada no login Google ({type(exc).__name__}). Veja o terminal do servidor."
        )


@app.get("/auth/google/pending/{token}")
@limiter.limit("30/minute")
async def auth_google_pending(request: Request, response: Response, token: str):
    """Devolve dados do pré-cadastro (email + nome sugerido) pra preencher o form."""
    from db import get_pending_google_signup

    pending = await asyncio.to_thread(get_pending_google_signup, token)
    if not pending:
        raise HTTPException(status_code=404, detail="Cadastro expirado ou inválido.")

    _no_store(response)
    return {
        "email": pending["email"],
        "name_hint": pending["name_hint"] or "",
    }


@app.post("/auth/google/complete-signup")
@limiter.limit("10/minute")
async def auth_google_complete_signup(
    request: Request,
    response: Response,
    body: GoogleSignupCompleteBody,
):
    """Finaliza o cadastro Google: cria conta com nome + telefone."""
    from db import consume_pending_google_signup

    if not body.accepted_terms:
        raise HTTPException(
            status_code=400,
            detail="É necessário aceitar a Política de Privacidade.",
        )

    try:
        result = await asyncio.to_thread(
            consume_pending_google_signup,
            body.token, body.name, body.phone,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    user_id = int(result["user_id"])
    email = result["email"]

    jwt_token, jti, refresh = _issue_session_token(user_id, email, request)
    _set_auth_cookie(response, jwt_token)
    _set_refresh_cookie(response, refresh)
    _set_dashboard_cookie(response, user_id, jti=jti)

    await log_auth_login_event(
        email,
        True,
        user_id=user_id,
        ip_address=get_remote_address(request),
        user_agent=request.headers.get("user-agent"),
    )

    wa_link = _build_whatsapp_onboarding_link(user_id)

    return {
        "user_id": user_id,
        "email": email,
        "link_code": result["link_code"],
        "whatsapp_link": wa_link,
        "dashboard_url": _post_login_url(),
    }


# ─── Billing (Stripe) ────────────────────────────────────────────────────────

class CreateCheckoutBody(BaseModel):
    interval: str = "monthly"  # "monthly" | "annual"


def _resolve_pro_price_id(interval: str) -> str:
    """Mapeia interval -> price ID configurado nas env vars.

    Mensal aceita fallback pro `STRIPE_PRICE_ID_PRO` legado pra não quebrar
    deploys que ainda não migraram. Anual exige a env var nova.
    """
    if interval == "monthly":
        return STRIPE_PRICE_ID_PRO_MENSAL or STRIPE_PRICE_ID_PRO
    if interval == "annual":
        return STRIPE_PRICE_ID_PRO_ANUAL
    return ""


@app.post("/billing/create-checkout")
async def billing_create_checkout(
    payload: CreateCheckoutBody | None = None,
    user_id: int = Depends(_get_current_user),
):
    """
    Cria uma sessão de checkout no Stripe para upgrade para o plano Pro.
    Body opcional: {"interval": "monthly" | "annual"} (default monthly).
    Requer: STRIPE_SECRET_KEY + price ID do interval escolhido.
    """
    interval = (payload.interval if payload else "monthly")
    if interval not in ("monthly", "annual"):
        raise HTTPException(status_code=400, detail="interval inválido (use 'monthly' ou 'annual').")

    price_id = _resolve_pro_price_id(interval)
    if not STRIPE_SECRET_KEY or not price_id:
        raise HTTPException(status_code=503, detail="Pagamentos ainda não configurados.")

    import stripe
    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from db import get_auth_user, set_stripe_customer

    stripe.api_key = STRIPE_SECRET_KEY

    user = get_auth_user(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")

    # Recupera ou cria o customer no Stripe
    customer_id = user.get("stripe_customer_id")
    if not customer_id:
        customer = stripe.Customer.create(
            email=user["email"],
            metadata={"finbot_user_id": str(user_id)},
            address={"country": "BR"},
            preferred_locales=["pt-BR"],
        )
        customer_id = customer.id
        set_stripe_customer(user_id, customer_id)

    session = stripe.checkout.Session.create(
        customer=customer_id,
        payment_method_types=["card"],
        line_items=[{"price": price_id, "quantity": 1}],
        mode="subscription",
        locale="pt-BR",
        success_url=f"{DASHBOARD_URL}/home?upgrade=success",
        cancel_url=f"{DASHBOARD_URL}/home?upgrade=cancelled",
        metadata={"finbot_user_id": str(user_id), "interval": interval},
        subscription_data={
            "trial_period_days": 7,
            "metadata": {"finbot_user_id": str(user_id), "interval": interval},
        },
    )

    return {"checkout_url": session.url, "interval": interval}


@app.post("/billing/webhook")
async def billing_webhook(request: Request):
    """
    Recebe eventos do Stripe (checkout.session.completed, customer.subscription.*).
    Atualiza o plano do usuário no banco.
    """
    if not STRIPE_SECRET_KEY:
        raise HTTPException(status_code=503, detail="Stripe não configurado.")

    import stripe
    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from db import update_user_plan, get_user_by_stripe_customer, set_payment_status

    stripe.api_key = STRIPE_SECRET_KEY
    payload    = await request.body()
    sig_header = request.headers.get("stripe-signature", "")

    try:
        event = stripe.Webhook.construct_event(payload, sig_header, STRIPE_WEBHOOK_SECRET)
    except stripe.error.SignatureVerificationError:
        await log_system_event(
            "error",
            "billing_signature_invalid",
            "Webhook do Stripe rejeitado por assinatura invalida.",
            source="billing",
        )
        raise HTTPException(status_code=400, detail="Assinatura inválida.")

    await log_system_event(
        "info",
        "billing_webhook_received",
        f"Webhook do Stripe recebido: {event['type']}",
        source="billing",
        details={"event_type": event["type"]},
    )

    # Stripe SDK v8+ StripeObject nao herda de dict, entao .get nao existe.
    # Usar acesso por colchete + try/except (suporta tanto StripeObject quanto dict puro).
    def _g(obj, key, default=None):
        if obj is None:
            return default
        try:
            v = obj[key]
        except (KeyError, TypeError, AttributeError):
            return default
        return v if v is not None else default

    def _resolve_user(obj) -> int | None:
        metadata = _g(obj, "metadata", {})
        uid = _g(metadata, "finbot_user_id")
        if uid:
            return int(uid)
        cid = _g(obj, "customer")
        if cid:
            return get_user_by_stripe_customer(cid)
        return None

    def _subscription_period_end(sub):
        # API >= 2025-09 movido pra subscription.items.data[].current_period_end.
        ts = _g(sub, "current_period_end")
        if ts is None:
            items_obj = _g(sub, "items", {})
            data = _g(items_obj, "data", []) or []
            if data:
                ts = _g(data[0], "current_period_end")
        if ts is None:
            return None
        return datetime.fromtimestamp(ts, tz=timezone.utc)

    def _invoice_subscription_id(invoice) -> str | None:
        sub_id = _g(invoice, "subscription")
        if sub_id:
            return sub_id if isinstance(sub_id, str) else _g(sub_id, "id")
        # API >= 2025-09 movido pra invoice.parent.subscription_details.subscription.
        parent = _g(invoice, "parent", {})
        details = _g(parent, "subscription_details", {})
        ref = _g(details, "subscription")
        if ref is None:
            return None
        return ref if isinstance(ref, str) else _g(ref, "id")

    async def _user_email(uid: int) -> str:
        try:
            from db import get_auth_user as _gau
            data = await asyncio.to_thread(_gau, int(uid))
            return ((data or {}).get("email") or "").strip()
        except Exception:
            return ""

    async def _fire_email(uid: int, fn, *args):
        """Envia email transacional em background — falha silenciosa pra nao quebrar webhook."""
        try:
            email = await _user_email(uid)
            if not email:
                return
            await asyncio.to_thread(fn, email, *args, DASHBOARD_URL)
        except Exception as exc:
            print(f"[billing] email {fn.__name__} falhou user={uid}: {exc}")

    if event["type"] == "checkout.session.completed":
        session = event["data"]["object"]
        user_id = _resolve_user(session)
        sub_id  = _g(session, "subscription")
        # Trial 7d: subscription nasce status=trialing, sem invoice paga.
        # Promover ja agora pra user nao ficar Free durante o trial.
        if user_id and sub_id:
            sub = stripe.Subscription.retrieve(sub_id)
            expires_dt = _subscription_period_end(sub)
            sub_status = _g(sub, "status") or "trialing"
            update_user_plan(user_id, "pro", expires_dt)
            set_payment_status(user_id, sub_status)
            await log_system_event(
                "info",
                "billing_checkout_completed",
                f"Checkout concluido; plano pro ate {expires_dt.date() if expires_dt else 'sem data'}.",
                source="billing",
                user_id=user_id,
                details={
                    "plan": "pro",
                    "expires_at": expires_dt.isoformat() if expires_dt else None,
                    "status": sub_status,
                },
            )
            # Email de boas-vindas Pro (item 37)
            from core.services.email_service import send_pro_welcome_email
            await _fire_email(user_id, send_pro_welcome_email, expires_dt)
            # Notificação admin (Slack/Discord webhook)
            try:
                from core.services.admin_notify import notify_new_pro
                _admin_email = await _user_email(user_id)
                _line_items = _g(session, "line_items") or {}
                _interval = None  # interval (monthly/annual) vem do price; best-effort
                await asyncio.to_thread(
                    notify_new_pro,
                    user_id=user_id, email=_admin_email,
                    plan="pro", status=sub_status, expires_at=expires_dt,
                    interval=_interval,
                )
            except Exception as exc:
                print(f"[billing] admin notify falhou user={user_id}: {exc}")
        elif user_id:
            await log_system_event(
                "info",
                "billing_checkout_completed",
                "Checkout do Stripe concluido (sem subscription).",
                source="billing",
                user_id=user_id,
            )

    elif event["type"] in ("invoice.paid", "invoice.payment_succeeded"):
        invoice  = event["data"]["object"]
        user_id  = _resolve_user(invoice)
        sub_id   = _invoice_subscription_id(invoice)
        if user_id and sub_id:
            sub = stripe.Subscription.retrieve(sub_id)
            expires_dt = _subscription_period_end(sub)
            sub_status = _g(sub, "status") or "active"
            update_user_plan(user_id, "pro", expires_dt)
            set_payment_status(user_id, sub_status)
            print(f"[billing] user {user_id} → pro até {expires_dt.date() if expires_dt else 'sem data'}")
            await log_system_event(
                "info",
                "billing_plan_updated",
                "Plano do usuario atualizado para pro.",
                source="billing",
                user_id=user_id,
                details={"plan": "pro", "expires_at": expires_dt.isoformat() if expires_dt else None, "status": sub_status},
            )
            # Email de confirmacao de cobranca (item 39) — so quando valor > 0
            # (invoices do trial vem com amount_paid=0 e nao precisam de notificacao).
            amount_cents = _g(invoice, "amount_paid") or 0
            if amount_cents and amount_cents > 0:
                amount_brl = float(amount_cents) / 100.0
                from core.services.email_service import send_pro_charged_email
                await _fire_email(user_id, send_pro_charged_email, amount_brl, expires_dt)

    elif event["type"] == "customer.subscription.trial_will_end":
        # Stripe dispara ~3 dias antes do trial acabar. Email de aviso (item 38)
        # — fonte primária; scheduler interno fica como fallback se o webhook
        # falhar. Dedup via system_event_logs evita duplicar com o scheduler.
        sub = event["data"]["object"]
        user_id = _resolve_user(sub)
        if user_id:
            from core.observability import recent_event_exists
            if not recent_event_exists("trial_ending_email_sent", user_id, within_days=6):
                expires_dt = _subscription_period_end(sub)
                from core.services.email_service import send_trial_ending_email
                await _fire_email(user_id, send_trial_ending_email, expires_dt)
                await log_system_event(
                    "info",
                    "trial_ending_email_sent",
                    "Email de trial ending enviado (webhook trial_will_end).",
                    source="billing",
                    user_id=user_id,
                )

    elif event["type"] == "invoice.payment_failed":
        # Stripe vai retentar (smart retries). NAO movemos pra free aqui;
        # so marca past_due. O downgrade definitivo acontece em
        # customer.subscription.deleted quando a sub for de fato cancelada.
        invoice = event["data"]["object"]
        user_id = _resolve_user(invoice)
        if user_id:
            set_payment_status(user_id, "past_due")
            await log_system_event(
                "warning",
                "billing_payment_failed",
                "Falha de pagamento; assinatura past_due (Stripe vai retentar).",
                source="billing",
                user_id=user_id,
            )
            # Email com link pra atualizar cartao (item 40)
            from core.services.email_service import send_payment_failed_email
            try:
                email = await _user_email(user_id)
                if email:
                    await asyncio.to_thread(send_payment_failed_email, email, DASHBOARD_URL)
            except Exception as exc:
                print(f"[billing] email payment_failed falhou user={user_id}: {exc}")
            # Notificação admin
            try:
                from core.services.admin_notify import notify_payment_failed
                _attempt = _g(invoice, "attempt_count")
                await asyncio.to_thread(
                    notify_payment_failed,
                    user_id=user_id, email=email, attempt_count=_attempt,
                )
            except Exception as exc:
                print(f"[billing] admin notify payment_failed falhou user={user_id}: {exc}")

    elif event["type"] == "customer.subscription.deleted":
        obj     = event["data"]["object"]
        user_id = _resolve_user(obj)
        if user_id:
            # Captura expires_at ANTES de zerar plan_expires_at (pro email
            # mostrar ate quando o user mantem acesso aos recursos Pro).
            from db import get_auth_user as _gau
            user_snapshot = await asyncio.to_thread(_gau, int(user_id))
            expires_for_email = (user_snapshot or {}).get("plan_expires_at")
            update_user_plan(user_id, "free", None)
            set_payment_status(user_id, "canceled")
            print(f"[billing] user {user_id} → free (cancelado)")
            await log_system_event(
                "warning",
                "billing_subscription_canceled",
                "Assinatura cancelada; usuario voltou para free.",
                source="billing",
                user_id=user_id,
            )
            # Email de confirmacao de cancelamento (item 41)
            from core.services.email_service import send_subscription_canceled_email
            try:
                email = await _user_email(user_id)
                if email:
                    await asyncio.to_thread(send_subscription_canceled_email, email, expires_for_email, DASHBOARD_URL)
            except Exception as exc:
                print(f"[billing] email canceled falhou user={user_id}: {exc}")
            # Notificação admin
            try:
                from core.services.admin_notify import notify_subscription_canceled
                await asyncio.to_thread(
                    notify_subscription_canceled,
                    user_id=user_id, email=email, expires_at=expires_for_email,
                )
            except Exception as exc:
                print(f"[billing] admin notify canceled falhou user={user_id}: {exc}")

    return {"received": True}


@app.post("/billing/portal")
async def billing_portal(user_id: int = Depends(_get_current_user)):
    """
    Cria uma sessão no Stripe Customer Portal para o usuário gerenciar
    a assinatura (cancelar, trocar cartão, ver faturas).
    """
    if not STRIPE_SECRET_KEY:
        raise HTTPException(status_code=503, detail="Pagamentos ainda não configurados.")

    import stripe
    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from db import get_auth_user

    stripe.api_key = STRIPE_SECRET_KEY
    user = get_auth_user(user_id)
    if not user or not user.get("stripe_customer_id"):
        raise HTTPException(status_code=404, detail="Sem assinatura ativa.")

    portal = stripe.billing_portal.Session.create(
        customer=user["stripe_customer_id"],
        return_url=f"{DASHBOARD_URL}/app",
    )
    return {"portal_url": portal.url}


# ─── Chat IA (Pro v1 Fase 2 — Bloco A) ──────────────────────────────────────

AI_CHAT_MONTHLY_LIMIT = int(os.getenv("AI_CHAT_MONTHLY_LIMIT", "1000"))


class AIChatBody(BaseModel):
    message: str


@app.post("/ai/chat")
async def ai_chat(
    body: AIChatBody,
    user_id: int = Depends(require_pro_feature("ai_chat")),
):
    """
    Chat conversacional com o Piggy (IA Pro). Function calling + confirmação
    humana obrigatória em writes. Rate limit mensal (default 100 msgs).

    Body: {"message": "texto"}
    Response: {"reply": "...", "usage": {"used": N, "limit": N}}
    """
    text = (body.message or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="message vazio.")
    if len(text) > 2000:
        raise HTTPException(status_code=400, detail="message muito longo (máx 2000 chars).")

    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from core.services.ai_chat import chat as ai_chat_run

    reply = await asyncio.to_thread(
        ai_chat_run,
        user_id,
        text,
        monthly_limit=AI_CHAT_MONTHLY_LIMIT,
        platform="dashboard",
    )
    used_after = await asyncio.to_thread(_db_ai_usage, user_id)
    return {"reply": reply, "usage": {"used": used_after, "limit": AI_CHAT_MONTHLY_LIMIT}}


@app.get("/ai/messages")
async def ai_messages(
    user_id: int = Depends(require_pro_feature("ai_chat")),
    limit: int = 30,
):
    """
    Retorna as últimas mensagens visíveis do chat IA pra renderizar no widget.
    Filtra tool calls e mensagens system — só user/assistant com texto.
    """
    if not 1 <= limit <= 200:
        limit = 30

    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from db import ai_get_recent_messages

    rows = await asyncio.to_thread(ai_get_recent_messages, user_id, limit)
    out = []
    for r in rows:
        role = r.get("role")
        content = (r.get("content") or "").strip()
        if role not in ("user", "assistant") or not content:
            continue
        # Pula assistants vazios (só tool_calls, sem texto pro usuário)
        if role == "assistant" and r.get("tool_calls"):
            continue
        out.append({
            "role": role,
            "content": content,
            "created_at": r["created_at"].isoformat() if r.get("created_at") else None,
        })
    used_after = await asyncio.to_thread(_db_ai_usage, user_id)
    return {"messages": out, "usage": {"used": used_after, "limit": AI_CHAT_MONTHLY_LIMIT}}


def _db_ai_usage(user_id: int) -> int:
    """Lê o contador sem incrementar (pra devolver no response)."""
    from db import ai_get_usage_this_month
    return ai_get_usage_this_month(user_id)


@app.get("/conta")
async def conta_redirect(request: Request):
    """
    Atalho público (GET) usado em invoices, recibos e emails do Stripe:
    `pigbankai.com/conta` → leva o usuário direto para o ponto certo.

    - Não autenticado → landing com flag `login_required=conta`.
    - Autenticado sem assinatura ativa → página de planos.
    - Autenticado com `stripe_customer_id` → Stripe Customer Portal.
    """
    token = _get_auth_token_from_request(request, None)
    payload = _decode_jwt(token) if token else None
    if not payload or payload.get("type") != "auth":
        return RedirectResponse(url=_dashboard_url("/?login_required=conta"), status_code=302)

    user_id = int(payload["sub"])
    jti = payload.get("jti")
    if jti:
        session = await asyncio.to_thread(get_active_session, jti)
        if not session or int(session.get("user_id") or 0) != user_id:
            return RedirectResponse(url=_dashboard_url("/?login_required=conta"), status_code=302)

    if not STRIPE_SECRET_KEY:
        return RedirectResponse(url=_dashboard_url("/precos"), status_code=302)

    import stripe
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from db import get_auth_user

    stripe.api_key = STRIPE_SECRET_KEY
    user = get_auth_user(user_id)
    if not user or not user.get("stripe_customer_id"):
        return RedirectResponse(url=_dashboard_url("/precos"), status_code=302)

    portal = stripe.billing_portal.Session.create(
        customer=user["stripe_customer_id"],
        return_url=f"{DASHBOARD_URL}/settings",
    )
    return RedirectResponse(url=portal.url, status_code=302)


# ─── Magic link de acesso ao dashboard ───────────────────────────────────────

@app.get("/d/{code}")
async def dashboard_short_link(
    request: Request,
    code: str,
    view: str | None = None,
    next: str | None = None,
):
    """
    Resolve um magic link gerado pelo bot.
    O link é de uso único e cria uma sessão curta no navegador.
    `next` (opcional): rota interna pra redirecionar apos logar
    (ex: /precos, /conta). Tem prioridade sobre `view`.
    """
    from db import consume_dashboard_session

    user_id = consume_dashboard_session(code)
    if not user_id:
        expired_html = """<!DOCTYPE html>
<html lang="pt-BR">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Link expirado</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{background:#070b14;font-family:-apple-system,BlinkMacSystemFont,"SF Pro Display","Segoe UI",sans-serif;
display:flex;align-items:center;justify-content:center;min-height:100vh;color:rgba(255,255,255,.85)}
.box{text-align:center;max-width:400px;padding:48px 32px}
.icon{font-size:3.5rem;margin-bottom:20px}
h2{font-size:1.4rem;font-weight:600;margin-bottom:10px}
p{color:rgba(255,255,255,.5);line-height:1.7;margin-bottom:28px}
a{display:inline-block;padding:11px 28px;background:rgba(124,58,237,.45);
border:1px solid rgba(124,58,237,.55);border-radius:14px;color:white;
text-decoration:none;font-size:.9rem;transition:background .2s}
a:hover{background:rgba(124,58,237,.65)}
</style></head>
<body><div class="box">
<div class="icon">🔒</div>
<h2>Link expirado ou inválido</h2>
<p>Este link de acesso ao dashboard expirou ou já foi usado.<br>
Solicite um novo link digitando <strong style="color:rgba(255,255,255,.8)">dashboard</strong> no bot.<br>
Os links expiram em __MAGIC_LINK_MINUTES__ minutos e funcionam uma única vez.</p>
<a href="/">← Página inicial</a>
</div></body></html>""".replace("__MAGIC_LINK_MINUTES__", str(DASHBOARD_MAGIC_LINK_MINUTES))
        return HTMLResponse(content=expired_html, status_code=401)

    # next: rotas internas permitidas pra evitar open-redirect. Tem prioridade sobre view.
    _ALLOWED_NEXT_PREFIXES = ("/precos", "/conta", "/app", "/home", "/settings")
    if next and next.startswith("/") and any(next.startswith(p) for p in _ALLOWED_NEXT_PREFIXES):
        redirect_url = next
    else:
        target_view = view if view in {"overview", "investments", "open-finance"} else None
        redirect_url = "/settings?view=open-finance" if target_view == "open-finance" else (
            f"/app?view={urllib.parse.quote(target_view)}" if target_view else "/app"
        )
    response = RedirectResponse(url=redirect_url, status_code=302)
    # Magic-link tambem cria auth_session — aparece em "Dispositivos conectados"
    # e pode ser revogado individualmente como qualquer outra sessao.
    ip = get_remote_address(request) or None
    ua = request.headers.get("user-agent") or None
    jti = await asyncio.to_thread(create_session, int(user_id), ip=ip, user_agent=ua)
    _set_dashboard_cookie(response, int(user_id), jti=jti)
    # Tambem seta auth_token (cookie principal) com o mesmo jti — permite acessar
    # rotas que exigem auth completa (/conta, /api/me, etc) sem precisar logar
    # de novo. Sem isso, ?next=/conta caia em /?login_required=conta porque
    # /conta so olha pro auth_token, nao pro dashboard_token.
    # Magic link também emite refresh_token (sessão de 14d com idle 7d).
    try:
        from db import get_auth_user
        from core.refresh_tokens import create_refresh_token
        u = await asyncio.to_thread(get_auth_user, int(user_id))
        email = (u or {}).get("email") or ""
        if email:
            auth_jwt = _make_jwt(int(user_id), email, jti=jti)
            _set_auth_cookie(response, auth_jwt)
            refresh = await asyncio.to_thread(
                create_refresh_token, int(user_id), jti, ip=ip, user_agent=ua,
            )
            _set_refresh_cookie(response, refresh)
    except Exception:
        # Fail-soft — o dashboard ainda funciona com o dashboard_token, mas o
        # user vai "deslogar" ao navegar pra /conta. Logado pra ser visível.
        logging.getLogger(__name__).warning(
            "magic link: não emitiu auth/refresh token pro user %s — sessão degradada (só dashboard_token)",
            user_id, exc_info=True,
        )
    return response


# ─── Páginas estáticas, assets, SEO e health ─────────────────────────────────
# Movidas pra frontend/routes/static_pages.py (refactor Fase 1, Etapa 1).
app.include_router(static_pages_router)

# ─── Unsubscribe ─────────────────────────────────────────────────────────────

import hashlib as _hashlib
import hmac as _hmac
import base64 as _base64

from core.services.email_service import make_unsub_url  # noqa: E402


def _verify_unsub_token(user_id: int, email: str, token: str) -> bool:
    """Verifica token de unsubscribe usando a mesma lógica do email_service."""
    secret   = (JWT_SECRET or "pigbank-unsub").encode()
    payload  = f"{user_id}:{email}".encode()
    sig      = _hmac.new(secret, payload, _hashlib.sha256).digest()
    expected = _base64.urlsafe_b64encode(sig).decode().rstrip("=")
    return _hmac.compare_digest(expected, token)


@app.get("/unsubscribe")
async def unsubscribe(uid: int, token: str):
    # busca o email pelo user_id
    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT email FROM auth_accounts WHERE user_id = %s", (uid,)
            )
            row = await cur.fetchone()

    if not row:
        return HTMLResponse("<h2>Link inválido.</h2>", status_code=400)

    email = row["email"]
    if not _verify_unsub_token(uid, email, token):
        return HTMLResponse("<h2>Link inválido ou expirado.</h2>", status_code=400)

    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "UPDATE auth_accounts SET engagement_opt_out = true WHERE user_id = %s",
                (uid,),
            )
        await conn.commit()

    return HTMLResponse("""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width,initial-scale=1.0"/>
  <title>Descadastro — PigBank AI</title>
  <style>
    body{margin:0;padding:0;background:#0a0d18;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
         color:#e2e8f0;display:flex;align-items:center;justify-content:center;min-height:100vh}
    .card{background:#0f1320;border:1px solid rgba(255,255,255,.1);border-radius:20px;
          padding:48px 40px;text-align:center;max-width:440px}
    .icon{font-size:56px;margin-bottom:16px}
    h1{margin:0 0 12px;font-size:22px;color:#fff}
    p{color:rgba(255,255,255,.6);line-height:1.7;margin:0 0 24px}
    a{color:#7c3aed;text-decoration:none}
  </style>
</head>
<body>
  <div class="card">
    <div class="icon">🐷</div>
    <h1>Descadastro confirmado</h1>
    <p>Você não vai mais receber os emails de dicas e insights do Piggy.<br/>
       Seus emails de segurança (código de verificação, redefinição de senha) continuam normais.</p>
    <p style="font-size:13px">Arrependeu? Mande "reativar emails" pro bot!
  </div>
</body>
</html>
""")


# ─── HTTP API routes ─────────────────────────────────────────────────────────

@app.get("/data/{user_id}")
async def get_data(
    request: Request,
    user_id: int,
    year: int = None,
    month: int = None,
    page: int = 1,
    limit: int = 25,
    filter_type: str = "all",
    q: str = "",
):
    _authorize_dashboard_access(request, user_id)
    return await get_financial_data(user_id, year, month, page, limit, filter_type, q)

@app.get("/history/{user_id}")
async def monthly_history(request: Request, user_id: int, months: int = 6):
    _authorize_dashboard_access(request, user_id)
    if not 1 <= months <= 24:
        raise HTTPException(status_code=400, detail="months must be 1-24")
    # Free: limita janela ao history_days do plano (~1 mes). Nao retorna 403 pra
    # nao quebrar dashboard — apenas capa silenciosamente. Frontend pode ler o
    # plano e mostrar CTA "ver mais com Pro".
    from core.services.plan_service import is_pro
    if months > 1 and not is_pro(user_id):
        months = 1
    data = await get_monthly_history(user_id, months)
    return {"data": data}

class LaunchCreatePayload(BaseModel):
    tipo: str  # 'receita' | 'despesa' | 'credito'
    valor: float
    alvo: str | None = None
    nota: str | None = None
    categoria: str | None = None
    card_id: int | None = None    # obrigatório quando tipo='credito'
    parcelas: int | None = None   # opcional pra tipo='credito' (1 ou null = à vista)


@app.post("/launches/{user_id}")
async def create_launch_route(request: Request, user_id: int, payload: LaunchCreatePayload):
    """Cria um lançamento manual.

    - `receita` / `despesa` → cria em `launches` + atualiza saldo (mesmo fluxo do bot)
    - `credito` → cria em `credit_transactions` na fatura aberta do cartão
      escolhido (mesmo fluxo do `gastei X no cartao Y` do WhatsApp)
    """
    _authorize_dashboard_access(request, user_id)

    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from core.services.category_service import infer_category, learn_from_inference
    from utils_text import is_internal_category, canonicalize_category_label

    tipo = (payload.tipo or "").strip().lower()
    if tipo not in ("receita", "despesa", "credito"):
        raise HTTPException(status_code=400, detail="tipo deve ser 'receita', 'despesa' ou 'credito'.")

    try:
        valor = float(payload.valor)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Valor inválido.")
    if valor <= 0:
        raise HTTPException(status_code=400, detail="O valor deve ser maior que zero.")

    alvo = (payload.alvo or "").strip() or None
    nota_in = (payload.nota or "").strip() or None

    # Resolve categoria — explícita do form ou inferência (mesmo fluxo do bot).
    explicit = (payload.categoria or "").strip() or None

    # ── Crédito → add_credit_purchase (à vista) ou installments (parcelado) ─
    if tipo == "credito":
        from db import add_credit_purchase, add_credit_purchase_installments, get_card_by_id
        from utils_date import today_tz

        card_id = payload.card_id
        if not card_id:
            raise HTTPException(status_code=400, detail="Selecione um cartão para a compra no crédito.")

        # Valida parcelas
        n_parc = 1
        if payload.parcelas is not None:
            try:
                n_parc = int(payload.parcelas)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="Número de parcelas inválido.")
            if n_parc < 1 or n_parc > 60:
                raise HTTPException(status_code=400, detail="Parcelas deve ser entre 1 e 60.")

        card = await asyncio.to_thread(get_card_by_id, int(user_id), int(card_id))
        if not card:
            raise HTTPException(status_code=400, detail="Cartão não encontrado.")
        card_name = card.get("name") or "cartão"

        nota = nota_in or alvo or f"compra no crédito ({card_name})"
        inferred = await asyncio.to_thread(infer_category, int(user_id), nota, explicit)
        categoria = canonicalize_category_label(inferred.category) or "outros"

        purchased_at = await asyncio.to_thread(today_tz)

        # Parcelado (2x+) → cria N transações em N faturas futuras
        if n_parc > 1:
            try:
                result = await asyncio.to_thread(
                    add_credit_purchase_installments,
                    int(user_id),
                    int(card_id),
                    valor,
                    categoria,
                    nota,
                    purchased_at,
                    n_parc,
                )
                await asyncio.to_thread(
                    learn_from_inference,
                    int(user_id),
                    nota,
                    categoria,
                    target_hint=alvo or card_name,
                    reason=inferred.reason,
                )
            except ValueError as exc:
                raise HTTPException(status_code=400, detail=str(exc)) from exc
            except Exception as exc:
                raise HTTPException(status_code=500, detail=f"Erro ao registrar parcelamento: {exc}") from exc

            info, total = (result[0], result[1]) if isinstance(result, tuple) else (result, valor)
            return {
                "ok": True,
                "tipo": "credito",
                "mode": "installments",
                "installments_total": n_parc,
                "group_id": info.get("group_id"),
                "tx_ids": info.get("tx_ids"),
                "card_id": int(card_id),
                "card_name": card_name,
                "valor_total": float(total),
                "categoria": categoria,
                "alvo": alvo or card_name,
                "nota": nota,
            }

        # À vista (1x)
        try:
            tx_id, due, bill_id = await asyncio.to_thread(
                add_credit_purchase,
                int(user_id),
                int(card_id),
                valor,
                categoria,
                nota,
                purchased_at,
            )
            await asyncio.to_thread(
                learn_from_inference,
                int(user_id),
                nota,
                categoria,
                target_hint=alvo or card_name,
                reason=inferred.reason,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Erro ao registrar compra no crédito: {exc}") from exc

        return {
            "ok": True,
            "tipo": "credito",
            "credit_transaction_id": int(tx_id),
            "bill_id": int(bill_id),
            "card_id": int(card_id),
            "card_name": card_name,
            "valor": float(valor),
            "categoria": categoria,
            "alvo": alvo or card_name,
            "nota": nota,
            "due_amount": float(due),
        }

    # ── Receita / Despesa → fluxo padrão de launches ──────────────────────
    from db import add_launch_and_update_balance

    nota = nota_in or alvo or ("receita registrada pelo dashboard" if tipo == "receita" else "despesa registrada pelo dashboard")
    inferred = await asyncio.to_thread(infer_category, int(user_id), nota, explicit)
    categoria = canonicalize_category_label(inferred.category) or "outros"
    is_internal = is_internal_category(categoria)

    try:
        launch_id, user_seq, new_balance = await asyncio.to_thread(
            add_launch_and_update_balance,
            int(user_id),
            tipo,
            valor,
            alvo,
            nota,
            categoria,
            None,  # criado_em → now()
            is_internal,
        )
        await asyncio.to_thread(
            learn_from_inference,
            int(user_id),
            nota,
            categoria,
            target_hint=alvo,
            reason=inferred.reason,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Erro ao registrar lançamento: {exc}") from exc

    return {
        "ok": True,
        "launch_id": int(launch_id),
        "user_seq": int(user_seq),
        "tipo": tipo,
        "valor": float(valor),
        "categoria": categoria,
        "alvo": alvo,
        "nota": nota,
        "new_balance": float(new_balance),
        "is_internal_movement": is_internal,
    }


class LaunchEditPayload(BaseModel):
    categoria: str | None = None
    nota: str | None = None
    criado_em: str | None = None


@app.patch("/launches/{user_id}/{launch_id}")
async def update_launch_route(
    request: Request,
    user_id: int,
    launch_id: int,
    payload: LaunchEditPayload,
):
    """Atualiza campos editáveis de um lançamento (categoria e/ou descrição).

    Não altera saldo/efeitos. Lançamentos de cartão de crédito não passam
    por aqui — eles ficam em `credit_transactions`.
    """
    _authorize_dashboard_access(request, user_id)

    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from utils_text import canonicalize_category_label

    categoria_norm: str | None = None
    if payload.categoria is not None:
        raw = payload.categoria.strip()
        if not raw:
            raise HTTPException(status_code=400, detail="Categoria não pode ser vazia.")
        categoria_norm = canonicalize_category_label(raw) or raw.lower()

    nota_norm: str | None = None
    if payload.nota is not None:
        nota_norm = payload.nota.strip()
        if len(nota_norm) > 200:
            raise HTTPException(status_code=400, detail="Descrição muito longa (máx. 200 caracteres).")

    criado_em_dt: datetime | None = None
    if payload.criado_em is not None and payload.criado_em.strip():
        from utils_date import _tz
        raw_dt = payload.criado_em.strip().replace("Z", "+00:00")
        try:
            criado_em_dt = datetime.fromisoformat(raw_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Data inválida.")
        if criado_em_dt.tzinfo is None:
            criado_em_dt = criado_em_dt.replace(tzinfo=_tz())

    if categoria_norm is None and nota_norm is None and criado_em_dt is None:
        raise HTTPException(status_code=400, detail="Nada para atualizar.")

    changed = await asyncio.to_thread(
        update_launch_fields,
        user_id,
        launch_id,
        categoria=categoria_norm,
        nota=nota_norm,
        criado_em=criado_em_dt,
    )
    if not changed:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado.")
    _invalidate_dashboard_current_cache(user_id)
    return {
        "ok": True,
        "launch_id": launch_id,
        "categoria": categoria_norm,
        "nota": nota_norm,
        "criado_em": criado_em_dt.isoformat() if criado_em_dt else None,
    }


@app.delete("/launches/{user_id}/{launch_id}")
async def delete_launch_route(
    request: Request,
    user_id: int,
    launch_id: int,
):
    """Apaga um lançamento e reverte seus efeitos no saldo.

    Usa `delete_launch_and_rollback` em uma única transação. Se o lançamento
    não pertence ao usuário ou já foi apagado, devolve 404.
    """
    _authorize_dashboard_access(request, user_id)

    try:
        await asyncio.to_thread(delete_launch_and_rollback, user_id, int(launch_id))
    except LookupError:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado.")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Erro ao apagar lançamento: {exc}") from exc

    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "launch_id": int(launch_id)}


@app.patch("/credit-transactions/{user_id}/{tx_id}")
async def update_credit_transaction_route(
    request: Request,
    user_id: int,
    tx_id: int,
    payload: LaunchEditPayload,
):
    """Atualiza categoria e/ou descrição de uma compra no cartão de crédito.

    Não altera valor, cartão ou data — esses mudariam o saldo da fatura ou a
    janela de fechamento; pra mexer nesses campos, o user deve apagar e
    recriar.
    """
    _authorize_dashboard_access(request, user_id)

    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from utils_text import canonicalize_category_label

    categoria_norm: str | None = None
    if payload.categoria is not None:
        raw = payload.categoria.strip()
        if not raw:
            raise HTTPException(status_code=400, detail="Categoria não pode ser vazia.")
        categoria_norm = canonicalize_category_label(raw) or raw.lower()

    nota_norm: str | None = None
    if payload.nota is not None:
        nota_norm = payload.nota.strip()
        if len(nota_norm) > 200:
            raise HTTPException(status_code=400, detail="Descrição muito longa (máx. 200 caracteres).")

    if categoria_norm is None and nota_norm is None:
        raise HTTPException(status_code=400, detail="Nada para atualizar.")

    changed = await asyncio.to_thread(
        update_credit_transaction_fields,
        user_id,
        tx_id,
        categoria=categoria_norm,
        nota=nota_norm,
    )
    if not changed:
        raise HTTPException(status_code=404, detail="Compra no crédito não encontrada.")
    _invalidate_dashboard_current_cache(user_id)
    return {
        "ok": True,
        "tx_id": tx_id,
        "categoria": categoria_norm,
        "nota": nota_norm,
    }


@app.delete("/credit-transactions/{user_id}/{tx_id}")
async def delete_credit_transaction_route(
    request: Request,
    user_id: int,
    tx_id: int,
):
    """Apaga uma compra no cartão de crédito.

    Se a compra faz parte de um parcelamento (group_id), TODO o grupo é
    desfeito — `undo_credit_transaction` já implementa esse comportamento.
    Retorna `mode` ('single' ou 'group') pro front exibir feedback adequado.
    """
    _authorize_dashboard_access(request, user_id)

    try:
        result = await asyncio.to_thread(undo_credit_transaction, user_id, int(tx_id))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Erro ao apagar compra: {exc}") from exc

    if result is None:
        raise HTTPException(status_code=404, detail="Compra no crédito não encontrada.")

    # `undo_credit_transaction` retorna mode="single"; `undo_installment_group`
    # retorna {group_id, removed_count, removed_total} sem mode — normalizamos aqui.
    mode = result.get("mode") or ("group" if result.get("group_id") else "single")

    _invalidate_dashboard_current_cache(user_id)
    return {
        "ok": True,
        "tx_id": int(tx_id),
        "mode": mode,
        "removed_total": float(result.get("removed_total") or 0),
        "removed_count": int(result.get("removed_count") or 1),
    }


class PocketCreatePayload(BaseModel):
    name: str
    description: str | None = None
    interest_enabled: bool = True
    interest_rate: float = 1.0


class PocketMovePayload(BaseModel):
    amount: float
    nota: str | None = None


@app.post("/pockets/{user_id}")
async def create_pocket_route(request: Request, user_id: int, payload: PocketCreatePayload):
    """Cria uma caixinha (pocket) com saldo zero."""
    _authorize_dashboard_access(request, user_id)

    # Free: respeita pockets_max do plano (1). Pro: ilimitado.
    from core.services.plan_service import get_user_limits
    from db.pockets import list_pockets
    limits = get_user_limits(user_id)
    pockets_max = limits["pockets_max"]
    if pockets_max is not None:
        existing = await asyncio.to_thread(list_pockets, user_id)
        if len(existing) >= pockets_max:
            raise HTTPException(
                status_code=403,
                detail={"error": "pro_required", "feature": "pockets_unlimited"},
            )

    name = (payload.name or "").strip()
    description = (payload.description or "").strip() or None
    interest_rate = float(payload.interest_rate or 1.0)
    if not name:
        raise HTTPException(status_code=400, detail="Nome da caixinha é obrigatório.")
    if len(name) > 80:
        raise HTTPException(status_code=400, detail="Nome muito longo (máx. 80 caracteres).")
    if description and len(description) > 200:
        raise HTTPException(status_code=400, detail="Descrição muito longa (máx. 200 caracteres).")
    if interest_rate <= 0:
        raise HTTPException(status_code=400, detail="Rendimento inválido.")

    try:
        launch_id, pocket_id, canon = await asyncio.to_thread(
            create_pocket,
            user_id,
            name,
            None,
            description,
            interest_enabled=payload.interest_enabled,
            interest_rate=interest_rate,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    _invalidate_dashboard_current_cache(user_id)
    return {
        "ok": True,
        "created": launch_id is not None,
        "pocket": {
            "id": int(pocket_id),
            "name": canon,
            "description": description,
            "interest_enabled": bool(payload.interest_enabled),
            "interest_rate": interest_rate,
            "interest_period": "cdi",
        },
    }


class PocketMetaPayload(BaseModel):
    name: str | None = None
    description: str | None = None
    target_amount: float | None = None
    target_date: str | None = None
    emoji: str | None = None
    color: str | None = None
    status: str | None = None
    interest_enabled: bool | None = None
    interest_rate: float | None = None
    clear_target: bool = False


@app.patch("/pockets/{user_id}/{pocket_id}/meta")
async def update_pocket_meta_route(
    request: Request, user_id: int, pocket_id: int, payload: PocketMetaPayload
):
    """PATCH em metadata da caixinha — incluindo target_amount/date pra virar meta.

    Pra remover a meta (mantendo a caixinha), passar `clear_target=true`.
    """
    _authorize_dashboard_access(request, user_id)
    from db.pockets import update_pocket_meta

    try:
        row = await asyncio.to_thread(
            update_pocket_meta,
            user_id, pocket_id,
            name=payload.name, description=payload.description,
            target_amount=payload.target_amount, target_date=payload.target_date,
            emoji=payload.emoji, color=payload.color, status=payload.status,
            interest_enabled=payload.interest_enabled, interest_rate=payload.interest_rate,
            clear_target=payload.clear_target,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if not row:
        raise HTTPException(status_code=404, detail="Caixinha não encontrada.")
    _invalidate_dashboard_current_cache(user_id)
    pocket = dict(row)
    if pocket.get("target_amount") is not None:
        pocket["target_amount"] = float(pocket["target_amount"])
    if pocket.get("target_date"):
        pocket["target_date"] = pocket["target_date"].isoformat()
    if pocket.get("balance") is not None:
        pocket["balance"] = float(pocket["balance"])
    if pocket.get("interest_rate") is not None:
        pocket["interest_rate"] = float(pocket["interest_rate"])
    if pocket.get("last_interest_date"):
        pocket["last_interest_date"] = pocket["last_interest_date"].isoformat()
    return {"ok": True, "pocket": pocket}


@app.get("/goals/{user_id}/status")
async def goals_status_route(request: Request, user_id: int):
    """Lista as caixinhas COM meta (target_amount NOT NULL) + cálculo de ritmo.

    Retorna pra cada meta: pct_complete, days_left, monthly_pace_needed,
    monthly_pace_current (média dos últimos 3 meses), projected_at_current_pace,
    status_indicator ('ahead'|'on_track'|'tight'|'behind').
    """
    _authorize_dashboard_access(request, user_id)
    from db.pockets import list_pockets
    from db.connection import get_conn
    from datetime import date

    def _compute():
        pockets = list_pockets(user_id)
        goals = []
        today = date.today()
        for p in pockets:
            ta = p.get("target_amount")
            saved = float(p.get("balance") or 0)
            is_goal = ta is not None
            if not is_goal:
                # Caixinha sem meta: retorna info mínima
                goals.append({
                    "id": p["id"],
                    "name": p["name"],
                    "balance": saved,
                    "target_amount": None,
                    "target_date": None,
                    "emoji": p.get("emoji"),
                    "color": p.get("color"),
                    "status": p.get("status") or "active",
                    "description": p.get("description"),
                    "interest_enabled": bool(p.get("interest_enabled")),
                    "interest_rate": float(p.get("interest_rate") or 1),
                    "interest_period": p.get("interest_period") or "cdi",
                    "is_goal": False,
                    "pct_complete": None,
                    "remaining": None,
                    "days_left": None,
                    "monthly_pace_needed": None,
                    "monthly_pace_current": None,
                    "projected_months": None,
                    "indicator": "no_target",
                })
                continue
            tgt = float(ta)
            pct = (saved / tgt * 100.0) if tgt > 0 else 0.0
            td = p.get("target_date")
            days_left = (td - today).days if td else None
            remaining = max(0.0, tgt - saved)
            monthly_pace_needed = None
            if td and days_left and days_left > 0 and remaining > 0:
                months_left = max(1, days_left / 30.0)
                monthly_pace_needed = remaining / months_left

            # Ritmo atual: médio dos últimos 90 dias (depositos - saques)
            monthly_pace_current = None
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        select
                          coalesce(sum(case when tipo = 'deposito_caixinha' then valor else 0 end), 0) -
                          coalesce(sum(case when tipo = 'saque_caixinha' then valor else 0 end), 0)
                          as net
                        from launches
                        where user_id=%s and alvo=%s
                          and criado_em >= now() - interval '90 days'
                        """,
                        (user_id, p["name"]),
                    )
                    r = cur.fetchone()
                    net90 = float(r["net"] or 0)
                    monthly_pace_current = net90 / 3.0  # média mensal

            # Status indicator
            indicator = "active"
            if pct >= 100:
                indicator = "achieved"
            elif days_left is not None and days_left < 0:
                indicator = "behind"
            elif monthly_pace_needed and monthly_pace_current is not None:
                if monthly_pace_current >= monthly_pace_needed * 1.1:
                    indicator = "ahead"
                elif monthly_pace_current >= monthly_pace_needed * 0.9:
                    indicator = "on_track"
                elif monthly_pace_current >= monthly_pace_needed * 0.5:
                    indicator = "tight"
                else:
                    indicator = "behind"

            # Projeção: a esse ritmo, quanto tempo até atingir?
            projected_months = None
            if monthly_pace_current and monthly_pace_current > 0 and remaining > 0:
                projected_months = remaining / monthly_pace_current

            goals.append({
                "id": p["id"],
                "name": p["name"],
                "balance": saved,
                "target_amount": tgt,
                "target_date": td.isoformat() if td else None,
                "emoji": p.get("emoji"),
                "color": p.get("color"),
                "status": p.get("status") or "active",
                "description": p.get("description"),
                "interest_enabled": bool(p.get("interest_enabled")),
                "interest_rate": float(p.get("interest_rate") or 1),
                "interest_period": p.get("interest_period") or "cdi",
                "is_goal": True,
                "pct_complete": round(pct, 1),
                "remaining": round(remaining, 2),
                "days_left": days_left,
                "monthly_pace_needed": round(monthly_pace_needed, 2) if monthly_pace_needed else None,
                "monthly_pace_current": round(monthly_pace_current, 2) if monthly_pace_current is not None else None,
                "projected_months": round(projected_months, 1) if projected_months else None,
                "indicator": indicator,
            })
        return goals

    goals = await asyncio.to_thread(_compute)
    return {"ok": True, "goals": goals}


@app.delete("/pockets/{user_id}/{pocket_name:path}")
async def delete_pocket_route(request: Request, user_id: int, pocket_name: str):
    """Exclui uma caixinha (apenas se o saldo estiver zerado)."""
    _authorize_dashboard_access(request, user_id)
    name = urllib.parse.unquote(pocket_name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nome da caixinha é obrigatório.")

    try:
        launch_id, canon = await asyncio.to_thread(delete_pocket, int(user_id), name)
    except LookupError as exc:
        raise HTTPException(status_code=404, detail="Caixinha não encontrada.") from exc
    except ValueError as exc:
        msg = str(exc)
        if msg == "POCKET_NOT_ZERO":
            raise HTTPException(
                status_code=400,
                detail="Zere o saldo (saque para a conta) antes de remover a caixinha.",
            ) from exc
        raise HTTPException(status_code=400, detail=msg) from exc

    _invalidate_dashboard_current_cache(int(user_id))
    return {"ok": True, "launch_id": launch_id, "name": canon}


def _pocket_move_error(exc: ValueError) -> HTTPException:
    msg = str(exc)
    mapping = {
        "AMOUNT_INVALID": "O valor precisa ser maior que zero.",
        "INSUFFICIENT_ACCOUNT": "Saldo da conta principal não cobre esse valor.",
        "INSUFFICIENT_POCKET": "A caixinha não tem esse valor disponível.",
    }
    return HTTPException(status_code=400, detail=mapping.get(msg, msg))


@app.post("/pockets/{user_id}/{pocket_name:path}/deposit")
async def pocket_deposit_route(request: Request, user_id: int, pocket_name: str, payload: PocketMovePayload):
    """Conta principal → caixinha."""
    _authorize_dashboard_access(request, user_id)
    name = urllib.parse.unquote(pocket_name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nome da caixinha é obrigatório.")
    nota = (payload.nota or "").strip() or None
    try:
        launch_id, new_acc, new_pocket, canon = await asyncio.to_thread(
            pocket_deposit_from_account, int(user_id), name, payload.amount, nota,
        )
    except LookupError as exc:
        raise HTTPException(status_code=404, detail="Caixinha não encontrada.") from exc
    except ValueError as exc:
        raise _pocket_move_error(exc) from exc

    _invalidate_dashboard_current_cache(int(user_id))
    return json.loads(jdump({
        "ok": True,
        "launch_id": launch_id,
        "name": canon,
        "account_balance": new_acc,
        "pocket_balance": new_pocket,
    }))


@app.post("/pockets/{user_id}/{pocket_name:path}/withdraw")
async def pocket_withdraw_route(request: Request, user_id: int, pocket_name: str, payload: PocketMovePayload):
    """Caixinha → conta principal."""
    _authorize_dashboard_access(request, user_id)
    name = urllib.parse.unquote(pocket_name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nome da caixinha é obrigatório.")
    nota = (payload.nota or "").strip() or None
    try:
        launch_id, new_acc, new_pocket, canon, taxes = await asyncio.to_thread(
            pocket_withdraw_to_account, int(user_id), name, payload.amount, nota,
        )
    except LookupError as exc:
        raise HTTPException(status_code=404, detail="Caixinha não encontrada.") from exc
    except ValueError as exc:
        raise _pocket_move_error(exc) from exc

    _invalidate_dashboard_current_cache(int(user_id))
    return json.loads(jdump({
        "ok": True,
        "launch_id": launch_id,
        "name": canon,
        "account_balance": new_acc,
        "pocket_balance": new_pocket,
        "tax_summary": taxes,
    }))


@app.get("/pockets/{user_id}/{pocket_name:path}/history")
async def get_pocket_history_route(request: Request, user_id: int, pocket_name: str, limit: int = 100):
    """Histórico de depósitos e saques de uma caixinha específica."""
    _authorize_dashboard_access(request, user_id)
    name = (pocket_name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nome da caixinha é obrigatório.")
    limit = max(min(int(limit or 100), 500), 1)

    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                """
                SELECT id, name, balance, target_amount, target_date, emoji, color, status,
                       description, interest_enabled, interest_rate, interest_period,
                       interest_tax_profile, last_interest_date
                FROM pockets
                WHERE user_id = %s AND lower(name) = lower(%s)
                LIMIT 1
                """,
                (int(user_id), name),
            )
            pocket_row = await cur.fetchone()
            if not pocket_row:
                raise HTTPException(status_code=404, detail="Caixinha não encontrada.")

            canon = pocket_row["name"]

            await cur.execute(
                """
                SELECT id, tipo, valor, alvo, nota, criado_em
                FROM launches
                WHERE user_id = %s
                  AND lower(alvo) = lower(%s)
                  AND tipo IN ('deposito_caixinha', 'saque_caixinha', 'criar_caixinha')
                ORDER BY criado_em DESC, id DESC
                LIMIT %s
                """,
                (int(user_id), canon, limit),
            )
            rows = await cur.fetchall()

    history = []
    deposits_total = 0.0
    withdrawals_total = 0.0
    for r in (rows or []):
        v = float(r["valor"] or 0)
        tipo = r["tipo"]
        if tipo == "deposito_caixinha":
            deposits_total += v
        elif tipo == "saque_caixinha":
            withdrawals_total += v
        history.append({
            "id": int(r["id"]),
            "tipo": tipo,
            "valor": v,
            "nota": r.get("nota"),
            "criado_em": r["criado_em"].isoformat() if r.get("criado_em") else None,
        })

    return {
        "ok": True,
        "pocket": {
            "id": int(pocket_row["id"]),
            "name": canon,
            "balance": float(pocket_row["balance"] or 0),
            "target_amount": float(pocket_row["target_amount"]) if pocket_row.get("target_amount") is not None else None,
            "target_date": pocket_row["target_date"].isoformat() if pocket_row.get("target_date") else None,
            "emoji": pocket_row.get("emoji"),
            "color": pocket_row.get("color"),
            "status": pocket_row.get("status") or "active",
            "description": pocket_row.get("description"),
            "interest_enabled": bool(pocket_row.get("interest_enabled")),
            "interest_rate": float(pocket_row.get("interest_rate") or 1),
            "interest_period": pocket_row.get("interest_period") or "cdi",
            "interest_tax_profile": pocket_row.get("interest_tax_profile"),
            "last_interest_date": pocket_row["last_interest_date"].isoformat()
                                  if pocket_row.get("last_interest_date") else None,
        },
        "totals": {
            "deposits": deposits_total,
            "withdrawals": withdrawals_total,
            "count": len(history),
        },
        "history": history,
    }


class CardCreatePayload(BaseModel):
    name: str
    closing_day: int
    due_day: int
    color: str | None = None
    flag: str | None = None
    last4: str | None = None
    credit_limit: float | None = None


@app.post("/cards/{user_id}")
async def create_card_route(request: Request, user_id: int, payload: CardCreatePayload):
    """Cria um cartão de crédito."""
    _authorize_dashboard_access(request, user_id)

    # Free: respeita cards_max do plano (1). Pro: ilimitado.
    from core.services.plan_service import get_user_limits
    from db.cards import list_cards
    limits = get_user_limits(user_id)
    cards_max = limits["cards_max"]
    if cards_max is not None:
        existing = await asyncio.to_thread(list_cards, user_id)
        if len(existing) >= cards_max:
            raise HTTPException(
                status_code=403,
                detail={"error": "pro_required", "feature": "cards_unlimited"},
            )

    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nome do cartão é obrigatório.")
    if len(name) > 80:
        raise HTTPException(status_code=400, detail="Nome muito longo (máx. 80 caracteres).")
    if not (1 <= payload.closing_day <= 31):
        raise HTTPException(status_code=400, detail="Dia de fechamento deve estar entre 1 e 31.")
    if not (1 <= payload.due_day <= 31):
        raise HTTPException(status_code=400, detail="Dia de vencimento deve estar entre 1 e 31.")
    if payload.credit_limit is not None and payload.credit_limit < 0:
        raise HTTPException(status_code=400, detail="Limite não pode ser negativo.")

    try:
        card_id = await asyncio.to_thread(
            create_card, user_id, name, payload.closing_day, payload.due_day,
            color=payload.color, flag=payload.flag, last4=payload.last4,
            credit_limit=payload.credit_limit,
        )
    except ValueError as exc:
        msg = str(exc)
        if msg.startswith("nome_duplicado:"):
            raise HTTPException(status_code=409, detail=f"Já existe um cartão chamado \"{name}\".") from exc
        if msg.startswith("color_invalido:"):
            raise HTTPException(status_code=400, detail="Cor inválida. Use: purple, coral, gold, green, blue ou gray.") from exc
        if msg.startswith("flag_invalida:"):
            raise HTTPException(status_code=400, detail="Bandeira inválida. Use: Visa, Mastercard, Elo, Amex, Hipercard ou Outros.") from exc
        if msg.startswith("last4_invalido:"):
            raise HTTPException(status_code=400, detail="Últimos 4 dígitos devem ser exatamente 4 números.") from exc
        raise HTTPException(status_code=400, detail=msg) from exc

    _invalidate_dashboard_current_cache(user_id)
    return {
        "ok": True,
        "card": {
            "id": int(card_id),
            "name": name,
            "closing_day": payload.closing_day,
            "due_day": payload.due_day,
            "color": payload.color,
            "flag": payload.flag,
            "last4": payload.last4,
            "credit_limit": payload.credit_limit,
        },
    }


class CardUpdatePayload(BaseModel):
    name: str | None = None
    closing_day: int | None = None
    due_day: int | None = None
    color: str | None = None
    flag: str | None = None
    last4: str | None = None
    credit_limit: float | None = None
    clear_last4: bool = False
    clear_limit: bool = False


class CardReorderPayload(BaseModel):
    ordered_ids: list[int]


@app.get("/cards/{user_id}/summary")
async def cards_summary_route(request: Request, user_id: int):
    """Lista todos os cartões com dados agregados (fatura aberta, limite usado).
    Alimenta a view /app#cards do dashboard.

    Performance: faz 1 query Postgres com subqueries agregadas em vez de
    N+1 (era ~22 queries × latência por chamada).
    """
    _authorize_dashboard_access(request, user_id)
    from db.connection import get_conn

    def _build():
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    select
                      c.id, c.name, c.color, c.flag, c.last4,
                      c.closing_day, c.due_day, c.credit_limit,
                      (u.default_card_id = c.id) as is_default,
                      coalesce((
                        select b.total - coalesce(b.paid_amount, 0)
                          from credit_bills b
                         where b.card_id = c.id and b.user_id = c.user_id and b.status = 'open'
                         order by b.period_end desc
                         limit 1
                      ), 0) as open_due,
                      coalesce((
                        select b.total
                          from credit_bills b
                         where b.card_id = c.id and b.user_id = c.user_id and b.status = 'open'
                         order by b.period_end desc
                         limit 1
                      ), 0) as open_total,
                      coalesce((
                        select b.paid_amount
                          from credit_bills b
                         where b.card_id = c.id and b.user_id = c.user_id and b.status = 'open'
                         order by b.period_end desc
                         limit 1
                      ), 0) as open_paid,
                      (
                        select b.id
                          from credit_bills b
                         where b.card_id = c.id and b.user_id = c.user_id and b.status = 'open'
                         order by b.period_end desc
                         limit 1
                      ) as open_bill_id,
                      (
                        select b.period_end
                          from credit_bills b
                         where b.card_id = c.id and b.user_id = c.user_id and b.status = 'open'
                         order by b.period_end desc
                         limit 1
                      ) as open_period_end,
                      coalesce((
                        select sum(b.total - coalesce(b.paid_amount, 0))
                          from credit_bills b
                         where b.card_id = c.id and b.user_id = c.user_id
                      ), 0) as credit_used
                    from credit_cards c
                    left join users u on u.id = c.user_id
                    where c.user_id = %s
                    order by c.name
                    """,
                    (user_id,),
                )
                rows = cur.fetchall()

        out = []
        for r in rows:
            credit_limit = float(r["credit_limit"]) if r.get("credit_limit") is not None else None
            usage = float(r["credit_used"] or 0)
            available = (credit_limit - usage) if credit_limit is not None else None
            out.append({
                "id": int(r["id"]),
                "name": r["name"],
                "color": r.get("color"),
                "flag": r.get("flag"),
                "last4": r.get("last4"),
                "closing_day": r["closing_day"],
                "due_day": r["due_day"],
                "credit_limit": credit_limit,
                "is_default": bool(r.get("is_default")),
                "open_bill": {
                    "id": int(r["open_bill_id"]) if r.get("open_bill_id") else None,
                    "total": float(r["open_total"] or 0),
                    "paid_amount": float(r["open_paid"] or 0),
                    "due_amount": float(r["open_due"] or 0),
                    "period_end": r["open_period_end"].isoformat() if r.get("open_period_end") else None,
                },
                "next_bill": {"total": 0.0, "period_end": None},  # TODO Sprint 2: calcular se necessário
                "credit_used": usage,
                "credit_available": available,
            })
        return out

    cards = await asyncio.to_thread(_build)
    return {"ok": True, "cards": cards}


@app.patch("/cards/{user_id}/reorder")
async def reorder_cards_route(request: Request, user_id: int, payload: CardReorderPayload):
    """Salva nova ordem manual dos cartões (drag-to-reorder).
    Recebe ordered_ids = [id_primeiro, id_segundo, ...] e grava display_order
    sequencial (0..N-1). Cartões fora da lista mantêm o que já tinham."""
    _authorize_dashboard_access(request, user_id)

    if not payload.ordered_ids:
        raise HTTPException(status_code=400, detail="Lista de cartões vazia.")
    if len(payload.ordered_ids) > 200:
        raise HTTPException(status_code=400, detail="Muitos cartões na lista.")

    from db.cards import reorder_cards

    updated = await asyncio.to_thread(reorder_cards, user_id, payload.ordered_ids)
    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "updated": updated}


@app.patch("/cards/{user_id}/{card_id}")
async def update_card_route(request: Request, user_id: int, card_id: int, payload: CardUpdatePayload):
    """Edita campos de um cartão (name, dias, color, flag, last4, credit_limit).
    Use clear_last4=true ou clear_limit=true pra apagar explicitamente."""
    _authorize_dashboard_access(request, user_id)

    from db.cards import update_card_meta, get_card_by_id

    if payload.name is not None:
        n = (payload.name or "").strip()
        if not n:
            raise HTTPException(status_code=400, detail="Nome do cartão é obrigatório.")
        if len(n) > 80:
            raise HTTPException(status_code=400, detail="Nome muito longo (máx. 80 caracteres).")
    if payload.closing_day is not None and not (1 <= payload.closing_day <= 31):
        raise HTTPException(status_code=400, detail="Dia de fechamento deve estar entre 1 e 31.")
    if payload.due_day is not None and not (1 <= payload.due_day <= 31):
        raise HTTPException(status_code=400, detail="Dia de vencimento deve estar entre 1 e 31.")
    if payload.credit_limit is not None and payload.credit_limit < 0:
        raise HTTPException(status_code=400, detail="Limite não pode ser negativo.")

    try:
        updated = await asyncio.to_thread(
            update_card_meta, user_id, card_id,
            name=payload.name,
            closing_day=payload.closing_day,
            due_day=payload.due_day,
            color=payload.color,
            flag=payload.flag,
            last4=payload.last4,
            credit_limit=payload.credit_limit,
            clear_last4=payload.clear_last4,
            clear_limit=payload.clear_limit,
        )
    except ValueError as exc:
        msg = str(exc)
        if msg.startswith("nome_duplicado:"):
            raise HTTPException(status_code=409, detail="Já existe outro cartão com esse nome.") from exc
        if msg.startswith("color_invalido:"):
            raise HTTPException(status_code=400, detail="Cor inválida.") from exc
        if msg.startswith("flag_invalida:"):
            raise HTTPException(status_code=400, detail="Bandeira inválida.") from exc
        if msg.startswith("last4_invalido:"):
            raise HTTPException(status_code=400, detail="Últimos 4 dígitos devem ser exatamente 4 números.") from exc
        raise HTTPException(status_code=400, detail=msg) from exc

    if not updated:
        raise HTTPException(status_code=404, detail="Cartão não encontrado ou nenhum campo alterado.")

    _invalidate_dashboard_current_cache(user_id)
    fresh = await asyncio.to_thread(get_card_by_id, user_id, card_id)
    return {
        "ok": True,
        "card": {
            "id": int(fresh["id"]),
            "name": fresh["name"],
            "closing_day": fresh["closing_day"],
            "due_day": fresh["due_day"],
            "color": fresh.get("color"),
            "flag": fresh.get("flag"),
            "last4": fresh.get("last4"),
            "credit_limit": float(fresh["credit_limit"]) if fresh.get("credit_limit") is not None else None,
        },
    }


@app.get("/cards/{user_id}/{card_id}/delete-impact")
async def card_delete_impact_route(request: Request, user_id: int, card_id: int):
    """Retorna o que será apagado se o cartão for excluído.
    Frontend usa pra mostrar confirmação ao usuário."""
    _authorize_dashboard_access(request, user_id)
    from db.cards import get_card_delete_impact, get_card_by_id

    card = await asyncio.to_thread(get_card_by_id, user_id, card_id)
    if not card:
        raise HTTPException(status_code=404, detail="Cartão não encontrado.")
    impact = await asyncio.to_thread(get_card_delete_impact, user_id, card_id)
    return {
        "ok": True,
        "card_name": card["name"],
        "impact": impact,
    }


@app.delete("/cards/{user_id}/{card_id}")
async def delete_card_route(request: Request, user_id: int, card_id: int):
    """Exclui um cartão e tudo vinculado (faturas, parcelamentos).
    ON DELETE CASCADE cuida das tabelas credit_bills/credit_transactions."""
    _authorize_dashboard_access(request, user_id)
    from db.cards import delete_card, get_card_by_id

    card = await asyncio.to_thread(get_card_by_id, user_id, card_id)
    if not card:
        raise HTTPException(status_code=404, detail="Cartão não encontrado.")

    deleted = await asyncio.to_thread(delete_card, user_id, card_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Cartão não encontrado.")

    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "deleted_card_id": card_id, "deleted_card_name": card["name"]}


# ─────────────────────────────────────────────────────────────────────────────
# Parcelamentos (Sprint 2 — 2026-05-14)
# ─────────────────────────────────────────────────────────────────────────────

class InstallmentUpdatePayload(BaseModel):
    nome: str | None = None
    categoria: str | None = None


@app.get("/installments/{user_id}/list")
async def installments_list_route(
    request: Request,
    user_id: int,
    sort: str = "urgency",
):
    """Lista parcelamentos do user com detalhe por parcela.
    Alimenta a view /app#installments do dashboard."""
    _authorize_dashboard_access(request, user_id)
    from db.cards import list_installment_groups_detailed

    if sort not in ("urgency", "recent"):
        sort = "urgency"
    groups = await asyncio.to_thread(list_installment_groups_detailed, user_id, sort)
    return {"ok": True, "installments": groups}


@app.get("/installments/{user_id}/{group_id}/delete-impact")
async def installment_delete_impact_route(request: Request, user_id: int, group_id: str):
    """Retorna o impacto de excluir o parcelamento (sem excluir).
    Frontend usa pra escolher mensagem do modal: com vs sem parcelas pagas."""
    _authorize_dashboard_access(request, user_id)
    from db.cards import get_installment_group_delete_impact

    impact = await asyncio.to_thread(get_installment_group_delete_impact, user_id, group_id)
    if not impact:
        raise HTTPException(status_code=404, detail="Parcelamento não encontrado.")
    return {"ok": True, "impact": impact}


@app.post("/installments/{user_id}/{group_id}/anticipate")
async def installment_anticipate_route(request: Request, user_id: int, group_id: str):
    """Antecipa a próxima parcela pendente: paga à vista da conta corrente.
    Deleta a tx do parcelamento + reduz fatura aberta + cria launch de despesa."""
    _authorize_dashboard_access(request, user_id)
    from db.cards import anticipate_installment

    result = await asyncio.to_thread(anticipate_installment, user_id, group_id)
    if not result:
        raise HTTPException(
            status_code=404,
            detail="Sem parcelas pendentes pra antecipar (parcelamento já quitado).",
        )
    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "result": result}


@app.delete("/installments/{user_id}/{group_id}")
async def installment_delete_route(request: Request, user_id: int, group_id: str):
    """Exclui parcelamento. Comportamento Option B:
    - tx em faturas abertas: deletadas (open bill cai, saldo do mês volta)
    - tx em faturas pagas: viram órfãs (group_id=null, nota+sufixo).
      Fatura paga intacta — dinheiro não volta pra conta."""
    _authorize_dashboard_access(request, user_id)
    from db.cards import undo_installment_group

    result = await asyncio.to_thread(undo_installment_group, user_id, group_id)
    if not result:
        raise HTTPException(status_code=404, detail="Parcelamento não encontrado.")
    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "result": result}


@app.patch("/installments/{user_id}/{group_id}")
async def installment_update_route(
    request: Request, user_id: int, group_id: str,
    payload: InstallmentUpdatePayload,
):
    """Edita nome (nota) e/ou categoria de todas as parcelas do grupo."""
    _authorize_dashboard_access(request, user_id)
    from db.cards import update_installment_group_meta

    if payload.nome is not None:
        n = (payload.nome or "").strip()
        if len(n) > 200:
            raise HTTPException(status_code=400, detail="Nome muito longo (máx. 200 caracteres).")
    if payload.categoria is not None:
        c = (payload.categoria or "").strip()
        if len(c) > 80:
            raise HTTPException(status_code=400, detail="Categoria muito longa (máx. 80 caracteres).")

    updated = await asyncio.to_thread(
        update_installment_group_meta, user_id, group_id,
        nome=payload.nome, categoria=payload.categoria,
    )
    if not updated:
        raise HTTPException(status_code=404, detail="Parcelamento não encontrado ou nenhum campo alterado.")
    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True}


def _months_pt():
    return [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
    ]


def _serialize_bill(row: dict) -> dict:
    total = float(row.get("total") or 0)
    paid = float(row.get("paid_amount") or 0)
    due = max(0.0, total - paid)
    pe = row.get("period_end")
    label = ""
    if pe:
        label = f"{_months_pt()[pe.month - 1]}/{pe.year}"
    return {
        "id": int(row["id"]),
        "card_id": int(row["card_id"]) if row.get("card_id") is not None else None,
        "card_name": row.get("card_name") or "",
        "period_start": pe and row.get("period_start").isoformat() if row.get("period_start") else None,
        "period_end": pe.isoformat() if pe else None,
        "label": label,
        "status": row.get("status") or "open",
        "total": total,
        "paid_amount": paid,
        "due_amount": due,
    }


@app.get("/bills/{user_id}")
async def list_bills_route(
    request: Request,
    user_id: int,
    card_id: int | None = None,
    include_closed: bool = False,
):
    """Lista faturas do usuário (cartão e valores).

    - Sem params: só abertas com saldo > 0 (default histórico, usado por
      `onCardRowClick` e modal de pagamento).
    - `card_id`: filtra por um cartão específico.
    - `include_closed=true`: inclui também `paid` e `closed` (usado pelas
      setas de navegação no modal de fatura pra ver meses passados/
      próximos cheios).
    """
    _authorize_dashboard_access(request, user_id)
    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))

    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            status_filter = "" if include_closed else "AND b.status = 'open'"
            card_filter = "AND b.card_id = %s" if card_id else ""
            params: list = [int(user_id)]
            if card_id:
                params.append(int(card_id))
            await cur.execute(
                f"""
                SELECT b.id, b.card_id, c.name AS card_name,
                       b.period_start, b.period_end, b.status,
                       b.total, COALESCE(b.paid_amount, 0) AS paid_amount
                FROM credit_bills b
                JOIN credit_cards c ON c.id = b.card_id
                WHERE b.user_id = %s
                  {status_filter}
                  {card_filter}
                ORDER BY b.period_end ASC, c.name ASC
                """,
                params,
            )
            rows = await cur.fetchall()

            await cur.execute(
                "SELECT balance FROM accounts WHERE user_id=%s", (int(user_id),)
            )
            bal_row = await cur.fetchone()

    bills = [_serialize_bill(dict(r)) for r in (rows or [])]
    if not include_closed:
        # Comportamento original — só esconde bills "vazias" no fluxo padrão.
        bills = [b for b in bills if b["due_amount"] > 0 or b["total"] > 0]

    balance = float(bal_row["balance"]) if bal_row else 0.0
    return {"ok": True, "balance": balance, "bills": bills}


@app.get("/bills/{user_id}/{bill_id}")
async def get_bill_detail_route(request: Request, user_id: int, bill_id: int):
    """Detalhe da fatura: período, totais e lista de transações."""
    _authorize_dashboard_access(request, user_id)
    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                """
                SELECT b.id, b.card_id, c.name AS card_name,
                       b.period_start, b.period_end, b.status,
                       b.total, COALESCE(b.paid_amount, 0) AS paid_amount
                FROM credit_bills b
                JOIN credit_cards c ON c.id = b.card_id
                WHERE b.user_id=%s AND b.id=%s
                LIMIT 1
                """,
                (int(user_id), int(bill_id)),
            )
            bill_row = await cur.fetchone()
            if not bill_row:
                raise HTTPException(status_code=404, detail="Fatura não encontrada.")

            await cur.execute(
                """
                SELECT id, valor, categoria, nota, purchased_at, is_refund,
                       installment_no, installments_total
                FROM credit_transactions
                WHERE user_id=%s AND bill_id=%s
                ORDER BY purchased_at DESC, id DESC
                """,
                (int(user_id), int(bill_id)),
            )
            tx_rows = await cur.fetchall()

    transactions = []
    for t in (tx_rows or []):
        transactions.append({
            "id": int(t["id"]),
            "valor": float(t["valor"] or 0),
            "categoria": t.get("categoria"),
            "nota": t.get("nota"),
            "purchased_at": t["purchased_at"].isoformat() if t.get("purchased_at") else None,
            "is_refund": bool(t.get("is_refund")),
            "installment_no": t.get("installment_no"),
            "installments_total": t.get("installments_total"),
        })

    return {"ok": True, "bill": _serialize_bill(dict(bill_row)), "transactions": transactions}


class PayBillPayload(BaseModel):
    amount: float | None = None  # None = paga total em aberto


@app.post("/bills/{user_id}/{bill_id}/pay")
async def pay_bill_route(
    request: Request,
    user_id: int,
    bill_id: int,
    payload: PayBillPayload,
):
    """Paga (parcial ou total) uma fatura. Bloqueia se saldo insuficiente.
    Reusa `pay_bill_amount` (mesmo fluxo do `pagar fatura` do WhatsApp)."""
    _authorize_dashboard_access(request, user_id)
    import sys
    sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
    from db import pay_bill_amount

    # Carrega fatura + saldo
    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                """
                SELECT b.id, b.card_id, c.name AS card_name,
                       b.period_end, b.total,
                       COALESCE(b.paid_amount, 0) AS paid_amount,
                       b.status
                FROM credit_bills b
                JOIN credit_cards c ON c.id = b.card_id
                WHERE b.user_id=%s AND b.id=%s
                LIMIT 1
                """,
                (int(user_id), int(bill_id)),
            )
            bill = await cur.fetchone()
            if not bill:
                raise HTTPException(status_code=404, detail="Fatura não encontrada.")

            await cur.execute(
                "SELECT balance FROM accounts WHERE user_id=%s", (int(user_id),)
            )
            acc = await cur.fetchone()
    balance = float(acc["balance"]) if acc else 0.0

    total = float(bill["total"] or 0)
    paid = float(bill["paid_amount"] or 0)
    due = max(0.0, total - paid)
    if due <= 0:
        raise HTTPException(status_code=400, detail="Esta fatura já está paga.")

    requested_amount = payload.amount
    if requested_amount is None:
        amount = due
    else:
        try:
            amount = float(requested_amount)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Valor inválido.")
        if amount <= 0:
            raise HTTPException(status_code=400, detail="O valor deve ser maior que zero.")
        if amount > due + 0.005:
            raise HTTPException(
                status_code=400,
                detail=f"Valor maior que o em aberto ({_fmt_money(due)})."
                if False else f"Valor maior que o em aberto. Em aberto: R$ {due:.2f}",
            )

    # Saldo insuficiente bloqueia
    if balance < amount - 0.005:
        raise HTTPException(
            status_code=400,
            detail=f"Saldo insuficiente. Saldo atual: R$ {balance:.2f}, valor pedido: R$ {amount:.2f}.",
        )

    card_name = bill["card_name"] or "cartão"
    try:
        res = await asyncio.to_thread(
            pay_bill_amount, int(user_id), int(bill["card_id"]), card_name, float(amount), int(bill_id)
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Erro ao pagar fatura: {exc}") from exc

    if isinstance(res, dict) and res.get("error") == "amount_too_high":
        raise HTTPException(status_code=400, detail="Valor maior que o em aberto.")
    if isinstance(res, dict) and res.get("error") == "invalid_amount":
        raise HTTPException(status_code=400, detail="Valor inválido.")
    if not res:
        raise HTTPException(status_code=400, detail="Nada para pagar nessa fatura.")

    # Re-lê o estado da fatura pós-pagamento
    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT status, total, COALESCE(paid_amount,0) AS paid_amount FROM credit_bills WHERE id=%s",
                (int(bill_id),),
            )
            after = await cur.fetchone()
    new_total = float(after["total"] or 0) if after else total
    new_paid = float(after["paid_amount"] or 0) if after else paid + amount
    new_due = max(0.0, new_total - new_paid)

    return {
        "ok": True,
        "paid": float(res.get("paid", amount)),
        "launch_id": int(res.get("launch_id")) if res.get("launch_id") is not None else None,
        "new_balance": float(res.get("new_balance", balance - amount)),
        "card_name": card_name,
        "bill_id": int(bill_id),
        "bill_status": (after or {}).get("status") or "open",
        "bill_total": new_total,
        "bill_paid_amount": new_paid,
        "bill_due_amount": new_due,
    }


# ─── Analytics routes (Sprint 6) ─────────────────────────────────────────────
# 5 endpoints separados pra alimentar a view Análises. Separados (em vez de
# 1 unificado) porque o painel personalizável vai deixar o user escolher quais
# widgets ver — assim cada widget pode buscar só seu dado.

def _parse_date_param(value: str | None, name: str) -> date | None:
    """Parsea 'YYYY-MM-DD' → date. None se vazio. 400 se inválido."""
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Parâmetro '{name}' inválido (esperado YYYY-MM-DD).")


def _resolve_analytics_window(months: int, from_str: str | None, to_str: str | None):
    """Wrapper de resolve_window que parseia strings de query."""
    from db import resolve_window
    fd = _parse_date_param(from_str, "from")
    td = _parse_date_param(to_str, "to")
    return resolve_window(months=months, from_date=fd, to_date=td)


@app.get("/analytics/{user_id}/kpis")
async def analytics_kpis_route(
    request: Request,
    user_id: int,
    months: int = 6,
    from_: str | None = Query(None, alias="from"),
    to: str | None = None,
):
    """KPIs do período: receita, despesa, líquido, taxa de poupança + comparativo
    com período anterior. Default = últimos 6 meses."""
    _authorize_dashboard_access(request, user_id)
    from db import compute_kpis
    fd, td = _resolve_analytics_window(months, from_, to)
    result = await asyncio.to_thread(compute_kpis, user_id, fd, td)
    return {"ok": True, "kpis": result}


@app.get("/analytics/{user_id}/evolution")
async def analytics_evolution_route(
    request: Request,
    user_id: int,
    months: int = 6,
):
    """Evolução mensal de receita/despesa/líquido nos últimos N meses (sempre
    buckets mensais terminando no mês atual)."""
    _authorize_dashboard_access(request, user_id)
    from db import compute_evolution
    n = max(1, min(int(months or 6), 36))
    result = await asyncio.to_thread(compute_evolution, user_id, n)
    return {"ok": True, "evolution": result, "months": n}


@app.get("/analytics/{user_id}/categories")
async def analytics_categories_route(
    request: Request,
    user_id: int,
    months: int = 6,
    from_: str | None = Query(None, alias="from"),
    to: str | None = None,
    limit: int = 10,
):
    """Distribuição de despesas por categoria no período. Inclui pct, count,
    e emoji/color quando user customizou via user_categories."""
    _authorize_dashboard_access(request, user_id)
    from db import compute_categories
    fd, td = _resolve_analytics_window(months, from_, to)
    lim = max(1, min(int(limit or 10), 50))
    result = await asyncio.to_thread(compute_categories, user_id, fd, td, lim)
    return {"ok": True, "categories": result, "window": {"from": fd.isoformat(), "to": td.isoformat()}}


@app.get("/analytics/{user_id}/weekday-pattern")
async def analytics_weekday_route(
    request: Request,
    user_id: int,
    months: int = 6,
    from_: str | None = Query(None, alias="from"),
    to: str | None = None,
):
    """Padrão de gasto por dia da semana (seg→dom). Total, count e média
    diária por DOW no período. Inclui credit_transactions por purchased_at
    (não bill.period_end — aqui interessa o dia da compra real)."""
    _authorize_dashboard_access(request, user_id)
    from db import compute_weekday_pattern
    fd, td = _resolve_analytics_window(months, from_, to)
    result = await asyncio.to_thread(compute_weekday_pattern, user_id, fd, td)
    return {"ok": True, "weekdays": result, "window": {"from": fd.isoformat(), "to": td.isoformat()}}


@app.get("/analytics/{user_id}/top-merchants")
async def analytics_top_merchants_route(
    request: Request,
    user_id: int,
    months: int = 6,
    from_: str | None = Query(None, alias="from"),
    to: str | None = None,
    limit: int = 10,
):
    """Top estabelecimentos no período. Junta launches (alvo/nota) +
    credit_transactions (nota). Agrupa por chave normalizada (lower/trim).
    Retorna sources = {debito, credito} pra desbobinar na UI se quiser."""
    _authorize_dashboard_access(request, user_id)
    from db import compute_top_merchants
    fd, td = _resolve_analytics_window(months, from_, to)
    lim = max(1, min(int(limit or 10), 50))
    result = await asyncio.to_thread(compute_top_merchants, user_id, fd, td, lim)
    return {"ok": True, "merchants": result, "window": {"from": fd.isoformat(), "to": td.isoformat()}}


# ─── Sprint 7: Insights proativos + padrões comportamentais (IA) ─────────────

@app.get("/insights/{user_id}/current")
async def insights_current_route(request: Request, user_id: int, force: bool = False):
    """Insights acionáveis do Piggy, gerados via LLM (gpt-4o-mini).

    Recebe estado financeiro ATUAL (orçamentos, recorrentes, metas, KPIs) e
    devolve 3-5 insights priorizados por severidade. Cache 6h por user.

    Fallback: se OPENAI_API_KEY ausente ou LLM falha, usa heurística antiga
    (`compute_active_insights`) pra não deixar o card vazio.

    `?force=true` ignora cache (útil só pra debug).
    """
    _authorize_dashboard_access(request, user_id)
    from core.ai_patterns import generate_ai_insights
    result = await asyncio.to_thread(generate_ai_insights, user_id, force=force)
    return {"ok": True, "insights": result or []}


@app.get("/analytics/{user_id}/patterns")
async def analytics_patterns_route(
    request: Request,
    user_id: int,
    force: bool = False,
):
    """Padrões comportamentais via LLM (gpt-4o-mini). Cache 24h.

    O LLM recebe métricas agregadas (gastos por hora, weekend split, salary
    burn, top merchants, top categorias) e devolve narrativas descobertas
    dinamicamente (variam por user). Retorna lista de items `{icon, title,
    subtitle, tone}` no campo `patterns`.

    Fallback: se LLM indisponível, retorna lista vazia (frontend mostra
    empty state). NÃO retorna a agregação bruta — só formato narrativo.

    `?force=true` ignora cache (útil só pra debug).
    """
    _authorize_dashboard_access(request, user_id)
    from core.ai_patterns import generate_ai_patterns
    result = await asyncio.to_thread(generate_ai_patterns, user_id, force=force)
    return {"ok": True, "patterns": result or []}


@app.get("/debug/ai/{user_id}/payload")
async def debug_ai_payload_route(
    request: Request,
    user_id: int,
    kind: str = "patterns",
):
    """[DEBUG] Retorna o JSON EXATO que vai pro LLM, sem chamar a LLM.

    Serve pra diagnosticar alucinações: se o LLM cita "delivery" mas o
    payload não tem nada de delivery, sabemos que é invenção.

    `kind` ∈ {'patterns', 'insights'}. Protegido pelo mesmo auth do dashboard.
    """
    _authorize_dashboard_access(request, user_id)
    from core.ai_patterns import _collect_patterns_data, _collect_insights_data
    if kind == "insights":
        data = await asyncio.to_thread(_collect_insights_data, user_id)
    else:
        data = await asyncio.to_thread(_collect_patterns_data, user_id)
    return {"ok": True, "kind": kind, "payload": data}


# ─── History route (Sprint 6) ────────────────────────────────────────────────

@app.get("/history/{user_id}/list")
async def history_list_route(
    request: Request,
    user_id: int,
    from_: str | None = Query(None, alias="from"),
    to: str | None = None,
    categoria: str | None = None,
    tipo: str | None = None,
    q: str | None = None,
    uncategorized: bool = False,
    refunds_only: bool = False,
    page: int = 1,
    limit: int = 50,
):
    """Timeline paginada de lançamentos. Junta launches + credit_transactions
    (alocadas por bill.period_end). Filtros: faixa de datas, categoria, tipo,
    busca textual livre (q — AND entre palavras, OR entre campos alvo/nota/
    categoria/card-name), uncategorized (só sem categoria), refunds_only
    (só estornos)."""
    _authorize_dashboard_access(request, user_id)
    from db import list_history
    fd = _parse_date_param(from_, "from")
    td = _parse_date_param(to, "to")
    result = await asyncio.to_thread(
        list_history,
        user_id, fd, td, categoria, tipo, q,
        bool(uncategorized), bool(refunds_only),
        page, limit,
    )
    return {"ok": True, **result}


@app.get("/history/{user_id}/quick-stats")
async def history_quick_stats_route(
    request: Request,
    user_id: int,
    months: int = 6,
    from_: str | None = Query(None, alias="from"),
    to: str | None = None,
):
    """Atalhos de filtro pro topo da view Histórico:
      - uncategorized_count: lançamentos sem categoria no período
      - largest_expense: o maior lançamento individual
      - refunds_count: estornos no cartão no período
      - recent_7d_count: atividade dos últimos 7 dias (sempre 7d,
        independente do período principal)

    Cada um vira um card clicável que aplica filtro correspondente."""
    _authorize_dashboard_access(request, user_id)
    from db import compute_history_quick_stats
    fd, td = _resolve_analytics_window(months, from_, to)
    result = await asyncio.to_thread(compute_history_quick_stats, user_id, fd, td)
    return {"ok": True, **result, "window": {"from": fd.isoformat(), "to": td.isoformat()}}


MAX_OFX_BYTES = 8 * 1024 * 1024  # 8 MB — extratos OFX raramente passam disso


@app.post("/ofx/import/{user_id}")
@limiter.limit("5/hour")
async def ofx_import_route(request: Request, user_id: int):
    """
    Upload de arquivo OFX via dashboard. Aceita multipart/form-data com campo
    `file`. Detecta automaticamente extrato bancario vs fatura de cartao e
    roteia pro service correto. Pro-only.
    """
    _authorize_dashboard_access(request, user_id)
    _require_pro(user_id, "ofx_import")

    form = await request.form()
    upload = form.get("file")
    if upload is None or not hasattr(upload, "read"):
        raise HTTPException(status_code=400, detail="Arquivo OFX nao enviado (campo 'file' ausente).")

    filename = (getattr(upload, "filename", "") or "arquivo.ofx").strip()
    if not filename.lower().endswith(".ofx"):
        raise HTTPException(status_code=400, detail="Arquivo precisa ter extensao .ofx.")

    # Allowlist de content-type: browsers mapeiam .ofx de formas variadas
    # (octet-stream, text/plain, xml...), então a lista é generosa — o objetivo
    # é só barrar spoof óbvio (image/*, video/*, application/pdf etc).
    content_type = (getattr(upload, "content_type", "") or "").lower().split(";")[0].strip()
    if content_type not in {
        "", "application/x-ofx", "text/x-ofx", "application/ofx",
        "text/plain", "application/octet-stream", "text/xml", "application/xml",
    }:
        raise HTTPException(status_code=400, detail="Content-Type inválido para arquivo OFX.")

    raw = await upload.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Arquivo OFX vazio.")
    if len(raw) > MAX_OFX_BYTES:
        raise HTTPException(status_code=413, detail=f"Arquivo grande demais (max {MAX_OFX_BYTES // (1024*1024)} MB).")

    from ofx_import import detect_ofx_type
    from core.services.ofx_service import handle_ofx_import, handle_credit_ofx_import

    ofx_type = detect_ofx_type(raw)
    if ofx_type == "credit_card":
        message = await asyncio.to_thread(handle_credit_ofx_import, str(user_id), raw, filename)
    elif ofx_type == "bank":
        message = await asyncio.to_thread(handle_ofx_import, str(user_id), raw, filename)
    else:
        raise HTTPException(
            status_code=400,
            detail="Nao consegui identificar o tipo de OFX (extrato bancario ou fatura de cartao).",
        )

    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "type": ofx_type, "message": message}


def _mask_email(email: str) -> str:
    local, _, domain = (email or "").partition("@")
    if not domain:
        return "seu email"
    return f"{local[:1]}***@{domain}"


@app.post("/export/{user_id}")
@limiter.limit("3/minute")
async def export_email(request: Request, user_id: int, year: int = None, month: int = None):
    """Gera o extrato do mês (PDF + XLSX + CSV) e envia pro email cadastrado."""
    _authorize_dashboard_access(request, user_id)
    _require_pro(user_id, "export")
    now = datetime.now(timezone.utc)
    y = year  or now.year
    m = month or now.month

    csv_content = await build_csv(user_id, y, m)
    if csv_content is None:
        raise HTTPException(
            status_code=404,
            detail="Nenhum lançamento encontrado neste mês para exportar.",
        )
    xlsx_bytes = await build_xlsx(user_id, y, m)
    pdf_bytes  = await build_pdf(user_id, y, m)

    from db.privacy import get_user_email
    to_email = await asyncio.to_thread(get_user_email, user_id)
    if not to_email:
        raise HTTPException(
            status_code=400,
            detail="Você não tem um email cadastrado para receber o extrato.",
        )

    import base64
    tag = f"{y:04d}_{m:02d}"
    attachments = [
        {"filename": f"extrato_{tag}.pdf",   "content": base64.b64encode(pdf_bytes).decode(),
         "content_type": "application/pdf"},
        {"filename": f"financas_{tag}.xlsx", "content": base64.b64encode(xlsx_bytes).decode(),
         "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"filename": f"financas_{tag}.csv",  "content": base64.b64encode(csv_content.encode("utf-8")).decode(),
         "content_type": "text/csv"},
    ]

    mes_label = f"{_MESES_PT[m]} de {y}"
    subject = f"Seu extrato PigBank — {mes_label}"
    from core.services.email_service import send_email, _base_html
    inner = (
        "<p>Oi! 🐷</p>"
        f"<p>Segue em anexo o seu extrato de <strong>{mes_label}</strong>:</p>"
        "<ul>"
        "<li><strong>PDF</strong> — resumo pra ler ou imprimir</li>"
        "<li><strong>XLSX / CSV</strong> — pra abrir em planilha</li>"
        "</ul>"
        "<p>Qualquer dúvida, fala com a gente em "
        "<a href=\"mailto:suporte@pigbankai.com\">suporte@pigbankai.com</a>.</p>"
    )
    html = _base_html(subject, inner)

    ok = await asyncio.to_thread(
        send_email, to_email, subject, html, attachments=attachments
    )
    if not ok:
        raise HTTPException(
            status_code=502,
            detail="Não consegui enviar o email agora. Tente novamente em instantes.",
        )

    return {"ok": True, "email": _mask_email(to_email)}

# ─── Budget routes ────────────────────────────────────────────────────────────

@app.get("/budgets/{user_id}")
async def get_budgets(request: Request, user_id: int):
    _authorize_dashboard_access(request, user_id)
    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT categoria, budget FROM category_budgets WHERE user_id = %s ORDER BY categoria",
                (user_id,),
            )
            rows = await cur.fetchall()
    return {"budgets": [dict(r) for r in rows]}

class BudgetPayload(BaseModel):
    categoria: str
    budget: float

FREE_BUDGETS_LIMIT = 3


@app.post("/budgets/{user_id}")
async def set_budget(request: Request, user_id: int, payload: BudgetPayload):
    _authorize_dashboard_access(request, user_id)
    if payload.budget <= 0:
        raise HTTPException(status_code=400, detail="budget must be > 0")

    from core.services.plan_service import is_pro

    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            # Pro gate: Free pode ter até FREE_BUDGETS_LIMIT orçamentos.
            # Update de orçamento existente não conta — só novo INSERT.
            if not is_pro(user_id):
                await cur.execute(
                    "SELECT 1 FROM category_budgets "
                    "WHERE user_id=%s AND lower(categoria)=lower(%s)",
                    (user_id, payload.categoria),
                )
                is_existing = await cur.fetchone() is not None
                if not is_existing:
                    await cur.execute(
                        "SELECT COUNT(*) AS n FROM category_budgets WHERE user_id=%s",
                        (user_id,),
                    )
                    row = await cur.fetchone()
                    current = int((row.get("n") if isinstance(row, dict) else row[0]) or 0)
                    if current >= FREE_BUDGETS_LIMIT:
                        raise HTTPException(
                            status_code=403,
                            detail={
                                "error": "pro_required",
                                "feature": "budgets_limit",
                                "limit": FREE_BUDGETS_LIMIT,
                            },
                        )

            await cur.execute(
                """
                INSERT INTO category_budgets (user_id, categoria, budget)
                VALUES (%s, %s, %s)
                ON CONFLICT (user_id, categoria)
                DO UPDATE SET budget = EXCLUDED.budget
                """,
                (user_id, payload.categoria, payload.budget),
            )
        await conn.commit()
    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "categoria": payload.categoria, "budget": payload.budget}

@app.delete("/budgets/{user_id}/{categoria}")
async def delete_budget(request: Request, user_id: int, categoria: str):
    _authorize_dashboard_access(request, user_id)
    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "DELETE FROM category_budgets WHERE user_id = %s AND categoria = %s",
                (user_id, categoria),
            )
        await conn.commit()
    return {"ok": True}


@app.get("/budgets/{user_id}/status")
async def budgets_status_route(request: Request, user_id: int, month: str | None = None):
    """Semáforo de orçamentos do mês: gasto vs limite com cor por categoria.

    `month` no formato 'YYYY-MM'. Default = mês corrente.
    Resposta inclui emoji/cor da categoria (joins com user_categories).
    """
    _authorize_dashboard_access(request, user_id)
    from db.budgets import get_budgets_status_for_month

    status = await asyncio.to_thread(get_budgets_status_for_month, user_id, month)
    return {"ok": True, **status}


# ─── Category metadata routes (Sprint 3) ─────────────────────────────────────

class CategoryCreatePayload(BaseModel):
    name: str
    emoji: str | None = None
    color: str | None = None


class CategoryUpdatePayload(BaseModel):
    name: str | None = None
    emoji: str | None = None
    color: str | None = None


@app.get("/categories/{user_id}")
async def categories_list_route(
    request: Request, user_id: int, include_archived: bool = True
):
    """Lista categorias do user (metadata visual). Faz seed lazy das 14 canônicas."""
    _authorize_dashboard_access(request, user_id)
    from db.categories import list_user_categories_full

    cats = await asyncio.to_thread(list_user_categories_full, user_id, include_archived)
    return {"ok": True, "categories": cats}


@app.post("/categories/{user_id}")
async def categories_create_route(
    request: Request, user_id: int, payload: CategoryCreatePayload
):
    """Cria categoria custom. Pro-only — Free fica com as 14 canônicas."""
    _authorize_dashboard_access(request, user_id)
    _require_pro(user_id, "custom_categories")
    from db.categories import create_user_category

    try:
        cat = await asyncio.to_thread(
            create_user_category, user_id, payload.name, payload.emoji, payload.color
        )
    except ValueError as exc:
        code = str(exc)
        if code == "CATEGORIA_DUPLICADA":
            raise HTTPException(status_code=409, detail="Já existe uma categoria com esse nome.")
        if code == "CATEGORIA_INVALIDA":
            raise HTTPException(status_code=400, detail="Nome de categoria inválido.")
        raise HTTPException(status_code=400, detail=code)
    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "category": cat}


@app.patch("/categories/{user_id}/{cat_id}")
async def categories_update_route(
    request: Request, user_id: int, cat_id: int, payload: CategoryUpdatePayload
):
    """Edita nome (faz UPDATE em cascata em launches/cards/etc), emoji ou cor."""
    _authorize_dashboard_access(request, user_id)
    from db.categories import update_user_category

    try:
        cat = await asyncio.to_thread(
            update_user_category,
            user_id, cat_id,
            new_name=payload.name, emoji=payload.emoji, color=payload.color,
        )
    except ValueError as exc:
        code = str(exc)
        if code == "CATEGORIA_NAO_ENCONTRADA":
            raise HTTPException(status_code=404, detail="Categoria não encontrada.")
        if code == "CATEGORIA_DUPLICADA":
            raise HTTPException(status_code=409, detail="Já existe uma categoria com esse nome.")
        if code == "CATEGORIA_INVALIDA":
            raise HTTPException(status_code=400, detail="Nome de categoria inválido.")
        raise HTTPException(status_code=400, detail=code)
    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "category": cat}


@app.post("/categories/{user_id}/{cat_id}/archive")
async def categories_archive_route(request: Request, user_id: int, cat_id: int):
    """Arquiva categoria: some dos dropdowns mas continua no histórico."""
    _authorize_dashboard_access(request, user_id)
    from db.categories import set_user_category_archived

    try:
        cat = await asyncio.to_thread(set_user_category_archived, user_id, cat_id, True)
    except ValueError:
        raise HTTPException(status_code=404, detail="Categoria não encontrada.")
    return {"ok": True, "category": cat}


@app.post("/categories/{user_id}/{cat_id}/unarchive")
async def categories_unarchive_route(request: Request, user_id: int, cat_id: int):
    _authorize_dashboard_access(request, user_id)
    from db.categories import set_user_category_archived

    try:
        cat = await asyncio.to_thread(set_user_category_archived, user_id, cat_id, False)
    except ValueError:
        raise HTTPException(status_code=404, detail="Categoria não encontrada.")
    return {"ok": True, "category": cat}


@app.delete("/categories/{user_id}/{cat_id}")
async def categories_delete_route(request: Request, user_id: int, cat_id: int):
    """Deleta categoria. Bloqueado se: (1) é system, (2) tem lançamentos. Sugerir arquivar."""
    _authorize_dashboard_access(request, user_id)
    from db.categories import delete_user_category

    try:
        await asyncio.to_thread(delete_user_category, user_id, cat_id)
    except ValueError as exc:
        code = str(exc)
        if code == "CATEGORIA_NAO_ENCONTRADA":
            raise HTTPException(status_code=404, detail="Categoria não encontrada.")
        if code == "CATEGORIA_SISTEMA_INDELETAVEL":
            raise HTTPException(
                status_code=400,
                detail="Categorias do sistema só podem ser arquivadas, não excluídas.",
            )
        if code == "CATEGORIA_COM_LANCAMENTOS":
            raise HTTPException(
                status_code=400,
                detail="Essa categoria tem lançamentos. Arquive em vez de excluir.",
            )
        raise HTTPException(status_code=400, detail=code)
    return {"ok": True}


# ─── Account setup routes (Sprint Wizard) ────────────────────────────────────

class InitialBalancePayload(BaseModel):
    amount: float


@app.get("/account/{user_id}/setup-status")
async def setup_status_route(request: Request, user_id: int):
    """Retorna se a conta está virgem (balance=0, nenhum launch, nenhum cartão).
    Usado pelo Wizard de Setup pra decidir se mostra automático."""
    _authorize_dashboard_access(request, user_id)

    def _query():
        from db.connection import get_conn
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT balance FROM accounts WHERE user_id=%s", (user_id,))
                row = cur.fetchone()
                balance = float(row["balance"]) if row else 0.0
                cur.execute("SELECT COUNT(*) AS n FROM launches WHERE user_id=%s", (user_id,))
                n_launches = int(cur.fetchone()["n"] or 0)
                cur.execute("SELECT COUNT(*) AS n FROM credit_cards WHERE user_id=%s", (user_id,))
                n_cards = int(cur.fetchone()["n"] or 0)
        return {"balance": balance, "n_launches": n_launches, "n_cards": n_cards}

    data = await asyncio.to_thread(_query)
    is_virgin = data["balance"] == 0 and data["n_launches"] == 0 and data["n_cards"] == 0
    return {"ok": True, "is_virgin": is_virgin, **data}


@app.post("/account/{user_id}/initial-balance")
async def set_initial_balance_route(request: Request, user_id: int, payload: InitialBalancePayload):
    """Cria lançamento 'Saldo inicial' (receita interna) que atualiza accounts.balance.

    Bloqueia se balance != 0 (use ajuste manual depois). Idempotente no sentido de
    que se chamar com amount=0 não cria nada.
    """
    _authorize_dashboard_access(request, user_id)
    if payload.amount < 0:
        raise HTTPException(status_code=400, detail="Saldo não pode ser negativo.")

    from db.accounts import get_balance, add_launch_and_update_balance

    current = await asyncio.to_thread(get_balance, user_id)
    if float(current) != 0:
        raise HTTPException(
            status_code=409,
            detail="Conta já tem saldo. Saldo inicial só pode ser definido em conta virgem.",
        )
    if payload.amount == 0:
        return {"ok": True, "balance": 0.0, "launch_id": None}

    launch_id, _user_seq, new_bal = await asyncio.to_thread(
        add_launch_and_update_balance,
        user_id, "receita", float(payload.amount), "setup", "Saldo inicial",
        categoria="saldo_inicial", is_internal_movement=True,
    )
    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "balance": float(new_bal), "launch_id": launch_id}


class AdjustBalancePayload(BaseModel):
    target_balance: float


@app.post("/account/{user_id}/adjust-balance")
async def adjust_balance_route(request: Request, user_id: int, payload: AdjustBalancePayload):
    """Ajusta saldo via lançamento de delta.

    Calcula `target_balance - saldo_atual` e cria um launch interno
    ('ajuste de saldo') com o valor da diferença. Saldo é atualizado pela
    própria função de lançamento. Mantém rastreabilidade — vê no histórico.

    Se delta == 0, não cria nada (idempotente).
    """
    _authorize_dashboard_access(request, user_id)
    from db.accounts import get_balance, add_launch_and_update_balance

    current = await asyncio.to_thread(get_balance, user_id)
    delta = float(payload.target_balance) - float(current)
    if abs(delta) < 0.005:
        return {"ok": True, "balance": float(current), "delta": 0.0, "launch_id": None}

    tipo = "receita" if delta > 0 else "despesa"
    valor = abs(delta)
    nota = "Ajuste de saldo manual"

    launch_id, _seq, new_bal = await asyncio.to_thread(
        add_launch_and_update_balance,
        user_id, tipo, valor, "ajuste", nota,
        categoria="ajuste", is_internal_movement=True,
    )
    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "balance": float(new_bal), "delta": delta, "launch_id": launch_id}


# ─── Recurring expenses / Gastos Fixos (Sprint 4) ────────────────────────────

class RecurringCreatePayload(BaseModel):
    name: str
    amount: float
    category: str
    due_day: int
    payment_type: str
    card_id: int | None = None
    is_essential: bool = False
    notes: str | None = None


class RecurringUpdatePayload(BaseModel):
    name: str | None = None
    amount: float | None = None
    category: str | None = None
    due_day: int | None = None
    payment_type: str | None = None
    card_id: int | None = None
    is_essential: bool | None = None
    is_active: bool | None = None
    notes: str | None = None


def _recurring_value_error(code: str) -> HTTPException:
    msg = {
        "NOME_INVALIDO": "Nome inválido.",
        "VALOR_INVALIDO": "Valor deve ser maior que zero.",
        "DIA_INVALIDO": "Dia do vencimento deve estar entre 1 e 31.",
        "FORMA_PAGAMENTO_INVALIDA": "Forma de pagamento inválida (use 'account' ou 'credit_card').",
        "CARTAO_OBRIGATORIO": "Cartão é obrigatório quando a forma de pagamento é cartão de crédito.",
        "CARTAO_NAO_ENCONTRADO": "Cartão não encontrado.",
        "RECORRENTE_NAO_ENCONTRADO": "Gasto fixo não encontrado.",
    }
    return HTTPException(status_code=400, detail=msg.get(code, code))


@app.get("/recurring-expenses/{user_id}")
async def recurring_list_route(request: Request, user_id: int, include_inactive: bool = False):
    """Lista os gastos fixos do user. Pro-only."""
    _authorize_dashboard_access(request, user_id)
    _require_pro(user_id, "recurring_expenses")

    from db.recurring import list_recurring_expenses

    items = await asyncio.to_thread(list_recurring_expenses, user_id, include_inactive)
    return {"ok": True, "recurring": items}


@app.post("/recurring-expenses/{user_id}")
async def recurring_create_route(request: Request, user_id: int, payload: RecurringCreatePayload):
    """Cria um gasto fixo. Pro-only."""
    _authorize_dashboard_access(request, user_id)
    _require_pro(user_id, "recurring_expenses")

    from db.recurring import create_recurring_expense

    try:
        item = await asyncio.to_thread(
            create_recurring_expense,
            user_id, payload.name, payload.amount, payload.category, payload.due_day,
            payload.payment_type, payload.card_id, payload.is_essential, payload.notes,
        )
    except ValueError as exc:
        raise _recurring_value_error(str(exc))
    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "recurring": item}


@app.patch("/recurring-expenses/{user_id}/{rec_id}")
async def recurring_update_route(
    request: Request, user_id: int, rec_id: int, payload: RecurringUpdatePayload
):
    """Edita um gasto fixo. Reajuste detectado quando amount muda."""
    _authorize_dashboard_access(request, user_id)
    _require_pro(user_id, "recurring_expenses")

    from db.recurring import update_recurring_expense

    try:
        item = await asyncio.to_thread(
            update_recurring_expense,
            user_id, rec_id,
            name=payload.name, amount=payload.amount, category=payload.category,
            due_day=payload.due_day, payment_type=payload.payment_type, card_id=payload.card_id,
            is_essential=payload.is_essential, is_active=payload.is_active, notes=payload.notes,
        )
    except ValueError as exc:
        code = str(exc)
        if code == "RECORRENTE_NAO_ENCONTRADO":
            raise HTTPException(status_code=404, detail="Gasto fixo não encontrado.")
        raise _recurring_value_error(code)
    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "recurring": item}


@app.post("/recurring-expenses/{user_id}/charges/{charge_id}/ack")
async def recurring_charge_ack_route(request: Request, user_id: int, charge_id: int):
    """Marca uma cobrança automática como vista (some do banner)."""
    _authorize_dashboard_access(request, user_id)
    _require_pro(user_id, "recurring_expenses")
    async with await db_connect() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "UPDATE recurring_charges SET acknowledged=true "
                "WHERE id=%s AND user_id=%s",
                (charge_id, user_id),
            )
        await conn.commit()
    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True}


@app.delete("/recurring-expenses/{user_id}/{rec_id}")
async def recurring_delete_route(request: Request, user_id: int, rec_id: int):
    """Exclui um gasto fixo. Lançamentos passados ficam intactos."""
    _authorize_dashboard_access(request, user_id)
    _require_pro(user_id, "recurring_expenses")

    from db.recurring import delete_recurring_expense

    try:
        await asyncio.to_thread(delete_recurring_expense, user_id, rec_id)
    except ValueError as exc:
        if str(exc) == "RECORRENTE_NAO_ENCONTRADO":
            raise HTTPException(status_code=404, detail="Gasto fixo não encontrado.")
        raise _recurring_value_error(str(exc))
    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True}


# ─── Investment routes ───────────────────────────────────────────────────────

class InvestmentCreatePayload(BaseModel):
    name: str
    rate: float
    period: str
    initial_amount: float | None = None
    asset_type: str | None = None
    indexer: str | None = None
    issuer: str | None = None
    purchase_date: date | None = None
    maturity_date: date | None = None
    interest_payment_frequency: str | None = None
    tax_profile: str | None = None
    note: str | None = None


class InvestmentMovementPayload(BaseModel):
    name: str
    amount: float
    note: str | None = None
    # Aporte com taxa específica (Tesouro IPCA+/Prefixado, Debêntures etc.).
    # Quando ausentes, o lote herda taxa/período do investimento.
    rate: float | None = None
    period: str | None = None
    purchase_date: date | None = None


class OpenFinanceMockConnectPayload(BaseModel):
    institution: str | None = None


class OpenFinancePluggyItemPayload(BaseModel):
    item: dict


class NotificationSettingsPayload(BaseModel):
    engagement_email_enabled: bool | None = None
    tip_email_enabled: bool | None = None
    insight_email_enabled: bool | None = None
    whatsapp_updates_enabled: bool | None = None
    daily_report_enabled: bool | None = None
    daily_report_hour: int | None = None
    daily_report_minute: int | None = None


def _investment_action_note(action: str, name: str, issuer: str | None = None, note: str | None = None) -> str:
    clean_name = (name or "").strip() or "investimento"
    clean_issuer = (issuer or "").strip()
    clean_note = (note or "").strip()
    suffix = f" ({clean_issuer})" if clean_issuer and clean_issuer.lower() not in clean_name.lower() else ""
    description = f"{action} {clean_name}{suffix}"
    if clean_note:
        description += f" - {clean_note}"
    return description


@app.get("/investments/{user_id}/rates")
async def investment_rates(request: Request, user_id: int):
    _authorize_dashboard_access(request, user_id)
    rates = await asyncio.to_thread(get_dashboard_market_rates)
    return {"market_rates": rates}


@app.post("/investments/{user_id}")
async def create_investment_route(request: Request, user_id: int, payload: InvestmentCreatePayload):
    _authorize_dashboard_access(request, user_id)
    _require_pro(user_id, "investments")

    name = payload.name.strip()
    period = payload.period.strip().lower()
    if not name:
        raise HTTPException(status_code=400, detail="Nome do investimento é obrigatório.")
    if period not in {"daily", "monthly", "yearly", "cdi", "cdi_spread", "ipca_spread", "selic_spread"}:
        raise HTTPException(status_code=400, detail="Indexador inválido.")
    if payload.rate <= 0 and period != "selic_spread":
        raise HTTPException(status_code=400, detail="Taxa deve ser maior que zero.")

    try:
        create_note = _investment_action_note("Criou investimento", name, payload.issuer, payload.note)
        initial_note = _investment_action_note("Aporte inicial em", name)
        launch_id, inv_id, canon = await asyncio.to_thread(
            create_investment_db,
            user_id,
            name,
            payload.rate,
            period,
            create_note,
            asset_type=payload.asset_type,
            indexer=payload.indexer,
            issuer=payload.issuer,
            purchase_date=payload.purchase_date,
            maturity_date=payload.maturity_date,
            interest_payment_frequency=payload.interest_payment_frequency,
            tax_profile=payload.tax_profile,
            initial_amount=payload.initial_amount,
            initial_note=initial_note,
        )
    except Exception as exc:
        message = "Saldo insuficiente na conta para o aporte inicial." if str(exc) == "INSUFFICIENT_ACCOUNT" else str(exc)
        raise HTTPException(status_code=400, detail=message) from exc

    _invalidate_dashboard_current_cache(user_id)
    return {
        "ok": True,
        "created": launch_id is not None,
        "investment": {"id": inv_id, "name": canon, "rate": payload.rate, "period": period},
    }


@app.post("/investments/{user_id}/deposit")
async def deposit_investment_route(request: Request, user_id: int, payload: InvestmentMovementPayload):
    _authorize_dashboard_access(request, user_id)
    if payload.amount <= 0:
        raise HTTPException(status_code=400, detail="Valor deve ser maior que zero.")
    try:
        launch_id, new_acc, new_inv, canon = await asyncio.to_thread(
            investment_deposit_from_account,
            user_id,
            payload.name.strip(),
            payload.amount,
            payload.note or _investment_action_note("Aporte em", payload.name),
            rate=payload.rate,
            period=payload.period,
            purchase_date=payload.purchase_date,
        )
    except LookupError as exc:
        raise HTTPException(status_code=404, detail="Investimento não encontrado.") from exc
    except ValueError as exc:
        if str(exc) == "INSUFFICIENT_ACCOUNT":
            message = "Saldo insuficiente na conta."
        elif str(exc) == "INVALID_RATE":
            message = "Taxa inválida para este aporte."
        elif str(exc) == "INVALID_PERIOD":
            message = "Indexador inválido para este aporte."
        elif str(exc) == "PURCHASE_DATE_FUTURE":
            message = "Data de compra não pode ser futura."
        else:
            message = str(exc)
        raise HTTPException(status_code=400, detail=message) from exc

    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "launch_id": launch_id, "account_balance": new_acc, "investment_balance": new_inv, "name": canon}


@app.post("/investments/{user_id}/withdraw")
async def withdraw_investment_route(request: Request, user_id: int, payload: InvestmentMovementPayload):
    _authorize_dashboard_access(request, user_id)
    if payload.amount <= 0:
        raise HTTPException(status_code=400, detail="Valor deve ser maior que zero.")
    try:
        launch_id, new_acc, new_inv, canon, tax_summary = await asyncio.to_thread(
            investment_withdraw_to_account,
            user_id,
            payload.name.strip(),
            payload.amount,
            payload.note or _investment_action_note("Resgate de", payload.name),
        )
    except LookupError as exc:
        raise HTTPException(status_code=404, detail="Investimento não encontrado.") from exc
    except ValueError as exc:
        message = "Saldo insuficiente no investimento." if str(exc) == "INSUFFICIENT_INVEST" else str(exc)
        raise HTTPException(status_code=400, detail=message) from exc

    _invalidate_dashboard_current_cache(user_id)
    return {
        "ok": True,
        "launch_id": launch_id,
        "account_balance": new_acc,
        "investment_balance": new_inv,
        "name": canon,
        "tax_summary": tax_summary,
    }


@app.delete("/investments/{user_id}/{name:path}")
async def delete_investment_route(request: Request, user_id: int, name: str):
    _authorize_dashboard_access(request, user_id)
    investment_name = urllib.parse.unquote(name).strip()
    try:
        launch_id, canon = await asyncio.to_thread(
            delete_investment,
            user_id,
            investment_name,
            "dashboard:delete",
        )
    except LookupError as exc:
        raise HTTPException(status_code=404, detail="Investimento não encontrado.") from exc
    except ValueError as exc:
        message = "Zere o saldo antes de remover o investimento." if str(exc) == "INV_NOT_ZERO" else str(exc)
        raise HTTPException(status_code=400, detail=message) from exc

    _invalidate_dashboard_current_cache(user_id)
    return {"ok": True, "launch_id": launch_id, "name": canon}


async def _get_notification_settings(user_id: int) -> dict:
    auth_user, daily_prefs = await asyncio.gather(
        asyncio.to_thread(get_auth_user, user_id),
        asyncio.to_thread(get_daily_report_prefs, user_id),
    )
    auth_user = auth_user or {}
    daily_prefs = daily_prefs or {}
    email = auth_user.get("email")
    phone = auth_user.get("phone_e164")
    email_available = bool(email)
    whatsapp_updates_available = bool(phone)
    engagement_opt_out = bool(auth_user.get("engagement_opt_out", False))
    tip_email_enabled = email_available and not engagement_opt_out and not bool(auth_user.get("tip_email_opt_out", False))
    insight_email_enabled = email_available and not engagement_opt_out and not bool(auth_user.get("insight_email_opt_out", False))
    whatsapp_updates_enabled = whatsapp_updates_available and not bool(auth_user.get("whatsapp_updates_opt_out", False))
    return {
        "ok": True,
        "email": email,
        "whatsapp_destination": phone,
        "email_notifications_available": email_available,
        "whatsapp_updates_available": whatsapp_updates_available,
        "engagement_email_enabled": tip_email_enabled or insight_email_enabled,
        "tip_email_enabled": tip_email_enabled,
        "insight_email_enabled": insight_email_enabled,
        "whatsapp_updates_enabled": whatsapp_updates_enabled,
        "daily_report_enabled": bool(daily_prefs.get("enabled", True)),
        "daily_report_hour": int(daily_prefs.get("hour", 9)),
        "daily_report_minute": int(daily_prefs.get("minute", 0)),
    }


async def _get_security_settings(user_id: int) -> dict:
    auth_user, identities = await asyncio.gather(
        asyncio.to_thread(get_auth_user, user_id),
        asyncio.to_thread(list_identities_by_user, user_id),
    )
    auth_user = auth_user or {}
    identities = identities or []
    whatsapp_identity = next((i for i in identities if i.get("provider") == "whatsapp"), None)
    phone = auth_user.get("phone_e164") or (whatsapp_identity or {}).get("external_id")
    return json.loads(jdump({
        "ok": True,
        "user_id": user_id,
        "email": auth_user.get("email"),
        "display_name": auth_user.get("display_name"),
        "phone": phone,
        "phone_status": auth_user.get("phone_status"),
        "phone_confirmed_at": auth_user.get("phone_confirmed_at"),
        "whatsapp_verified_at": auth_user.get("whatsapp_verified_at"),
        "plan": auth_user.get("plan"),
        "plan_expires_at": auth_user.get("plan_expires_at"),
        "created_at": auth_user.get("created_at"),
        "identities": identities,
    }))


@app.get("/settings/{user_id}/security")
async def security_settings_route(request: Request, user_id: int):
    _authorize_dashboard_access(request, user_id)
    return await _get_security_settings(user_id)


@app.patch("/settings/{user_id}/security/contact")
async def update_security_contact_route(
    request: Request,
    user_id: int,
    payload: SecurityContactPayload,
):
    _authorize_dashboard_access(request, user_id)
    auth_user = await asyncio.to_thread(get_auth_user, user_id)
    if not auth_user:
        raise HTTPException(status_code=400, detail="Esta conta ainda não tem login por e-mail configurado.")

    email = payload.email.strip().lower() if payload.email else None
    phone = (payload.phone or "").strip() or None

    display_name_raw = payload.display_name
    display_name_provided = display_name_raw is not None
    display_name: str | None = None
    if display_name_provided:
        display_name = display_name_raw.strip()
        if display_name == "":
            display_name = None  # remove o nome
        else:
            if len(display_name) > 50:
                raise HTTPException(status_code=400, detail="O nome deve ter no máximo 50 caracteres.")
            if len(display_name) < 2:
                raise HTTPException(status_code=400, detail="O nome deve ter pelo menos 2 caracteres.")

    if not email and not phone and not display_name_provided:
        raise HTTPException(status_code=400, detail="Informe e-mail, telefone ou nome.")
    if email and ("@" not in email or "." not in email.rsplit("@", 1)[-1]):
        raise HTTPException(status_code=400, detail="E-mail inválido.")

    normalized_phone = None
    if phone:
        try:
            normalized_phone = normalize_phone_e164(phone)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    old_email = (auth_user.get("email") or "").strip().lower() or None
    email_actually_changed = bool(email) and email != old_email

    try:
        async with await db_connect() as conn:
            async with conn.cursor() as cur:
                if email:
                    await cur.execute(
                        "UPDATE auth_accounts SET email = %s WHERE user_id = %s",
                        (email, user_id),
                    )
                if normalized_phone:
                    await cur.execute(
                        """
                        UPDATE auth_accounts
                        SET phone_e164 = %s,
                            phone_status = 'pending',
                            phone_confirmed_at = NULL,
                            whatsapp_verified_at = NULL
                        WHERE user_id = %s
                        """,
                        (normalized_phone, user_id),
                    )
                if display_name_provided:
                    await cur.execute(
                        "UPDATE auth_accounts SET display_name = %s WHERE user_id = %s",
                        (display_name, user_id),
                    )
            await conn.commit()
    except psycopg.errors.UniqueViolation as exc:
        raise HTTPException(status_code=409, detail="Este e-mail ou telefone já está em uso.") from exc

    if email_actually_changed:
        await asyncio.to_thread(
            record_audit_event,
            user_id,
            AuditEvent.EMAIL_CHANGED,
            request=request,
            details={"new_email": email},
        )

    return await _get_security_settings(user_id)


@app.post("/settings/{user_id}/password-reset")
@limiter.limit("3/minute")
async def security_password_reset_route(request: Request, user_id: int):
    _authorize_dashboard_access(request, user_id)
    auth_user = await asyncio.to_thread(get_auth_user, user_id)
    email = (auth_user or {}).get("email")
    if not email:
        raise HTTPException(status_code=400, detail="Adicione um e-mail antes de resetar a senha.")

    from db import create_password_reset_token
    from core.services.email_service import send_password_reset_email

    token = await asyncio.to_thread(create_password_reset_token, email)
    if not token:
        raise HTTPException(status_code=404, detail="Conta de e-mail não encontrada.")
    reset_url = f"{DASHBOARD_URL}/reset-password#token={token}"
    sent = await asyncio.to_thread(send_password_reset_email, email.strip().lower(), reset_url)
    if not sent:
        raise HTTPException(status_code=500, detail="Não foi possível enviar o e-mail de reset.")
    return {"ok": True, "message": "Enviamos um link de redefinição de senha para o seu e-mail."}


@app.get("/settings/{user_id}/activity")
async def security_activity_route(
    request: Request,
    user_id: int,
    limit: int = 10,
    before_id: int | None = None,
):
    """Lista os ultimos eventos de auditoria do usuario (Atividade da conta)."""
    _authorize_dashboard_access(request, user_id)
    rows = await asyncio.to_thread(list_audit_events, user_id, limit, before_id)
    next_before = rows[-1]["id"] if rows and len(rows) >= max(1, min(int(limit), 50)) else None
    return json.loads(jdump({"ok": True, "events": rows, "next_before": next_before}))


def _current_session_jti(request: Request) -> str | None:
    """Le o jti da sessao corrente a partir do cookie auth_token. None se ausente/legado."""
    token = _get_auth_token_from_request(request, None)
    if not token:
        return None
    payload = _decode_jwt(token)
    if not payload or payload.get("type") != "auth":
        return None
    return payload.get("jti")


@app.get("/settings/{user_id}/sessions")
async def security_sessions_list_route(request: Request, user_id: int):
    """Lista as sessoes ativas (dispositivos conectados) do usuario."""
    _authorize_dashboard_access(request, user_id)
    current_jti = _current_session_jti(request)
    rows = await asyncio.to_thread(list_user_sessions, user_id)
    sessions = []
    for r in rows:
        sessions.append({
            "jti": r["jti"],
            "device_label": device_label(r.get("user_agent")),
            "ip": r.get("ip"),
            "user_agent": r.get("user_agent"),
            "created_at": r.get("created_at"),
            "last_seen_at": r.get("last_seen_at"),
            "is_current": r["jti"] == current_jti,
        })
    return json.loads(jdump({"ok": True, "sessions": sessions, "current_jti": current_jti}))


@app.delete("/settings/{user_id}/sessions/{jti}")
async def security_session_revoke_route(request: Request, user_id: int, jti: str):
    """Revoga uma sessao especifica (que nao seja a corrente)."""
    _authorize_dashboard_access(request, user_id)
    current_jti = _current_session_jti(request)
    if current_jti and jti == current_jti:
        raise HTTPException(
            status_code=400,
            detail="Use o botão 'Sair' para encerrar a sessão atual.",
        )
    revoked = await asyncio.to_thread(revoke_session, user_id, jti)
    if not revoked:
        raise HTTPException(status_code=404, detail="Sessão não encontrada ou já encerrada.")
    return {"ok": True}


@app.delete("/settings/{user_id}/sessions")
async def security_sessions_revoke_others_route(request: Request, user_id: int):
    """Revoga todas as sessoes do usuario exceto a corrente."""
    _authorize_dashboard_access(request, user_id)
    current_jti = _current_session_jti(request)
    revoked_count = await asyncio.to_thread(revoke_other_sessions, user_id, current_jti)
    return {"ok": True, "revoked": revoked_count}


@app.get("/settings/{user_id}/notifications")
async def notification_settings_route(request: Request, user_id: int):
    _authorize_dashboard_access(request, user_id)
    return await _get_notification_settings(user_id)


@app.patch("/settings/{user_id}/notifications")
async def update_notification_settings_route(
    request: Request,
    user_id: int,
    payload: NotificationSettingsPayload,
):
    _authorize_dashboard_access(request, user_id)

    touches_email_prefs = (
        payload.engagement_email_enabled is not None
        or payload.tip_email_enabled is not None
        or payload.insight_email_enabled is not None
    )
    if touches_email_prefs:
        auth_user = await asyncio.to_thread(get_auth_user, user_id)
        if not auth_user or not auth_user.get("email"):
            raise HTTPException(status_code=400, detail="Vincule um e-mail para configurar notificações por e-mail.")

    if payload.engagement_email_enabled is not None:
        await asyncio.to_thread(set_engagement_opt_out, user_id, not payload.engagement_email_enabled)

    if payload.tip_email_enabled is not None:
        await asyncio.to_thread(set_tip_email_opt_out, user_id, not payload.tip_email_enabled)

    if payload.insight_email_enabled is not None:
        await asyncio.to_thread(set_insight_email_opt_out, user_id, not payload.insight_email_enabled)

    if payload.tip_email_enabled is not None or payload.insight_email_enabled is not None:
        await asyncio.to_thread(sync_engagement_opt_out, user_id)

    if payload.whatsapp_updates_enabled is not None:
        auth_user = await asyncio.to_thread(get_auth_user, user_id)
        if not auth_user or not auth_user.get("phone_e164"):
            raise HTTPException(status_code=400, detail="Vincule um WhatsApp para receber atualizações.")
        await asyncio.to_thread(set_whatsapp_updates_opt_out, user_id, not payload.whatsapp_updates_enabled)

    if payload.daily_report_hour is not None or payload.daily_report_minute is not None:
        current = await asyncio.to_thread(get_daily_report_prefs, user_id)
        hour = payload.daily_report_hour if payload.daily_report_hour is not None else int(current.get("hour", 9))
        minute = payload.daily_report_minute if payload.daily_report_minute is not None else int(current.get("minute", 0))
        if not 0 <= int(hour) <= 23:
            raise HTTPException(status_code=400, detail="Hora inválida.")
        if not 0 <= int(minute) <= 59:
            raise HTTPException(status_code=400, detail="Minuto inválido.")
        await asyncio.to_thread(set_daily_report_hour, user_id, int(hour), int(minute))

    if payload.daily_report_enabled is not None:
        await asyncio.to_thread(set_daily_report_enabled, user_id, payload.daily_report_enabled)

    return await _get_notification_settings(user_id)


@app.get("/open-finance/{user_id}")
async def open_finance_snapshot_route(request: Request, user_id: int):
    _authorize_dashboard_access(request, user_id)
    snapshot = await asyncio.to_thread(get_open_finance_snapshot, user_id)
    return json.loads(jdump({"ok": True, **snapshot}))


@app.post("/open-finance/{user_id}/connect-token")
async def open_finance_connect_token_route(request: Request, user_id: int):
    _authorize_dashboard_access(request, user_id)

    webhook_url = (os.getenv("PLUGGY_WEBHOOK_URL") or "").strip()
    if not webhook_url and DASHBOARD_URL.startswith("https://"):
        webhook_url = f"{DASHBOARD_URL}/open-finance/pluggy/webhook"

    try:
        token_data = await asyncio.to_thread(
            create_pluggy_connect_token,
            user_id,
            webhook_url or None,
        )
    except PluggyConfigError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except PluggyApiError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc

    return {
        "ok": True,
        "accessToken": token_data["accessToken"],
        "includeSandbox": PLUGGY_INCLUDE_SANDBOX,
        "provider": "pluggy",
    }


@app.post("/open-finance/{user_id}/pluggy-item")
async def open_finance_pluggy_item_route(request: Request, user_id: int, payload: OpenFinancePluggyItemPayload):
    _authorize_dashboard_access(request, user_id)
    try:
        connection = await asyncio.to_thread(save_pluggy_open_finance_item, user_id, payload.item)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    await asyncio.to_thread(
        record_audit_event,
        user_id,
        AuditEvent.OPEN_FINANCE_CONNECTED,
        request=request,
        details={"provider": "pluggy", "item_id": (connection or {}).get("item_id")},
    )

    snapshot = await asyncio.to_thread(get_open_finance_snapshot, user_id)
    return json.loads(jdump({"ok": True, "connection": connection, **snapshot}))


def _verify_pluggy_webhook_signature(raw_body: bytes, signature_header: str, secret: str) -> bool:
    signature = (signature_header or "").strip()
    if signature.startswith("sha256="):
        signature = signature.split("=", 1)[1]
    if not signature:
        return False

    expected = _hmac.new(secret.encode("utf-8"), raw_body, _hashlib.sha256).hexdigest()
    return _hmac.compare_digest(signature, expected)


@app.post("/open-finance/pluggy/webhook")
async def open_finance_pluggy_webhook(request: Request):
    """
    Recebe eventos da Pluggy e responde rapido.
    Trabalho pesado de sync deve rodar fora do request.
    """
    secret = (os.getenv("PLUGGY_WEBHOOK_SECRET") or "").strip()
    if not secret:
        raise HTTPException(status_code=503, detail="Webhook não configurado.")

    raw_body = await request.body()
    received_sig = request.headers.get("X-Pluggy-Signature") or ""
    if not _verify_pluggy_webhook_signature(raw_body, received_sig, secret):
        raise HTTPException(status_code=401, detail="Assinatura inválida.")

    try:
        event = json.loads(raw_body)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Webhook inválido.") from exc

    event_name = str(event.get("event") or event.get("type") or "")
    item_id = str(event.get("itemId") or event.get("item_id") or event.get("item", {}).get("id") or "")
    status_by_event = {
        "item/created": "UPDATING",
        "item/updated": "ACTIVE",
        "item/error": "ERROR",
        "item/deleted": "DELETED",
    }
    status = status_by_event.get(event_name)
    if item_id and status:
        await asyncio.to_thread(update_pluggy_open_finance_item_status, item_id, status, event)

    await log_system_event(
        "info" if event_name != "item/error" else "warning",
        "pluggy_webhook_received",
        f"Webhook Pluggy recebido: {event_name or 'evento desconhecido'}",
        source="open_finance",
        details={"event": event_name, "item_id": item_id},
    )
    return {"received": True}


@app.post("/open-finance/{user_id}/mock-connect")
async def open_finance_mock_connect_route(request: Request, user_id: int, payload: OpenFinanceMockConnectPayload):
    _authorize_dashboard_access(request, user_id)
    result = await asyncio.to_thread(
        create_mock_open_finance_connection,
        user_id,
        payload.institution or "nubank",
    )

    await asyncio.to_thread(
        record_audit_event,
        user_id,
        AuditEvent.OPEN_FINANCE_CONNECTED,
        request=request,
        details={"provider": "mock", "institution": payload.institution or "nubank"},
    )

    snapshot = await asyncio.to_thread(get_open_finance_snapshot, user_id)
    return json.loads(jdump({"ok": True, "sync": result, **snapshot}))


@app.delete("/open-finance/{user_id}")
async def open_finance_disconnect_route(request: Request, user_id: int):
    _authorize_dashboard_access(request, user_id)
    deleted = await asyncio.to_thread(disconnect_open_finance_connection, user_id)

    if deleted:
        await asyncio.to_thread(
            record_audit_event,
            user_id,
            AuditEvent.OPEN_FINANCE_DISCONNECTED,
            request=request,
        )

    return {"ok": True, "deleted": deleted}


# ─── WebSocket ────────────────────────────────────────────────────────────────

@app.websocket("/ws/{user_id}")
async def websocket_endpoint(ws: WebSocket, user_id: int):
    token = (ws.cookies.get(DASHBOARD_COOKIE_NAME) or "").strip()
    payload = decode_dashboard_token_full(token)
    if not payload or int(payload["user_id"]) != int(user_id):
        await ws.close(code=1008)
        return
    jti = payload.get("jti")
    if jti:
        session = await asyncio.to_thread(get_active_session, jti)
        if not session or int(session.get("user_id") or 0) != int(user_id):
            await ws.close(code=1008)
            return

    now = datetime.now(timezone.utc)
    if not await manager.connect(ws, user_id, now.year, now.month):
        return
    print(f"Connected: user={user_id} total={len(manager.active.get(user_id, {}))}")
    try:
        # Send initial snapshot with current month
        data = await get_financial_data(user_id, now.year, now.month)
        await ws.send_text(jdump({"type": "snapshot", "data": data}))

        while True:
            raw = await ws.receive_text()
            try:
                payload = json.loads(raw) if raw.strip().startswith("{") else {"type": raw}
                t       = payload.get("type")

                if t == "refresh":
                    y, m = manager.get_month(ws, user_id)
                    page  = int(payload.get("page", 1))
                    limit = int(payload.get("limit", 25))
                    filter_type = str(payload.get("filter_type", "all"))
                    query = str(payload.get("q", ""))
                    data = await get_financial_data(user_id, y, m, page, limit, filter_type, query)
                    await ws.send_text(jdump({"type": "update", "data": data}))

                elif t == "get_month":
                    # Data for a specific month (month selector navigation)
                    now = datetime.now(timezone.utc)
                    y   = int(payload.get("year", now.year))
                    m   = int(payload.get("month", now.month))
                    page  = int(payload.get("page", 1))
                    limit = int(payload.get("limit", 25))
                    filter_type = str(payload.get("filter_type", "all"))
                    query = str(payload.get("q", ""))

                    manager.set_month(ws, user_id, y, m)

                    data = await get_financial_data(user_id, y, m, page, limit, filter_type, query)
                    await ws.send_text(jdump({"type": "month_data", "data": data}))

                elif t == "get_history":
                    n       = min(max(int(payload.get("months", 6)), 1), 24)
                    history = await get_monthly_history(user_id, n)
                    await ws.send_text(jdump({"type": "history_data", "data": history}))

                elif t == "ping":
                    await ws.send_text(jdump({"type": "pong"}))

            except Exception as exc:
                print(f"[ws] error handling message type={payload.get('type') if 'payload' in dir() else '?'}: {exc}")

    except WebSocketDisconnect:
        pass
    finally:
        manager.disconnect(ws, user_id)
        print(f"Disconnected: user={user_id}")

# ─── Entry point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    uvicorn.run(
        "finance_bot_websocket_custom:app",
        host="0.0.0.0",
        port=8000,
        reload=False,
        log_level="info",
    )
